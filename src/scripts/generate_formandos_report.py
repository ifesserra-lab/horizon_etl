"""
Gera o relatório HTML 'Formandos × Pesquisa' para IFES Serra.

Fonte:
  - data/formandos/formados_2024_1.xlsx  (lista de formandos)
  - data/exports/advisorships_canonical.json
  - data/exports/research_groups_canonical.json
  - data/exports/researchers_canonical.json

Metodologia de matching:
  Formandos são cruzados com advisorships_canonical por nome (person_name).
  Apenas pessoas que aparecem como person_id em orientações são consideradas
  "registradas no SigPesq como alunos de pesquisa".

Uso:
  python -m src.scripts.generate_formandos_report
  python -m src.scripts.generate_formandos_report --semester 2024_2
"""

from __future__ import annotations

import argparse
import json
import unicodedata
from collections import Counter, defaultdict
from datetime import datetime
from pathlib import Path
from typing import Any

import re


def _match_key(name: str | None) -> str:
    """
    Accent-insensitive, case-insensitive key for name matching across sources.

    SigPesq / Lattes / planilhas grafam nomes com acentuação e espaçamento
    inconsistentes; comparar por ``lower().strip()`` perde casamentos legítimos
    (ex.: 'João' vs 'Joao'). Esta chave remove acentos e normaliza espaços.
    """
    s = (name or "").lower().strip()
    s = "".join(c for c in unicodedata.normalize("NFD", s)
                if unicodedata.category(c) != "Mn")
    return re.sub(r"\s+", " ", s)


def _token_set(name: str | None) -> set[str]:
    """{primeiro, último} nome normalizados — para casamento aproximado."""
    parts = _match_key(name).split()
    return {parts[0], parts[-1]} if len(parts) >= 2 else set()

import openpyxl

# ---------------------------------------------------------------------------
# String normalization
# ---------------------------------------------------------------------------

_LOWER_WORDS = frozenset({
    "de", "da", "do", "das", "dos", "e", "a", "o", "em", "com",
    "para", "por", "ou", "um", "uma", "no", "na", "nos", "nas",
})
# Acronyms: all ASCII uppercase/digits, max 6 chars (PIBIC, CAPES, IFES, IA…)
_ACRONYM_RE = re.compile(r"^[A-Z0-9]{2,6}$")
# Preserve as-is: explicitly known mixed-case terms
_PRESERVE = frozenset({"CNPq", "SigPesq", "LaTeX", "OpenCV", "OpenAI"})


def normalize_str(text: str) -> str:
    """
    Normalize a knowledge-area / label string to Title Case.

    Rules (priority order):
      1. Word in _PRESERVE whitelist → keep as-is.
      2. Word (lowered) is PT preposition/article and not first word → lowercase.
      3. Word is pure ASCII uppercase, 2-6 chars → treat as acronym, keep as-is.
      4. Otherwise → capitalize first char, lowercase rest.
    """
    if not text or not text.strip():
        return text
    words = text.strip().split()
    result = []
    for i, word in enumerate(words):
        core = word.strip(".,;:()")
        if core in _PRESERVE:
            result.append(word)
        elif core.lower() in _LOWER_WORDS and i != 0:
            result.append(core.lower())
        elif _ACRONYM_RE.match(core):
            result.append(word)
        else:
            result.append(word[0].upper() + word[1:].lower() if word else word)
    return " ".join(result)


def normalize_name(text: str) -> str:
    """
    Normalize a person name to Title Case (no acronym preservation).
    Handles ALL_CAPS and all-lower inputs.
    """
    if not text or not text.strip():
        return text
    words = text.strip().split()
    result = []
    for i, word in enumerate(words):
        core = word.strip(".,;:()")
        if core.lower() in _LOWER_WORDS and i != 0:
            result.append(core.lower())
        else:
            result.append(word[0].upper() + word[1:].lower() if word else word)
    return " ".join(result)

# ---------------------------------------------------------------------------
# Paths
# ---------------------------------------------------------------------------
BASE = Path(__file__).resolve().parents[2]  # horizon_etl/
DATA_FORMANDOS = BASE / "data" / "formandos"
DATA_EXPORTS = BASE / "data" / "exports"
OUT_DIR = DATA_EXPORTS / "formandos"
BOLSISTAS_FILE = DATA_EXPORTS / "bolsistas" / "ifes-campus-serra-bolsistas.json"

# Fontes integradas (cruzamento formandos × projetos × mestrado).
FAPES_PROJ_FILE = (
    DATA_EXPORTS / "projetos-fapes"
    / "ifes-campus-serra-projetos-concluidos-em-andamento.json"
)
FACTO_PROJ_FILE = DATA_EXPORTS / "projetos-facto" / "facto_projects_full.json"
MESTRADO_BASE_FILE = DATA_EXPORTS / "mestrado" / "ppcomp_base_analitico.json"

SEMESTER_FILE_MAP = {
    "2020_1": "formados_2020_1.xlsx",
    "2020_2": "formados_2020_2.xlsx",
    "2021_1": "formados_2021_1.xlsx",
    "2021_2": "formados_2021_2.xlsx",
    "2022_1": "formados_2022_1.xlsx",
    "2022_2": "formados_2022_2.xlsx",
    "2023_1": "formados_2023_1.xlsx",
    "2023_2": "formados_2023_2.xlsx",
    "2024_1": "formados_2024_1.xlsx",
    "2024_2": "formados_2024_2.xlsx",
    "2025_1": "formados_2025_1.xlsx",
    "2025_2": "formados_2025_2.xlsx",
}

# Admission-form column appears only in some files (2020–2023). We build a
# global matrícula→forma-de-ingresso map from every file that has it, so
# students from files lacking the column still get their admission category.
_ADMISSION_COL = "Desc_Forma_Ingresso_Matricula"

# Duração prevista do currículo, em semestres. Os cursos têm tamanhos distintos,
# então comparar semestres absolutos entre eles é enviesado — usamos o ATRASO
# (semestres além do previsto) como métrica comparável entre cursos.
CURSO_EXPECTED_SEMESTERS = {
    "Sistemas de Informação": 8,               # 4 anos
    "Engenharia de Controle e Automação": 12,  # 6 anos (curso noturno)
}
_DEFAULT_EXPECTED = 8


def expected_semesters(curso: str | None) -> int:
    """Semestres previstos no currículo do curso (SI=8, ECA=12)."""
    if curso in CURSO_EXPECTED_SEMESTERS:
        return CURSO_EXPECTED_SEMESTERS[curso]
    return 12 if "Controle" in (curso or "") else _DEFAULT_EXPECTED

# ---------------------------------------------------------------------------
# Data loading
# ---------------------------------------------------------------------------

def _parse_matricula(mat: str) -> dict | None:
    """Extract entry year and semester from matricula like '20181BSI0056'."""
    if not mat or len(mat) < 5:
        return None
    try:
        year = int(str(mat)[:4])
        sem = int(str(mat)[4])
        if sem not in (1, 2):
            return None
        return {"year": year, "semester": sem}
    except (ValueError, IndexError):
        return None


def _header_index(ws) -> dict[str, int]:
    """Map column header → 0-based index from row 1 (headers vary per file)."""
    header = next(ws.iter_rows(min_row=1, max_row=1, values_only=True))
    return {str(h).strip(): i for i, h in enumerate(header) if h is not None}


_admission_map_cache: dict[str, str] | None = None


def load_admission_map() -> dict[str, str]:
    """
    Build {matrícula → forma de ingresso} across ALL semester files.

    The admission column (Desc_Forma_Ingresso_Matricula) only exists in the
    2020–2023 exports. Since the join key is the matrícula (per user), we scan
    every file once and key on matrícula, so students whose graduation-semester
    file lacks the column still resolve their admission category.
    """
    global _admission_map_cache
    if _admission_map_cache is not None:
        return _admission_map_cache
    amap: dict[str, str] = {}
    for fname in SEMESTER_FILE_MAP.values():
        path = DATA_FORMANDOS / fname
        if not path.exists():
            continue
        wb = openpyxl.load_workbook(path, read_only=True)
        ws = wb.active
        cols = _header_index(ws)
        if _ADMISSION_COL not in cols or "Matrícula" not in cols:
            wb.close()
            continue
        mi, ai = cols["Matrícula"], cols[_ADMISSION_COL]
        for r in ws.iter_rows(values_only=True, min_row=2):
            mat = str(r[mi]).strip() if r[mi] is not None else ""
            adm = r[ai]
            if mat and adm and mat not in amap:
                amap[mat] = str(adm).strip()
        wb.close()
    _admission_map_cache = amap
    return amap


def load_formandos(semester: str) -> list[dict]:
    """Load one semester file, keyed by header names (column order varies)."""
    path = DATA_FORMANDOS / SEMESTER_FILE_MAP[semester]
    wb = openpyxl.load_workbook(path, read_only=True)
    ws = wb.active
    cols = _header_index(ws)
    ni = cols.get("Nome", 1)
    ci = cols.get("Curso", 3)
    mi = cols.get("Matrícula", 0)
    ai = cols.get(_ADMISSION_COL)  # may be absent
    amap = load_admission_map()
    out: list[dict] = []
    for r in ws.iter_rows(values_only=True, min_row=2):
        if not r[ni]:
            continue
        mat = str(r[mi]).strip() if r[mi] is not None else ""
        # admission: prefer in-file column, fall back to global matrícula map
        adm = (r[ai] if ai is not None and r[ai] else None) or amap.get(mat)
        out.append({
            "nome": r[ni], "curso": r[ci], "matricula": mat,
            "entry": _parse_matricula(mat),
            "admissao": str(adm).strip() if adm else None,
            "grad_semester": semester,
        })
    wb.close()
    return out


# ---------------------------------------------------------------------------
# Admission-form (cotas) classification
# ---------------------------------------------------------------------------

# Top-level group: a student is in exactly one.
ADM_GROUP_AMPLA = "Ampla Concorrência"
ADM_GROUP_COTA = "Cotas / Reserva de vagas"
ADM_GROUP_TRANSF = "Transferência"
ADM_GROUP_NA = "Sem informação"


def admission_group(desc: str | None) -> str:
    """Classify a forma-de-ingresso string into one top-level group."""
    if not desc:
        return ADM_GROUP_NA
    d = desc.lower()
    if "transfer" in d:
        return ADM_GROUP_TRANSF
    if any(k in d for k in ("ppi", "renda", "esc. públ", "esc. publ", "ação afirmativa")):
        return ADM_GROUP_COTA
    return ADM_GROUP_AMPLA


# Raw SIG label → human-readable PT. PPI = Pretos, Pardos e Indígenas;
# CD = Pessoa com Deficiência; OE = Outras Escolas; renda = ≤1,5 salário mínimo
# per capita; M* = modalidades ENEM da Lei de Cotas (12.711); PS = processo seletivo.
_ADM_LABELS: dict[str, str] = {
    "M9 - Enem - Ampla Concorrência": "Ampla Concorrência (ENEM)",
    "M8 - Enem - Esc. Públ.": "Escola Pública (ENEM)",
    "M6 - Enem - Esc. Públ. - PPI": "Escola Pública · PPI (ENEM)",
    "M7 - Enem - Esc. Públ. - CD": "Escola Pública · Pessoa c/ Deficiência (ENEM)",
    "M4 - Enem - Renda - Esc. Públ.": "Escola Pública · Renda ≤1,5 SM (ENEM)",
    "M2 - Enem - Renda - Esc. Públ - PPI": "Escola Pública · Renda ≤1,5 SM · PPI (ENEM)",
    "Ampla Concorrência": "Ampla Concorrência",
    "Transferência Externa": "Transferência Externa",
    "Transferência Interna": "Transferência Interna",
    "PS - Ação Afirmativa 1 - OE - CD": "Ação Afirmativa 1 · Outras Escolas · Pessoa c/ Deficiência",
    "PS - Ação Afirmativa 2 - OE": "Ação Afirmativa 2 · Outras Escolas",
    "PS - Ação Afirmativa 2 - PPI": "Ação Afirmativa 2 · PPI",
}


def admission_label(desc: str | None) -> str:
    """Human-readable PT label for a raw forma-de-ingresso string."""
    if not desc:
        return "Sem informação"
    return _ADM_LABELS.get(desc.strip(), desc.strip())


def admission_flags(desc: str | None) -> set[str]:
    """Quota attributes a student qualifies under (can be multiple)."""
    if not desc:
        return set()
    d = desc.lower()
    flags: set[str] = set()
    if "ppi" in d:
        flags.add("PPI")
    if "renda" in d:
        flags.add("Renda")
    if "esc. públ" in d or "esc. publ" in d:
        flags.add("Escola Pública")
    if d.endswith("- cd") or d.endswith(" cd") or "- cd" in d:
        flags.add("Pessoa c/ Deficiência")
    if "ação afirmativa" in d:
        flags.add("Ação Afirmativa")
    return flags


def load_json(name: str) -> Any:
    return json.loads((DATA_EXPORTS / name).read_text())


def load_bolsistas() -> dict:
    """Dados de bolsistas FAPES do IFES Serra: {bolsistas_unicos, alocacoes}."""
    if not BOLSISTAS_FILE.exists():
        return {"bolsistas_unicos": [], "alocacoes": []}
    data = json.loads(BOLSISTAS_FILE.read_text())
    return {
        "bolsistas_unicos": data.get("bolsistas_unicos", []),
        "alocacoes": data.get("alocacoes", []),
    }


def load_lattes() -> dict[str, list[dict]]:
    """
    Load all IC and TCC records from Lattes JSON files.
    Returns {'ic': [...], 'tcc': [...]}, each item has keys:
      orientando, titulo, ano_inicio, ano_conclusao, supervisor, status
    """
    lattes_dir = BASE / "data" / "lattes_json"
    ic: list[dict] = []
    tcc: list[dict] = []
    for f in sorted(lattes_dir.glob("*.json")):
        try:
            data = json.loads(f.read_text())
        except Exception:
            continue
        _ip = data.get("informacoes_pessoais") or {}
        nome_sup = normalize_name(
            _ip.get("nome_completo") or _ip.get("nome") or f.stem
        )
        for status in ("em_andamento", "concluidas"):
            d = (data.get("orientacoes") or {}).get(status, {}) or {}
            for item in d.get("iniciacao_cientifica", []):
                ic.append({
                    "orientando": (item.get("orientando") or "").strip(),
                    "titulo": item.get("titulo", ""),
                    "ano_inicio": item.get("ano_inicio"),
                    "ano_conclusao": item.get("ano_conclusao"),
                    "supervisor": nome_sup,
                    "status": status,
                })
            for item in d.get("tcc", []):
                tcc.append({
                    "orientando": (item.get("orientando") or "").strip(),
                    "titulo": item.get("titulo", ""),
                    "ano_inicio": item.get("ano_inicio"),
                    "ano_conclusao": item.get("ano_conclusao"),
                    "supervisor": nome_sup,
                    "status": status,
                })
    return {"ic": ic, "tcc": tcc}


# ---------------------------------------------------------------------------
# Fontes integradas: projetos FAPES, projetos FACTO, mestrado PPComp
# ---------------------------------------------------------------------------

def _xkey(name: str | None) -> str:
    """Chave de cruzamento entre fontes: ascii minúsculo, espaços colapsados.

    Junta nomes grafados de forma divergente entre SigPesq, FAPES e FACTO
    (acentos, caixa, espaços duplos).
    """
    if not name:
        return ""
    s = unicodedata.normalize("NFKD", name).encode("ascii", "ignore").decode()
    return re.sub(r"\s+", " ", s).strip().lower()


# Coordenadores excluídos da análise de projetos de pesquisa (recorte da gestão).
# Casado por substring sobre a _xkey do coordenador, para tolerar variações de grafia.
_EXCLUDED_COORD_SUBSTR = ("elton",)


def _is_excluded_coord(name: str | None) -> bool:
    k = _xkey(name)
    return any(sub in k for sub in _EXCLUDED_COORD_SUBSTR)


def _br_money(value: Any) -> float:
    """Converte '1.256.355,65' (pt-BR) ou número em float. 0.0 se vazio/inválido."""
    if value is None:
        return 0.0
    if isinstance(value, (int, float)):
        return float(value)
    try:
        return float(str(value).replace(".", "").replace(",", "."))
    except ValueError:
        return 0.0


def load_fapes_projects() -> dict:
    """Projetos FAPES do IFES Serra por status/prazo. {} se ausente."""
    if not FAPES_PROJ_FILE.exists():
        return {}
    return json.loads(FAPES_PROJ_FILE.read_text())


def load_facto_projects() -> dict:
    """Projetos da fundação FACTO (IFES institucional). {} se ausente."""
    if not FACTO_PROJ_FILE.exists():
        return {}
    return json.loads(FACTO_PROJ_FILE.read_text())


def load_mestrado_base() -> dict:
    """Base analítica do mestrado PPComp. {} se ausente."""
    if not MESTRADO_BASE_FILE.exists():
        return {}
    return json.loads(MESTRADO_BASE_FILE.read_text())


def compute_integrated(sup_formando_counts: dict[str, int],
                       ic_grad_invest: float,
                       bolsistas_alocado: float) -> dict:
    """Cruza formandos × projetos FAPES/FACTO × mestrado PPComp.

    Produz três blocos:
      - fapes_ecosystem: orientadores de formandos que coordenam projetos FAPES,
        com formandos mentorados × portfólio (orçamento, bolsas).
      - mestrado_pipeline: funil graduação → mestrado PPComp + bolsas de pós.
      - fomento_panel: escala de fomento — bolsa de IC (semente) vs projetos
        FAPES (R$48M) vs fundação FACTO.

    Cruzamento de orientadores por nome normalizado (_xkey). Junção de ALUNOS
    não é feita aqui — exige ID comum (Lattes/SigPesq), ver nota nas seções.
    """
    fapes = load_fapes_projects()
    facto = load_facto_projects()
    mestrado = load_mestrado_base()

    integrated: dict = {}

    # ---- A. Ecossistema do orientador (formandos × portfólio FAPES) ----
    if fapes:
        coord_portfolio: dict[str, dict] = {}
        for _cat, projetos in (fapes.get("projetos") or {}).items():
            for p in projetos:
                ck = _xkey(p.get("coordenador_nome"))
                if not ck or _is_excluded_coord(p.get("coordenador_nome")):
                    continue
                d = coord_portfolio.setdefault(ck, {
                    "nome": p.get("coordenador_nome"),
                    "n_proj": 0, "orcamento": 0.0, "bolsa_val": 0.0, "bolsa_q": 0,
                })
                d["n_proj"] += 1
                d["orcamento"] += _br_money(p.get("orcamento_contratado"))
                d["bolsa_val"] += _br_money(p.get("valor_bolsas"))
                d["bolsa_q"] += p.get("quantidade_bolsas") or 0

        # chave de cruzamento dos orientadores de formandos
        sup_xkey = {_xkey(name): (name, n) for name, n in sup_formando_counts.items()}

        rows = []
        for ck, port in coord_portfolio.items():
            if ck not in sup_xkey:
                continue  # coordenador FAPES que não orientou formando casado
            disp_name, n_form = sup_xkey[ck]
            rows.append({
                "nome": disp_name,
                "formandos": n_form,
                "n_proj": port["n_proj"],
                "orcamento": round(port["orcamento"], 2),
                "bolsa_val": round(port["bolsa_val"], 2),
                "bolsa_q": port["bolsa_q"],
            })
        rows.sort(key=lambda r: -r["orcamento"])

        resumo = fapes.get("resumo", {}).get("total_status_ou_prazo", {})
        campus_orc = resumo.get("orcamento_contratado_total", 0) or 0
        sum_orc = sum(r["orcamento"] for r in rows)
        integrated["fapes_ecosystem"] = {
            "rows": rows,
            "n_orientadores": len(rows),
            "n_coord_total": len(coord_portfolio),
            "total_orc": round(sum_orc, 2),
            "total_bolsa": round(sum(r["bolsa_val"] for r in rows), 2),
            "total_bolsa_q": sum(r["bolsa_q"] for r in rows),
            "campus_orc": campus_orc,
            "campus_bolsa": resumo.get("valor_bolsas_total", 0) or 0,
            "campus_proj": resumo.get("quantidade_projetos", 0) or 0,
            "pct_orc": round(sum_orc / campus_orc * 100, 1) if campus_orc else 0,
        }

    # ---- B. Pipeline graduação → mestrado PPComp ----
    if mestrado:
        pipe = mestrado.get("pipeline", {})
        bolsas = mestrado.get("bolsas", {})
        integrated["mestrado_pipeline"] = {
            "total": mestrado.get("total", 0),
            "defendidos": mestrado.get("defendidos", 0),
            "ativos": mestrado.get("ativos", 0),
            "evasao": mestrado.get("evasao", 0),
            "taxa_defesa": mestrado.get("taxa_defesa_desfecho", 0),
            "tempo_medio": mestrado.get("tempo_medio_anos", 0),
            "pipe_n": pipe.get("total", 0),
            "pipe_pct": pipe.get("pct", 0),
            "pipe_def": pipe.get("defendidos", 0),
            "pipe_ativos": pipe.get("ativos", 0),
            "pipe_situacao": pipe.get("por_situacao", {}),
            "bolsas_fapes": bolsas.get("fapes", 0),
            "bolsas_facto": bolsas.get("facto", 0),
            "bolsas_uniao": bolsas.get("uniao", 0),
            "bolsas_valor_fapes": bolsas.get("valor_fapes", 0),
            "bolsas_por_situacao": bolsas.get("por_situacao", {}),
        }

    # ---- C. Painel de fomento (escala: semente IC → projetos) ----
    panel: dict = {
        "ic_grad_invest": round(ic_grad_invest, 2),
        "bolsistas_alocado": round(bolsistas_alocado, 2),
    }
    if fapes:
        r = fapes.get("resumo", {})
        panel["fapes_total"] = r.get("total_status_ou_prazo", {})
        panel["fapes_concluidos"] = r.get("concluidos", {})
        panel["fapes_andamento"] = r.get("em_andamento_por_status_e_prazo", {})
        panel["fapes_prazo_encerrado"] = r.get("status_em_andamento_prazo_encerrado", {})
    if facto:
        # Recorte IFES Serra: a base FACTO é institucional (Reitoria + todos os
        # campi). Mantém apenas projetos cuja instituição executora menciona Serra.
        tot = 0.0
        n_ok = 0
        for p in facto.get("projects", []):
            if p.get("_error"):
                continue
            is_serra = False
            is_excluded = False
            p_tot = 0.0
            for fn, linhas in (p.get("csv") or {}).items():
                if "Informa" in fn:
                    for linha in linhas:
                        if "serra" in _xkey(linha.get("Instituição executora")):
                            is_serra = True
                        if _is_excluded_coord(linha.get("Coordenador")):
                            is_excluded = True
                        p_tot += _br_money(linha.get("Valor aprovado"))
            if is_serra and not is_excluded:
                n_ok += 1
                tot += p_tot
        panel["facto_aprovado"] = round(tot, 2)
        panel["facto_proj"] = n_ok
        panel["facto_portal"] = facto.get("_portal", "")
    if mestrado:
        panel["mestrado_fapes_val"] = mestrado.get("bolsas", {}).get("valor_fapes", 0)
    integrated["fomento_panel"] = panel

    return integrated


# ---------------------------------------------------------------------------
# Stats computation
# ---------------------------------------------------------------------------

def compute(formandos: list[dict], adv_projects: list[dict],
            rgs: list[dict], lattes: dict[str, list[dict]] | None = None,
            grad_semester: str = "", bolsistas: dict | None = None) -> dict:
    names_map: dict[str, str] = {
        f["nome"].lower().strip(): f["curso"] for f in formandos
    }
    entry_map: dict[str, dict] = {
        f["nome"].lower().strip(): f["entry"]
        for f in formandos
        if f.get("entry")
    }
    total = len(formandos)

    # accent-insensitive lookup: match key → canonical name_lower (do formando)
    mk_to_name: dict[str, str] = {
        _match_key(f["nome"]): f["nome"].lower().strip() for f in formandos
    }

    # ---- student matching: only person_ids that appear as advisee ----
    # Casamento por chave sem acento — recupera grafias divergentes SigPesq×planilha.
    name_to_pid: dict[str, int] = {}
    for proj in adv_projects:
        for adv in proj.get("advisorships", []):
            pid = adv.get("person_id")
            canonical = mk_to_name.get(_match_key(adv.get("person_name")))
            if canonical and pid:
                name_to_pid[canonical] = pid

    matched_pids: set[int] = set(name_to_pid.values())
    matched_names: set[str] = set(name_to_pid.keys())

    sem_registro = total - len(matched_pids)

    # ---- bolsistas FAPES que se formaram → também contam como pesquisa ----
    # Bolsa de pesquisa é evidência de participação, mesmo sem registro no SigPesq.
    _bolsistas_unicos = (bolsistas or {}).get("bolsistas_unicos", [])
    bolsista_formado_names: set[str] = set()
    for _b in _bolsistas_unicos:
        _c = mk_to_name.get(_match_key(_b.get("bolsista_pesquisador_nome")))
        if _c:
            bolsista_formado_names.add(_c)

    # conjunto unificado "participou de pesquisa" = SigPesq IC ∪ bolsistas FAPES
    research_names: set[str] = matched_names | bolsista_formado_names

    # ---- advisorship records for matched formandos ----
    person_projects: dict[int, set] = defaultdict(set)
    person_supervisors: dict[int, set] = defaultdict(set)
    person_fellowship_records: dict[int, list] = defaultdict(list)
    sponsor_persons: dict[str, set] = defaultdict(set)
    fellowship_persons: dict[str, set] = defaultdict(set)
    sup_unique: dict[str, set] = defaultdict(set)
    person_first_ic: dict[int, datetime] = {}  # earliest IC start per person

    pid_to_name: dict[int, str] = {v: k for k, v in name_to_pid.items()}

    for proj in adv_projects:
        for adv in proj.get("advisorships", []):
            pid = adv.get("person_id")
            if pid not in matched_pids:
                continue
            person_projects[pid].add(proj["id"])
            sname = normalize_name(adv.get("supervisor_name") or "")
            if sname:
                person_supervisors[pid].add(sname)
                sup_unique[sname].add(pid)

            # track earliest IC start
            try:
                ic_start = datetime.fromisoformat((adv.get("start_date") or "")[:10])
                if pid not in person_first_ic or ic_start < person_first_ic[pid]:
                    person_first_ic[pid] = ic_start
            except Exception:
                pass

            fel = adv.get("fellowship") or {}
            if fel and fel.get("name"):
                sponsor = normalize_str(fel.get("sponsor_name") or "Voluntário")
                fname = normalize_str(fel["name"])
                year = (adv.get("start_date") or "")[:4]
                days: int | None = None
                try:
                    s = datetime.fromisoformat((adv["start_date"] or "")[:10])
                    e = datetime.fromisoformat((adv["end_date"] or "")[:10])
                    days = (e - s).days
                except Exception:
                    pass
                person_fellowship_records[pid].append({
                    "year": year,
                    "sponsor": sponsor,
                    "fellowship": fname,
                    "days": days,
                    "value": fel.get("value") or 0.0,
                })
                sponsor_persons[sponsor].add(pid)
                fellowship_persons[fname].add(pid)

    # ---- research groups ----
    # names with a PAID fellowship (sponsor ≠ Voluntário), via fellowship records
    paid_names: set[str] = set()
    for _pidP, _recsP in person_fellowship_records.items():
        if any(r["sponsor"] != "Voluntário" for r in _recsP):
            paid_names.add(pid_to_name.get(_pidP, ""))

    # name → admission group (cota vs não-cota), from deduped formandos
    name_to_admgroup: dict[str, str] = {
        f["nome"].lower().strip(): admission_group(f.get("admissao"))
        for f in formandos
    }

    rg_names: set[str] = set()
    rg_top: Counter = Counter()
    rg_stats: dict[str, dict[str, int]] = defaultdict(
        lambda: {"total": 0, "paid": 0, "vol": 0, "cota": 0, "naocota": 0}
    )
    for rg in rgs:
        members_in = [
            m for m in (rg.get("members") or [])
            if _match_key(m.get("name")) in mk_to_name
        ]
        if members_in:
            gname = normalize_str(rg.get("name", ""))
            rg_top[gname] += len(members_in)
            for m in members_in:
                nk = mk_to_name[_match_key(m.get("name"))]
                rg_names.add(nk)
                rg_stats[gname]["total"] += 1
                if nk in paid_names:
                    rg_stats[gname]["paid"] += 1
                else:
                    rg_stats[gname]["vol"] += 1
                if name_to_admgroup.get(nk) == "Cotas / Reserva de vagas":
                    rg_stats[gname]["cota"] += 1
                else:
                    rg_stats[gname]["naocota"] += 1

    # add RG-only formandos to with_research
    rg_only_pids: set[int] = set()
    for n in rg_names - matched_names:
        # not already found via advisorships; we can't get pid easily
        pass

    # "com pesquisa" inclui SigPesq IC + bolsistas FAPES formados
    with_research = len(research_names)
    with_research_sigpesq = len(matched_pids)  # só via advisorship (SigPesq)
    sem_pesquisa = total - len(research_names)

    # ---- fellowship counts ----
    fellowship_counts = {k: len(v) for k, v in fellowship_persons.items()}
    sponsor_counts = {k: len(v) for k, v in sponsor_persons.items()}

    # ---- investment per sponsor ----
    sponsor_investment: dict[str, float] = defaultdict(float)
    for recs in person_fellowship_records.values():
        for r in recs:
            sponsor_investment[r["sponsor"]] += r["value"]

    # ---- sponsor × fellowship (unique persons) ----
    sponsor_fellowship_unique: dict[str, dict[str, int]] = defaultdict(lambda: defaultdict(int))
    sf_persons: dict[tuple, set] = defaultdict(set)
    for pid, recs in person_fellowship_records.items():
        for r in recs:
            sf_persons[(r["sponsor"], r["fellowship"])].add(pid)
    for (s, f), pids in sf_persons.items():
        sponsor_fellowship_unique[s][f] = len(pids)

    # ---- curso × sponsor ----
    curso_sponsor: dict[str, dict[str, int]] = defaultdict(lambda: defaultdict(int))
    curso_sponsor_sets: dict[str, dict[str, set]] = defaultdict(lambda: defaultdict(set))
    for pid, recs in person_fellowship_records.items():
        # find curso by name
        pid_name = next(
            (n for n, p in name_to_pid.items() if p == pid), None
        )
        if pid_name is None:
            continue
        curso = names_map.get(pid_name, "N/A")
        for r in recs:
            curso_sponsor_sets[curso][r["sponsor"]].add(pid)
    for curso, sps in curso_sponsor_sets.items():
        for s, pids in sps.items():
            curso_sponsor[curso][s] = len(pids)

    # ---- project distribution ----
    proj_dist = Counter(len(v) for v in person_projects.values())
    with_projects = len(person_projects)

    # ---- duration ----
    durations = [
        r["days"]
        for recs in person_fellowship_records.values()
        for r in recs
        if r["days"] and 30 < r["days"] < 800
    ]
    dur_mean = round(sum(durations) / len(durations)) if durations else 0
    dur_median = sorted(durations)[len(durations) // 2] if durations else 0
    dur_bins = Counter()
    for d in durations:
        if d <= 180:
            dur_bins["≤6m"] += 1
        elif d <= 365:
            dur_bins["7-12m"] += 1
        else:
            dur_bins[">12m"] += 1

    # ---- progressão ----
    prog: Counter = Counter()
    for pid, recs in person_fellowship_records.items():
        if len(recs) <= 1:
            continue
        sorted_recs = sorted(recs, key=lambda x: x["year"])
        sponsors = [r["sponsor"] for r in sorted_recs]
        paid = [s for s in sponsors if s != "Voluntário"]
        vol = [s for s in sponsors if s == "Voluntário"]
        if vol and paid:
            if sponsors[0] == "Voluntário":
                prog["vol→pago"] += 1
            elif sponsors[-1] == "Voluntário":
                prog["pago→vol"] += 1
            else:
                prog["misto"] += 1
        elif len(set(paid)) > 1:
            prog["multi-agencia"] += 1
    multi_bolsa = sum(
        1 for recs in person_fellowship_records.values() if len(recs) > 1
    )

    # ---- orientadores distribution ----
    sup_dist = Counter(len(v) for v in person_supervisors.values())
    top_sups = sorted(sup_unique.items(), key=lambda x: -len(x[1]))[:10]

    # ---- curso breakdown (with/without research) ----
    curso_with = Counter()
    curso_total = Counter(f["curso"] for f in formandos)
    for n in research_names:
        curso_with[names_map[n]] += 1

    # ---- RG top (with paid/volunteer split) ----
    rg_top_list = rg_top.most_common(8)
    rg_top_stats = [
        {"name": name, "total": rg_stats[name]["total"],
         "paid": rg_stats[name]["paid"], "vol": rg_stats[name]["vol"],
         "cota": rg_stats[name]["cota"], "naocota": rg_stats[name]["naocota"]}
        for name, _ in rg_top_list
    ]

    # ---- KA from RGs (deduplicated by normalized key) ----
    _KA_TRANSLATE: dict[str, str] = {
        "machine learning": "Aprendizado de Máquina",
        "deep learning": "Aprendizado Profundo",
        "computer vision": "Visão Computacional",
        "natural language processing": "Processamento de Linguagem Natural",
        "internet of things": "Internet das Coisas",
        "iot": "Internet das Coisas",
        "data science": "Ciência de Dados",
        "big data": "Grandes Volumes de Dados",
        "embedded systems": "Sistemas Embarcados",
        "control systems": "Sistemas de Controle",
        "automation": "Automação",
        "robotics": "Robótica",
        "software engineering": "Engenharia de Software",
    }
    _ka_raw: dict[str, int] = defaultdict(int)
    for rg in rgs:
        members_in = [
            m for m in (rg.get("members") or [])
            if _match_key(m.get("name")) in mk_to_name
        ]
        if not members_in:
            continue
        for ka in (rg.get("knowledge_areas") or []):
            raw_name = ka.get("name", "")
            translated = _KA_TRANSLATE.get(raw_name.lower().strip(), raw_name)
            key = normalize_str(translated)
            _ka_raw[key] += len(members_in)
    # merge near-duplicates: same normalized lower key
    _ka_merged: dict[str, int] = defaultdict(int)
    for k, v in _ka_raw.items():
        _ka_merged[k.lower()] = max(_ka_merged[k.lower()], v)
    # rebuild with proper casing (normalize again)
    ka_counter: Counter = Counter()
    for k, v in _ka_merged.items():
        ka_counter[normalize_str(k)] += v

    # ---- IC timing: when did students start IC relative to enrollment ----
    def _date_to_sem(year: int, month: int) -> tuple[int, int]:
        """Convert year/month to (year, semester): Jan-Jun → 1, Jul-Dec → 2."""
        return year, (1 if month <= 6 else 2)

    def _sem_diff(ey: int, es: int, iy: int, i_s: int) -> int:
        """Semesters from entry (ey/es) to IC start (iy/i_s)."""
        return (iy - ey) * 2 + (i_s - es)

    ic_period_dist: Counter = Counter()   # semesters after entry → count
    ic_year_dist: Counter = Counter()     # calendar year of first IC → count
    ic_timing_records: list[dict] = []    # per-person detail

    for pid, ic_date in person_first_ic.items():
        name = pid_to_name.get(pid, "")
        entry = entry_map.get(name)
        ic_year_dist[ic_date.year] += 1
        if entry:
            diff = _sem_diff(entry["year"], entry["semester"],
                             *_date_to_sem(ic_date.year, ic_date.month))
            if 0 <= diff <= 12:  # sanity: within 6 years
                ic_period_dist[diff] += 1
                ic_timing_records.append({
                    "name": normalize_name(name),
                    "entry": f"{entry['year']}/{entry['semester']}",
                    "ic_start": ic_date.strftime("%Y/%m"),
                    "semesters_after": diff,
                    "curso": names_map.get(name, ""),
                })

    ic_timing_records.sort(key=lambda x: x["semesters_after"])

    # average semesters to first IC (for those with entry data)
    timing_vals = [r["semesters_after"] for r in ic_timing_records]
    avg_timing = round(sum(timing_vals) / len(timing_vals), 1) if timing_vals else None

    ic_timing: dict = {
        "period_dist": dict(sorted(ic_period_dist.items())),
        "year_dist": dict(sorted(ic_year_dist.items())),
        "avg_semesters": avg_timing,
        "n_with_entry": len(ic_timing_records),
        "records": ic_timing_records,
    }

    # ---- graduation time (semesters from entry to graduation) ----
    # classify each formando: primary sponsor (paid wins) + fellowship names
    _PAID_SPONSORS = {"fapes", "ifes", "cnpq", "capes", "faperj", "fapesp", "fapemig"}
    _SPONSOR_DISPLAY = {"fapes": "Fapes", "ifes": "Ifes", "cnpq": "CNPq",
                        "capes": "CAPES", "faperj": "FAPERJ", "fapesp": "FAPESP", "fapemig": "FAPEMIG"}
    _person_primary_sponsor: dict[str, str] = {f["nome"].lower().strip(): "Sem IC" for f in formandos}
    _person_fel_type: dict[str, str] = {f["nome"].lower().strip(): "no_ic" for f in formandos}
    _person_fel_names: dict[str, set] = defaultdict(set)
    for _proj in adv_projects:
        for _adv in _proj.get("advisorships", []):
            _pn = mk_to_name.get(_match_key(_adv.get("person_name")))
            if not _pn or _pn not in _person_fel_type:
                continue
            _fel = _adv.get("fellowship") or {}
            _sponsor_raw = (_fel.get("sponsor_name") or "").strip()
            _sponsor_key = _sponsor_raw.lower()
            _fname = (_fel.get("name") or "").strip()
            if _fname:
                _person_fel_names[_pn].add(_fname)
            if _sponsor_key in _PAID_SPONSORS:
                _person_fel_type[_pn] = "paid"
                _person_primary_sponsor[_pn] = _SPONSOR_DISPLAY.get(_sponsor_key, _sponsor_raw)
            elif _person_fel_type[_pn] != "paid":
                _person_fel_type[_pn] = "volunteer"
                if _person_primary_sponsor[_pn] == "Sem IC":
                    _person_primary_sponsor[_pn] = "Voluntário"

    _grad_overall: list[int] = []
    _grad_by_curso: dict[str, list[int]] = defaultdict(list)
    _grad_ic: list[int] = []
    _grad_no_ic: list[int] = []
    _grad_ic_by_curso: dict[str, list[int]] = defaultdict(list)
    _grad_no_ic_by_curso: dict[str, list[int]] = defaultdict(list)
    _grad_paid: list[int] = []
    _grad_volunteer: list[int] = []
    _grad_paid_by_curso: dict[str, list[int]] = defaultdict(list)
    _grad_volunteer_by_curso: dict[str, list[int]] = defaultdict(list)
    _grad_by_sponsor: dict[str, list[int]] = defaultdict(list)
    _grad_by_sponsor_curso: dict[str, dict[str, list[int]]] = defaultdict(lambda: defaultdict(list))
    _grad_by_fel_name: dict[str, list[int]] = defaultdict(list)
    _grad_by_adm: dict[str, list[int]] = defaultdict(list)          # admission group → diffs
    _grad_by_adm_curso: dict[str, dict[str, list[int]]] = defaultdict(lambda: defaultdict(list))
    # ATRASO (diff − previsto): comparável entre cursos de tamanhos diferentes
    _grad_delay_overall: list[int] = []
    _grad_delay_by_adm: dict[str, list[int]] = defaultdict(list)
    _grad_delay_by_curso: dict[str, list[int]] = defaultdict(list)
    _grad_entry_years_by_curso: dict[str, list[int]] = defaultdict(list)  # impacto coorte
    _grad_per_person: dict[str, int] = {}  # name_lower → diff (shared by cross-analyses)

    if grad_semester:
        try:
            _gy, _gs = int(grad_semester[:4]), int(grad_semester[5])
        except (ValueError, IndexError):
            _gy, _gs = 0, 0

        if _gy:
            for f in formandos:
                entry = f.get("entry")
                if not entry:
                    continue
                _fgrad = f.get("grad_semester", grad_semester)
                try:
                    _fgy, _fgs = int(_fgrad[:4]), int(_fgrad[5])
                except Exception:
                    _fgy, _fgs = _gy, _gs
                diff = _sem_diff(entry["year"], entry["semester"], _fgy, _fgs)
                if 1 <= diff <= 50:  # sanity: at least 1 sem, at most 25 years
                    _grad_overall.append(diff)
                    _grad_by_curso[f["curso"]].append(diff)
                    _nk = f["nome"].lower().strip()
                    _ftype = _person_fel_type.get(_nk, "no_ic")
                    if _nk in matched_names:
                        _grad_ic.append(diff)
                        _grad_ic_by_curso[f["curso"]].append(diff)
                    else:
                        _grad_no_ic.append(diff)
                        _grad_no_ic_by_curso[f["curso"]].append(diff)
                    if _ftype == "paid":
                        _grad_paid.append(diff)
                        _grad_paid_by_curso[f["curso"]].append(diff)
                    elif _ftype == "volunteer":
                        _grad_volunteer.append(diff)
                        _grad_volunteer_by_curso[f["curso"]].append(diff)
                    # per-sponsor and per-fellowship-name tracking
                    _sp = _person_primary_sponsor.get(_nk, "Sem IC")
                    _grad_by_sponsor[_sp].append(diff)
                    _grad_by_sponsor_curso[_sp][f["curso"]].append(diff)
                    for _fn in _person_fel_names.get(_nk, set()):
                        _grad_by_fel_name[_fn].append(diff)
                    # admission group (cota vs ampla vs transf)
                    _ag = admission_group(f.get("admissao"))
                    _grad_by_adm[_ag].append(diff)
                    _grad_by_adm_curso[_ag][f["curso"]].append(diff)
                    # atraso = semestres além do previsto (comparável entre cursos)
                    _delay = diff - expected_semesters(f["curso"])
                    _grad_delay_overall.append(_delay)
                    _grad_delay_by_adm[_ag].append(_delay)
                    _grad_delay_by_curso[f["curso"]].append(_delay)
                    _grad_entry_years_by_curso[f["curso"]].append(entry["year"])
                    # shared lookup: name → diff (for cross-analyses)
                    _grad_per_person[_nk] = diff

    def _gstats(vals: list[int]) -> dict:
        if not vals:
            return {}
        sv = sorted(vals)
        n = len(sv)
        return {
            "n": n,
            "mean": round(sum(sv) / n, 1),
            "median": sv[n // 2],
            "min": sv[0],
            "max": sv[-1],
            "dist": dict(Counter(sv)),
        }

    def _gcategories(vals: list[int]) -> dict:
        transfers = [v for v in vals if v < 4]
        regular   = [v for v in vals if 4 <= v <= 24]
        extended  = [v for v in vals if v > 24]
        return {"transfers": len(transfers), "regular": len(regular), "extended": len(extended)}

    _all_cursos = sorted(set(list(_grad_by_curso.keys())))
    graduation_time = {
        "overall": _gstats(_grad_overall),
        "by_curso": {c: _gstats(v) for c, v in _grad_by_curso.items()},
        "categories": _gcategories(_grad_overall),
        "by_curso_categories": {c: _gcategories(v) for c, v in _grad_by_curso.items()},
        "ic_vs_no_ic": {
            "ic": _gstats(_grad_ic),
            "no_ic": _gstats(_grad_no_ic),
            "by_curso": {
                c: {
                    "ic": _gstats(_grad_ic_by_curso.get(c, [])),
                    "no_ic": _gstats(_grad_no_ic_by_curso.get(c, [])),
                }
                for c in _all_cursos
            },
        },
        "fellowship_type": {
            "paid":      _gstats(_grad_paid),
            "volunteer": _gstats(_grad_volunteer),
            "no_ic":     _gstats(_grad_no_ic),
            "by_curso": {
                c: {
                    "paid":      _gstats(_grad_paid_by_curso.get(c, [])),
                    "volunteer": _gstats(_grad_volunteer_by_curso.get(c, [])),
                    "no_ic":     _gstats(_grad_no_ic_by_curso.get(c, [])),
                }
                for c in _all_cursos
            },
        },
        "by_sponsor": {
            sp: {
                "overall": _gstats(vals),
                "by_curso": {c: _gstats(cv) for c, cv in _grad_by_sponsor_curso[sp].items()},
            }
            for sp, vals in _grad_by_sponsor.items()
        },
        "by_fel_name": {
            fn: _gstats(vals)
            for fn, vals in sorted(_grad_by_fel_name.items(), key=lambda x: -len(x[1]))
        },
        "by_admission": {
            ag: {
                "overall": _gstats(vals),
                "delay": _gstats(_grad_delay_by_adm.get(ag, [])),
                "by_curso": {c: _gstats(cv) for c, cv in _grad_by_adm_curso[ag].items()},
            }
            for ag, vals in _grad_by_adm.items()
        },
        "overall_delay": _gstats(_grad_delay_overall),
        "delay_by_curso": {c: _gstats(v) for c, v in _grad_delay_by_curso.items()},
        # impacto de coorte + leitura POR CURSO (isolado): cada curso na própria régua.
        # Cursos têm naturezas distintas (SI diurno 4a, ECA noturno 6a) — não comparar entre si.
        "cohort_impact": {
            c: {
                "n": len(_grad_delay_by_curso[c]),
                "expected": expected_semesters(c),
                "old_le_2015": sum(1 for y in _grad_entry_years_by_curso[c] if y <= 2015),
                "long_tail_ge6": sum(1 for d in _grad_delay_by_curso[c] if d >= 6),
                "on_time": sum(1 for d in _grad_delay_by_curso[c] if d <= 0),
                "early_plausible": sum(1 for d in _grad_delay_by_curso[c] if -2 <= d <= -1),
                "suspect": sum(1 for d in _grad_delay_by_curso[c] if d <= -3),
                "late": sum(1 for d in _grad_delay_by_curso[c] if d > 0),
                "atraso_mean": _gstats(_grad_delay_by_curso[c]).get("mean"),
                "atraso_median": _gstats(_grad_delay_by_curso[c]).get("median"),
            }
            for c in _all_cursos
        },
        "expected": dict(CURSO_EXPECTED_SEMESTERS),
    }

    # ---- cohort analysis (group formandos by entry year) ----
    _cohort_raw: dict[int, dict] = {}
    for _f in formandos:
        _entry = _f.get("entry")
        if not _entry:
            continue
        _yr = _entry["year"]
        _nk = _f["nome"].lower().strip()
        if _yr not in _cohort_raw:
            _cohort_raw[_yr] = {"total": 0, "ic": 0, "paid": 0, "volunteer": 0, "grad_times": []}
        _cohort_raw[_yr]["total"] += 1
        if _nk in research_names:   # pesquisa = SigPesq IC ∪ bolsistas FAPES (igual ao headline)
            _cohort_raw[_yr]["ic"] += 1
        _ft2 = _person_fel_type.get(_nk, "no_ic")
        if _ft2 == "paid":
            _cohort_raw[_yr]["paid"] += 1
        elif _ft2 == "volunteer":
            _cohort_raw[_yr]["volunteer"] += 1
        _gd = _grad_per_person.get(_nk)
        if _gd:
            _cohort_raw[_yr]["grad_times"].append(_gd)

    cohort_analysis = {
        yr: {
            "total": d["total"],
            "ic": d["ic"],
            "ic_pct": round(d["ic"] / d["total"] * 100, 1) if d["total"] else 0,
            "paid": d["paid"],
            "volunteer": d["volunteer"],
            "no_ic": d["total"] - d["ic"],
            "mean_grad": round(sum(d["grad_times"]) / len(d["grad_times"]), 1) if d["grad_times"] else None,
            "median_grad": sorted(d["grad_times"])[len(d["grad_times"]) // 2] if d["grad_times"] else None,
            "n_grad": len(d["grad_times"]),
        }
        for yr, d in sorted(_cohort_raw.items())
    }

    # ---- supervisor impact on graduation time ----
    _sup_student_diff: dict[str, dict[str, int]] = defaultdict(dict)  # sup → {name: diff}
    for _proj in adv_projects:
        for _adv in _proj.get("advisorships", []):
            _pid2 = _adv.get("person_id")
            if _pid2 not in matched_pids:
                continue
            _pname2 = pid_to_name.get(_pid2, "")
            _diff2 = _grad_per_person.get(_pname2)
            if _diff2 is None:
                continue
            _sname2 = normalize_name(_adv.get("supervisor_name") or "")
            if _sname2:
                _sup_student_diff[_sname2][_pname2] = _diff2

    supervisor_impact = {
        s: _gstats(list(diffs.values()))
        for s, diffs in sorted(
            _sup_student_diff.items(),
            key=lambda x: sum(x[1].values()) / len(x[1]) if x[1] else 99,
        )
        if len(diffs) >= 2
    }

    # ---- fellowship investment × graduation time ----
    _person_total_invest: dict[str, float] = defaultdict(float)
    for _pid3, _recs3 in person_fellowship_records.items():
        _pname3 = pid_to_name.get(_pid3, "")
        for _r in _recs3:
            if _r.get("value") and _r.get("days"):
                _person_total_invest[_pname3] += (_r["value"] or 0) * ((_r["days"] or 0) / 30.0)

    _val_groups: dict[str, list[int]] = defaultdict(list)
    for _pname4, _diff4 in _grad_per_person.items():
        if _pname4 not in matched_names:
            _val_groups["Sem IC"].append(_diff4)
            continue
        _invest = _person_total_invest.get(_pname4, 0.0)
        if _invest == 0:
            _val_groups["IC sem valor"].append(_diff4)
        elif _invest < 2000:
            _val_groups["R$ 1–2k"].append(_diff4)
        elif _invest < 5000:
            _val_groups["R$ 2–5k"].append(_diff4)
        else:
            _val_groups["R$ 5k+"].append(_diff4)

    fellowship_value_impact = {k: _gstats(v) for k, v in _val_groups.items() if v}

    # ---- IC continuity (1 vs 2+ projects) ----
    _ic_cont: dict[str, list[int]] = defaultdict(list)
    for _pid5, _projs5 in person_projects.items():
        _pname5 = pid_to_name.get(_pid5, "")
        _diff5 = _grad_per_person.get(_pname5)
        if _diff5 is None:
            continue
        _np = len(_projs5)
        _key5 = "1 projeto" if _np == 1 else ("2 projetos" if _np == 2 else "3+ projetos")
        _ic_cont[_key5].append(_diff5)

    ic_continuity = {k: _gstats(v) for k, v in _ic_cont.items() if v}

    # also no_ic for baseline
    ic_continuity["Sem IC"] = _gstats(_grad_no_ic)

    # ---- IC → TCC pipeline (Lattes) ----
    ic_tcc_pipeline: dict = {}
    if lattes:
        _tcc_by_name: dict[str, list[dict]] = defaultdict(list)
        for _r6 in lattes.get("tcc", []):
            _n6 = mk_to_name.get(_match_key(_r6.get("orientando")))
            if _n6:
                _tcc_by_name[_n6].append(_r6)

        _both_names6 = matched_names & set(_tcc_by_name.keys())
        _ic_only6 = len(matched_names - set(_tcc_by_name.keys()))
        _tcc_only6 = len(set(_tcc_by_name.keys()) - matched_names)

        _sigpesq_sups6: dict[str, set] = defaultdict(set)
        for _pid6, _sups6 in person_supervisors.items():
            _pname6 = pid_to_name.get(_pid6, "")
            _sigpesq_sups6[_pname6].update(_sups6)

        _same_sup6 = 0
        _diff_sup6 = 0
        _pipeline_recs6: list[dict] = []
        for _n6 in _both_names6:
            _sp_sups = {normalize_name(s).lower() for s in _sigpesq_sups6.get(_n6, set())}
            _tcc_sups = {normalize_name(_r.get("supervisor") or "").lower() for _r in _tcc_by_name[_n6] if _r.get("supervisor")}
            if _sp_sups & _tcc_sups:
                _same_sup6 += 1
                _pipeline_recs6.append({
                    "name": normalize_name(_n6),
                    "ic_sup": sorted(_sigpesq_sups6.get(_n6, set())),
                    "tcc_titles": [_r.get("titulo", "") for _r in _tcc_by_name[_n6]],
                })
            else:
                _diff_sup6 += 1

        ic_tcc_pipeline = {
            "ic_only": _ic_only6,
            "tcc_only": _tcc_only6,
            "both": len(_both_names6),
            "both_pct": round(len(_both_names6) / len(matched_names) * 100, 1) if matched_names else 0,
            "same_sup": _same_sup6,
            "diff_sup": _diff_sup6,
            "same_sup_pct": round(_same_sup6 / len(_both_names6) * 100, 1) if _both_names6 else 0,
            "pipeline_records": sorted(_pipeline_recs6, key=lambda x: x["name"])[:20],
        }

    # ---- Lattes cross-reference ----
    lattes_cross: dict = {}
    if lattes:
        ic_recs = lattes.get("ic", [])
        tcc_recs = lattes.get("tcc", [])

        lattes_ic_names: set[str] = {
            mk_to_name[_match_key(r["orientando"])]
            for r in ic_recs
            if _match_key(r["orientando"]) in mk_to_name
        }
        lattes_tcc_names: set[str] = {
            mk_to_name[_match_key(r["orientando"])]
            for r in tcc_recs
            if _match_key(r["orientando"]) in mk_to_name
        }
        lattes_any_names = lattes_ic_names | lattes_tcc_names

        # Per-person project counts from Lattes (accent-insensitive match)
        lattes_ic_count: dict[str, int] = defaultdict(int)
        for r in ic_recs:
            n = mk_to_name.get(_match_key(r["orientando"]))
            if n:
                lattes_ic_count[n] += 1

        lattes_tcc_count: dict[str, int] = defaultdict(int)
        for r in tcc_recs:
            n = mk_to_name.get(_match_key(r["orientando"]))
            if n:
                lattes_tcc_count[n] += 1

        # SigPesq per-person project count (already have person_projects)
        sigpesq_name_count: dict[str, int] = {}
        for n, pid in name_to_pid.items():
            sigpesq_name_count[n] = len(person_projects.get(pid, set()))

        # Union coverage
        union_names = matched_names | lattes_any_names
        new_via_lattes = sorted(lattes_any_names - matched_names)

        # Combined project count per person (IC deduplicated, TCC additive)
        combined_count: dict[str, int] = {}
        for n in union_names:
            sp = sigpesq_name_count.get(n, 0)
            lt = lattes_ic_count.get(n, 0)
            tcc = lattes_tcc_count.get(n, 0)
            # IC: take max (same project may appear in both sources)
            # TCC: additive (SigPesq doesn't track TCC)
            combined_count[n] = max(sp, lt) + tcc

        def _avg(vals: list) -> float:
            return round(sum(vals) / len(vals), 2) if vals else 0.0

        sp_vals = list(sigpesq_name_count.values())
        union_vals = list(combined_count.values())

        # Enriched: formandos where Lattes added TCC
        enriched = [
            {
                "name": normalize_name(n),
                "sigpesq": sigpesq_name_count.get(n, 0),
                "lattes_ic": lattes_ic_count.get(n, 0),
                "tcc": lattes_tcc_count.get(n, 0),
                "combined": combined_count[n],
            }
            for n in matched_names
            if combined_count.get(n, 0) > sigpesq_name_count.get(n, 0)
        ]
        enriched.sort(key=lambda x: -x["combined"])

        # Distribution: SigPesq vs Union
        sp_dist = dict(Counter(sp_vals))
        union_dist = dict(Counter(union_vals))

        lattes_cross = {
            "total_ic_lattes": len(ic_recs),
            "total_tcc_lattes": len(tcc_recs),
            "lattes_ic_formandos": len(lattes_ic_names),
            "lattes_tcc_formandos": len(lattes_tcc_names),
            "new_via_lattes": [normalize_name(n) for n in new_via_lattes],
            "sigpesq_n": len(matched_names),
            "union_n": len(union_names),
            "coverage_sigpesq_pct": round(len(matched_names) / total * 100, 1),
            "coverage_union_pct": round(len(union_names) / total * 100, 1),
            "avg_sigpesq": _avg(sp_vals),
            "avg_union": _avg(union_vals),
            "delta_avg": round(_avg(union_vals) - _avg(sp_vals), 2),
            "delta_pct": round((_avg(union_vals) / _avg(sp_vals) - 1) * 100, 1) if sp_vals else 0,
            "enriched_count": len(enriched),
            "enriched_top": enriched[:10],
            "sp_dist": sp_dist,
            "union_dist": union_dist,
        }

    # ---- IC recovery: recuperar alunos com IC fora do SigPesq via Lattes ----
    # Objetivo: dos formandos "sem IC" no SigPesq, quais aparecem com iniciação
    # científica no Lattes? Dois níveis de confiança:
    #   exato  = mesmo nome (sem acento)  → alta confiança
    #   provável = mesmo primeiro+último nome → revisar manualmente
    ic_recovery: dict = {}
    if lattes:
        _sem_ic_names = set(names_map) - matched_names  # formandos sem IC no SigPesq
        # nível 1: match exato (sem acento) com IC do Lattes
        _lat_ic_canon = {
            mk_to_name[_match_key(r["orientando"])]
            for r in lattes.get("ic", [])
            if _match_key(r["orientando"]) in mk_to_name
        }
        _recovered_exact = sorted(_lat_ic_canon & _sem_ic_names)
        # nível 2: match aproximado (primeiro+último nome) para os ainda sem IC
        _remaining = _sem_ic_names - set(_recovered_exact)
        _lat_ic_tokens = [
            _token_set(r["orientando"]) for r in lattes.get("ic", [])
        ]
        _lat_ic_tokens = [t for t in _lat_ic_tokens if t]
        _recovered_fuzzy = sorted(
            n for n in _remaining
            if _token_set(n) and any(_token_set(n) <= t for t in _lat_ic_tokens)
        )
        _combined_ic = matched_names | set(_recovered_exact) | set(_recovered_fuzzy)
        ic_recovery = {
            "sigpesq_ic": len(matched_names),
            "sem_ic_sigpesq": len(_sem_ic_names),
            "recovered_exact": [normalize_name(n) for n in _recovered_exact],
            "recovered_exact_n": len(_recovered_exact),
            "recovered_fuzzy": [normalize_name(n) for n in _recovered_fuzzy],
            "recovered_fuzzy_n": len(_recovered_fuzzy),
            "combined_ic_n": len(_combined_ic),
            "sem_ic_after": total - len(_combined_ic),
            "coverage_before_pct": round(len(matched_names) / total * 100, 1),
            "coverage_after_pct": round(len(_combined_ic) / total * 100, 1),
        }

    # ---- bolsistas FAPES × formandos: quais já se formaram + tipos de bolsa ----
    bolsistas_cross: dict = {}
    _unicos = (bolsistas or {}).get("bolsistas_unicos", [])
    _alocacoes = (bolsistas or {}).get("alocacoes", [])
    if _unicos:
        _form_by_name: dict[str, dict] = {
            f["nome"].lower().strip(): f for f in formandos
        }
        _matched_b: list[dict] = []
        _formado_canon: set[str] = set()  # canonical names dos bolsistas formados
        for b in _unicos:
            canon = mk_to_name.get(_match_key(b.get("bolsista_pesquisador_nome")))
            if not canon or canon not in _form_by_name:
                continue
            f = _form_by_name[canon]
            _formado_canon.add(canon)
            _matched_b.append({
                "nome": normalize_name(b.get("bolsista_pesquisador_nome", "")),
                "curso": f["curso"],
                "admissao_grupo": admission_group(f.get("admissao")),
                "com_ic_sigpesq": canon in matched_names,
                "valor_alocado": b.get("valor_alocado_total", 0) or 0,
                "valor_pago": b.get("valor_pago_total", 0) or 0,
                "total_projetos": b.get("total_projetos", 0) or 0,
                "total_alocacoes": b.get("total_alocacoes", 0) or 0,
            })
        _matched_b.sort(key=lambda x: -x["valor_alocado"])
        _by_curso = Counter(m["curso"] for m in _matched_b)
        _by_adm = Counter(m["admissao_grupo"] for m in _matched_b)

        # ---- tipos de bolsa FAPES dos formados (a partir das alocações) ----
        _bolsa_people: dict[str, set] = defaultdict(set)   # sigla → {nome canon}
        _bolsa_nome: dict[str, str] = {}
        _bolsa_valores: dict[str, list] = defaultdict(list)
        _bolsa_alocado: dict[str, float] = defaultdict(float)
        for a in _alocacoes:
            _cn = mk_to_name.get(_match_key(a.get("bolsista_pesquisador_nome")))
            if _cn not in _formado_canon:
                continue
            _sig = a.get("bolsa_sigla") or "?"
            _bolsa_people[_sig].add(_cn)
            _bolsa_nome.setdefault(_sig, a.get("bolsa_nome") or _sig)
            _v = a.get("bolsa_nivel_valor")
            if _v:
                _bolsa_valores[_sig].append(_v)
            _bolsa_alocado[_sig] += a.get("valor_alocado_total", 0) or 0
        por_bolsa = sorted(
            ({
                "sigla": _sig,
                "nome": _bolsa_nome.get(_sig, _sig),
                "formados": len(_people),
                "valor_min": min(_bolsa_valores[_sig]) if _bolsa_valores[_sig] else 0,
                "valor_max": max(_bolsa_valores[_sig]) if _bolsa_valores[_sig] else 0,
                "valor_alocado": round(_bolsa_alocado[_sig]),
            } for _sig, _people in _bolsa_people.items()),
            key=lambda x: -x["formados"],
        )

        bolsistas_cross = {
            "total_bolsistas": len(_unicos),
            "formaram": len(_matched_b),
            "nao_formandos": len(_unicos) - len(_matched_b),
            "com_ic_sigpesq": sum(1 for m in _matched_b if m["com_ic_sigpesq"]),
            "novos_pesquisa": sum(1 for m in _matched_b if not m["com_ic_sigpesq"]),
            "por_curso": dict(_by_curso),
            "por_ingresso": dict(_by_adm),
            "por_bolsa": por_bolsa,
            "valor_alocado_total": sum(m["valor_alocado"] for m in _matched_b),
            "valor_pago_total": sum(m["valor_pago"] for m in _matched_b),
            "registros": _matched_b,
        }

    # ---- admission form (cotas) analysis — deduped students only ----
    adm_group_counts: Counter = Counter()
    adm_category_counts: Counter = Counter()
    adm_flag_counts: Counter = Counter()
    adm_group_curso: dict[str, Counter] = defaultdict(Counter)
    adm_group_total: dict[str, int] = defaultdict(int)   # group → total students
    adm_group_ic: dict[str, int] = defaultdict(int)      # group → students with IC
    adm_flag_total: dict[str, int] = defaultdict(int)
    adm_flag_ic: dict[str, int] = defaultdict(int)
    adm_known = 0
    # fellowship cross: group → counts of paid / volunteer / no_ic + sponsor mix
    adm_grp_paid: Counter = Counter()
    adm_grp_vol: Counter = Counter()
    adm_grp_noic: Counter = Counter()
    adm_grp_sponsor: dict[str, Counter] = defaultdict(Counter)
    # name → set of paid sponsors (display-cased), from fellowship records
    name_sponsors: dict[str, set] = defaultdict(set)
    for _pidS, _recsS in person_fellowship_records.items():
        _nmS = pid_to_name.get(_pidS, "")
        for _rS in _recsS:
            if _rS["sponsor"] and _rS["sponsor"] != "Voluntário":
                name_sponsors[_nmS].add(_rS["sponsor"])
    for f in formandos:
        desc = f.get("admissao")
        g = admission_group(desc)
        adm_group_counts[g] += 1
        adm_group_curso[f["curso"]][g] += 1
        if desc:
            adm_known += 1
            adm_category_counts[normalize_str(desc) if False else desc] += 1
        nk = f["nome"].lower().strip()
        has_ic = nk in research_names
        adm_group_total[g] += 1
        if has_ic:
            adm_group_ic[g] += 1
        # fellowship type cross
        ftype = _person_fel_type.get(nk, "no_ic")
        if not has_ic:
            adm_grp_noic[g] += 1
        elif ftype == "paid":
            adm_grp_paid[g] += 1
            for _sp in name_sponsors.get(nk, set()):
                adm_grp_sponsor[g][_sp] += 1
        else:
            adm_grp_vol[g] += 1
        for fl in admission_flags(desc):
            adm_flag_counts[fl] += 1
            adm_flag_total[fl] += 1
            if has_ic:
                adm_flag_ic[fl] += 1

    admission = {
        "total": total,
        "known": adm_known,
        "unknown": total - adm_known,
        "group_counts": dict(adm_group_counts),
        "category_counts": dict(adm_category_counts.most_common()),
        "flag_counts": dict(adm_flag_counts.most_common()),
        "group_curso": {c: dict(v) for c, v in adm_group_curso.items()},
        "group_ic": {
            g: {"total": adm_group_total[g], "ic": adm_group_ic.get(g, 0),
                "pct": round(adm_group_ic.get(g, 0) / adm_group_total[g] * 100, 1)
                if adm_group_total[g] else 0.0}
            for g in adm_group_total
        },
        "flag_ic": {
            fl: {"total": adm_flag_total[fl], "ic": adm_flag_ic.get(fl, 0),
                 "pct": round(adm_flag_ic.get(fl, 0) / adm_flag_total[fl] * 100, 1)
                 if adm_flag_total[fl] else 0.0}
            for fl in adm_flag_total
        },
        "group_fellowship": {
            g: {
                "total": adm_group_total[g],
                "paid": adm_grp_paid.get(g, 0),
                "volunteer": adm_grp_vol.get(g, 0),
                "no_ic": adm_grp_noic.get(g, 0),
                "ic": adm_grp_paid.get(g, 0) + adm_grp_vol.get(g, 0),
                "pct_paid_total": round(adm_grp_paid.get(g, 0) / adm_group_total[g] * 100, 1)
                if adm_group_total[g] else 0.0,
                "pct_paid_ic": round(
                    adm_grp_paid.get(g, 0)
                    / (adm_grp_paid.get(g, 0) + adm_grp_vol.get(g, 0)) * 100, 1)
                if (adm_grp_paid.get(g, 0) + adm_grp_vol.get(g, 0)) else 0.0,
                "sponsors": dict(adm_grp_sponsor.get(g, Counter()).most_common()),
            }
            for g in adm_group_total
        },
    }

    # ---- integrated cross-source stats (FAPES / FACTO / mestrado PPComp) ----
    _sup_formando_counts = {name: len(pids) for name, pids in sup_unique.items()}
    _ic_grad_invest = sum(sponsor_investment.values())
    _bolsistas_alocado = bolsistas_cross.get("valor_alocado_total", 0) or 0
    integrated = compute_integrated(
        _sup_formando_counts, _ic_grad_invest, _bolsistas_alocado
    )

    return {
        "total": total,
        "admission": admission,
        "sem_registro": sem_registro,
        "with_research": with_research,
        "with_research_sigpesq": with_research_sigpesq,
        "research_bolsistas_only": len(bolsista_formado_names - matched_names),
        "with_paid_bolsa": len(paid_names | bolsista_formado_names),
        "sem_pesquisa": sem_pesquisa,
        "pct_research": round(with_research / total * 100, 1),
        "pct_research_sigpesq": round(with_research_sigpesq / total * 100, 1),
        "fellowship_counts": fellowship_counts,
        "sponsor_counts": sponsor_counts,
        "sponsor_investment": dict(sponsor_investment),
        "sponsor_fellowship_unique": {k: dict(v) for k, v in sponsor_fellowship_unique.items()},
        "curso_total": dict(curso_total),
        "curso_with": dict(curso_with),
        "curso_sponsor": {k: dict(v) for k, v in curso_sponsor.items()},
        "proj_dist": dict(proj_dist),
        "with_projects": with_projects,
        "durations_n": len(durations),
        "dur_mean": dur_mean,
        "dur_median": dur_median,
        "dur_bins": dict(dur_bins),
        "prog": dict(prog),
        "multi_bolsa": multi_bolsa,
        "sup_dist": dict(sup_dist),
        "top_sups": [(s, len(pids)) for s, pids in top_sups],
        "rg_top": rg_top_list,
        "rg_top_stats": rg_top_stats,
        "ka_top": ka_counter.most_common(15),
        "total_with_fellowship": len(set().union(*fellowship_persons.values())) if fellowship_persons else 0,
        "lattes_cross": lattes_cross,
        "ic_recovery": ic_recovery,
        "bolsistas_cross": bolsistas_cross,
        "ic_timing": ic_timing,
        "graduation_time": graduation_time,
        "cohort_analysis": cohort_analysis,
        "supervisor_impact": supervisor_impact,
        "fellowship_value_impact": fellowship_value_impact,
        "ic_continuity": ic_continuity,
        "ic_tcc_pipeline": ic_tcc_pipeline,
        "integrated": integrated,
    }


# ---------------------------------------------------------------------------
# HTML rendering helpers
# ---------------------------------------------------------------------------

CSS = """
:root {
  --bg:      #f4f8f5;
  --surface: #ffffff;
  --card:    #ffffff;
  --border:  #e2ebe4;
  --green:   #0f7a40;
  --green2:  #1f9d57;
  --amber:   #b8860b;
  --blue:    #2f6fb0;
  --red:     #c0392b;
  --gray:    #71857a;
  --text:    #16241a;
  --sub:     #5f7268;
  --font:    'Inter', 'Segoe UI', system-ui, -apple-system, sans-serif;
}
* { margin: 0; padding: 0; box-sizing: border-box; }
html { -webkit-print-color-adjust: exact; print-color-adjust: exact; }
@media print {
  body { background:#fff !important; }
  .section, .stat-card { box-shadow:none !important; }
}
body { background:var(--bg); color:var(--text); font-family:var(--font);
       min-height:100vh; padding:40px 24px 60px; }
header { text-align:center; margin-bottom:48px; }
header .eyebrow { font-size:11px; letter-spacing:3px; text-transform:uppercase;
                  color:var(--green); margin-bottom:10px; }
header h1 { font-size:clamp(22px,4vw,34px); font-weight:700; color:var(--text);
            margin-bottom:12px; }
header p { color:var(--sub); font-size:14px; }
.stats-grid { display:grid; grid-template-columns:repeat(4,1fr); gap:16px; margin-bottom:28px; }
.stat-card { position:relative; background:var(--card); border:1px solid var(--border);
             border-radius:12px; padding:22px 20px; overflow:hidden;
             box-shadow:0 1px 2px rgba(16,40,24,.04), 0 4px 14px rgba(16,40,24,.05); }
.stat-card::before { content:''; position:absolute; top:0; left:0; right:0; height:3px; background:var(--green); }
.stat-card .number { font-size:36px; font-weight:800; color:var(--green); line-height:1; letter-spacing:-.01em; }
.stat-card .pct { font-size:13px; color:var(--sub); margin-top:2px; }
.stat-card .label { font-size:12px; color:var(--sub); margin-top:8px; }
.section { background:var(--card); border:1px solid var(--border); border-radius:12px;
           padding:24px; margin-bottom:20px;
           box-shadow:0 1px 2px rgba(16,40,24,.04), 0 4px 14px rgba(16,40,24,.05); }
.section h2 { font-size:15px; font-weight:600; color:var(--text); margin-bottom:16px; }
.section .sub { font-size:12px; color:var(--sub); margin-bottom:16px; }
.bar-row { display:flex; align-items:center; gap:8px; margin-bottom:6px; }
.bar-row .lbl { font-size:11px; color:var(--text); width:180px; flex-shrink:0; }
.bar-row .lbl.sm { width:80px; }
.bar-row .lbl.md { width:120px; }
.bar-track { flex:1; height:12px; background:var(--border); border-radius:3px; overflow:hidden; }
.bar-track.sm { height:10px; }
.bar-fill { height:100%; }
.bar-row .val { font-size:11px; width:24px; }
.grid2 { display:grid; grid-template-columns:1fr 1fr; gap:16px; }
.grid3 { display:grid; grid-template-columns:repeat(3,1fr); gap:12px; }
.grid4 { display:grid; grid-template-columns:repeat(4,1fr); gap:12px; }
.mini-card { background:#f7faf8; border:1px solid var(--border); border-radius:8px; padding:16px; }
.mini-card .agency { font-size:11px; letter-spacing:2px; text-transform:uppercase; margin-bottom:8px; }
.mini-card .big { font-size:28px; font-weight:700; color:var(--green); line-height:1; }
.mini-card .tiny { font-size:11px; color:var(--sub); }
.mini-card .divider { border-top:1px solid var(--border); margin:10px 0; padding-top:8px; }
.pill-row { display:flex; justify-content:space-between; font-size:11px; margin-bottom:3px; }
.prog-card { background:#f7faf8; border:1px solid var(--border); border-radius:6px;
             padding:12px; text-align:center; }
.prog-card .pt { font-size:10px; color:var(--sub); text-transform:uppercase;
                 letter-spacing:1px; margin-bottom:6px; }
.prog-card .pn { font-size:24px; font-weight:700; }
.prog-card .ps { font-size:10px; color:var(--sub); margin-top:4px; }
.note { font-size:11px; color:var(--sub); line-height:1.6; padding:10px 12px;
        background:#f7faf8; border-left:3px solid var(--sub); border-radius:4px; margin-top:12px; }
.note strong { color:var(--text); }
.decision { font-size:11px; color:var(--text); line-height:1.65; padding:11px 13px;
            background:#eef4fb; border-left:3px solid var(--blue); border-radius:4px; margin-top:14px; }
.decision .d-lead { font-weight:700; color:var(--blue); text-transform:uppercase;
                    letter-spacing:.05em; font-size:10px; display:block; margin-bottom:4px; }
.decision strong { color:var(--text); }
.list-row { display:flex; justify-content:space-between; font-size:11px;
            padding:4px 8px; background:#f7faf8; border-radius:3px; margin-bottom:4px; }
footer { text-align:center; margin-top:48px; font-size:11px; color:var(--sub); }
"""


def bar(label: str, value: int, max_val: int, color: str,
        lbl_class: str = "") -> str:
    pct = value / max_val * 100 if max_val else 0
    return (
        f'<div class="bar-row">'
        f'<span class="lbl {lbl_class}">{label}</span>'
        f'<div class="bar-track"><div class="bar-fill" '
        f'style="width:{pct:.1f}%;background:{color};"></div></div>'
        f'<span class="val" style="color:{color};">{value}</span>'
        f'</div>'
    )


def mini_card_agency(name: str, color: str, count: int,
                     fellowship_breakdown: dict, investment: float) -> str:
    rows = "".join(
        f'<div class="pill-row"><span style="color:var(--text);">{k}</span>'
        f'<span style="color:var(--green);">{v}</span></div>'
        for k, v in sorted(fellowship_breakdown.items())
    )
    inv_str = (
        f'R$ {investment:,.0f}'.replace(",", ".")
        if investment > 0 else "R$ 0"
    )
    return f"""
    <div class="mini-card">
      <div class="agency" style="color:{color};">{name}</div>
      <div class="big">{count}</div>
      <div class="tiny">formandos</div>
      <div class="divider">
        <div style="font-size:10px;color:var(--sub);margin-bottom:6px;">Bolsas concedidas</div>
        {rows}
        <div style="margin-top:8px;">
          <div style="font-size:10px;color:var(--sub);">Investimento acumulado</div>
          <div style="font-size:13px;color:{color};font-weight:600;">
            {inv_str}<span style="font-size:10px;color:var(--sub);">/mês·bolsa</span>
          </div>
        </div>
      </div>
    </div>"""


# Texto interpretativo por seção: o que o dado significa + como usar para decidir.
# Indexado pelo título exato passado a section(); aplicado automaticamente.
DECISIONS: dict[str, str] = {
    "Distribuição por curso":
        "<strong>O que é:</strong> quantos formandos cada curso entrega e que fração fez pesquisa. "
        "<strong>Como decidir:</strong> compare a <em>proporção</em> com pesquisa (não o número absoluto — os cursos têm tamanhos diferentes). "
        "Curso com baixa adesão à IC é alvo prioritário de campanha de captação e de novas cotas de bolsa.",
    "Forma de ingresso (cotas)":
        "<strong>O que é:</strong> participação em pesquisa por forma de ingresso (cotistas × ampla concorrência). "
        "<strong>Como decidir:</strong> se cotistas têm menos IC ou menos bolsa <em>paga</em>, há barreira de equidade — "
        "direcione editais e bolsas afirmativas para igualar o acesso à pesquisa.",
    "Tipo de bolsa":
        "<strong>O que é:</strong> mix de modalidades (PIBIC/PIVIC/PIBITI…). "
        "<strong>Como decidir:</strong> muito voluntário (PIVIC/PIVITI) revela demanda reprimida por bolsa paga — "
        "use como argumento quantitativo para pleitear mais cotas remuneradas.",
    "Agências de fomento":
        "<strong>O que é:</strong> quem financia (CNPq, Ifes, Fapes) e o volume de cada uma. "
        "<strong>Como decidir:</strong> dependência forte de uma única agência é risco — diversifique. "
        "Onde o valor captado é baixo, concentre esforço de submissão.",
    "Curso × agência de fomento":
        "<strong>O que é:</strong> qual curso capta de qual agência. "
        "<strong>Como decidir:</strong> curso sem acesso a determinada agência deve ser orientado a submeter lá; "
        "replique a prática de quem já capta bem.",
    "Progressão de bolsa":
        "<strong>O que é:</strong> alunos que mudaram de status (voluntário→pago, entre agências). "
        "<strong>Como decidir:</strong> vol→pago é trajetória saudável de retenção; "
        "pago→voluntário pode sinalizar perda de financiamento a investigar.",
    "Orientadores por formando":
        "<strong>O que é:</strong> quantos orientadores diferentes cada aluno teve. "
        "<strong>Como decidir:</strong> 1 orientador indica vínculo estável; muitos podem indicar troca/abandono. "
        "Pesquisa concentrada em poucos orientadores é risco de sobrecarga e de dependência — distribua a mentoria.",
    "Grupos de pesquisa":
        "<strong>O que é:</strong> grupos que mais formam pesquisadores, com divisão pago/voluntário/cotista. "
        "<strong>Como decidir:</strong> grupos ativos são onde o investimento em bolsa rende mais; "
        "grupos com muito voluntário são candidatos diretos a novas cotas pagas.",
    "Áreas de conhecimento":
        "<strong>O que é:</strong> temas em que a pesquisa dos formandos se concentra. "
        "<strong>Como decidir:</strong> alinhe editais e contratação docente às áreas fortes; "
        "área ausente ou rarefeita é lacuna estratégica a cobrir.",
    "Produção científica":
        "<strong>O que é:</strong> artigos registrados no sistema para os formandos. "
        "<strong>Como decidir:</strong> baixa produção registrada quase sempre é subnotificação — "
        "cruze com o Lattes e melhore o cadastro antes de concluir que há pouca produção.",
    "Quando os formandos entraram na IC":
        "<strong>O que é:</strong> em que ponto do curso o aluno começou a iniciação científica. "
        "<strong>Como decidir:</strong> entrada tardia desperdiça tempo de pesquisa — "
        "antecipe a divulgação da IC para os primeiros períodos.",
    "Bolsistas FAPES formados":
        "<strong>O que é:</strong> bolsistas FAPES que já se formaram, por curso e tipo de bolsa. "
        "<strong>Como decidir:</strong> é a medida de retorno do investimento FAPES — "
        "bolsista que se formou é sucesso; cruze com a continuidade no mestrado para ver o efeito de longo prazo.",
    "Recuperação de alunos com IC via Lattes":
        "<strong>O que é:</strong> alunos com IC que o SigPesq não capturou, recuperados via Lattes. "
        "<strong>Como decidir:</strong> mede o buraco de registro do SigPesq — "
        "a cobertura real de pesquisa é maior que a do sistema; priorize a melhoria do cadastro.",
    "Cruzamento Lattes × SigPesq":
        "<strong>O que é:</strong> o quanto o Lattes amplia os projetos por formando além do SigPesq. "
        "<strong>Como decidir:</strong> delta grande = SigPesq subnotifica de forma relevante; "
        "justifica investir na integração automática das duas bases.",
    "Tempo de formação":
        "<strong>O que é:</strong> atraso sobre o currículo de <em>cada</em> curso (SI 8 sem · ECA 12 sem) — não o tempo absoluto. "
        "<strong>Como decidir:</strong> compare sempre pelo atraso (a régua absoluta mistura cursos e engana). "
        "Concentração de atraso longo pede revisão de oferta de disciplinas e pré-requisitos.",
    "Análise de coorte por ano de ingresso":
        "<strong>O que é:</strong> desempenho por turma de entrada, cada curso na própria régua. "
        "<strong>Como decidir:</strong> atraso que cresce em coortes recentes aponta problema sistêmico daquele período "
        "(mudança curricular, evasão) — investigue a turma, não o aluno.",
    "Impacto do orientador no tempo de formação":
        "<strong>O que é:</strong> tempo médio de formação dos alunos de cada orientador. "
        "<strong>Como decidir:</strong> <em>não</em> é ranking de qualidade — reflete o perfil dos alunos orientados tanto quanto o orientador. "
        "Use para identificar quem precisa de apoio e para difundir boas práticas dos mais eficientes.",
    "Investimento em bolsa × tempo de formação":
        "<strong>O que é:</strong> tempo de formação por faixa de valor recebido em bolsa. "
        "<strong>Como decidir:</strong> se mais investimento corresponde a formação mais rápida, "
        "há base para ampliar valor e duração das bolsas — não só o número de cotas.",
    "Continuidade na IC × tempo de formação":
        "<strong>O que é:</strong> desfecho de quem fez 1 projeto de IC versus vários. "
        "<strong>Como decidir:</strong> se a continuidade melhora o resultado, incentive a <em>renovação</em> de projetos, "
        "não apenas a primeira entrada na IC.",
    "Pipeline IC → TCC":
        "<strong>O que é:</strong> quantos seguem da IC para o TCC e se mantêm o mesmo orientador. "
        "<strong>Como decidir:</strong> alta continuidade com mesmo orientador indica vínculo forte e maturação; "
        "baixa continuidade é oportunidade de integrar IC e TCC numa trilha única.",
    "Bolsa × tempo de formação":
        "<strong>O que é:</strong> comparação do tempo de formação entre quem teve bolsa e quem não teve. "
        "<strong>Como decidir:</strong> se a bolsa acelera a formação, é argumento direto para ampliar a oferta de bolsas.",
    "Ecossistema do orientador — IC × projetos FAPES":
        "<strong>O que é:</strong> orientadores de IC dos formandos que também coordenam grandes projetos FAPES. "
        "<strong>Como decidir:</strong> são os professores-chave que ligam a graduação ao fomento de grande porte — "
        "proteja e amplie a capacidade deles. Forte concentração em poucos nomes é risco de continuidade a mitigar.",
    "Pipeline graduação → mestrado (PPComp)":
        "<strong>O que é:</strong> quantos formandos seguem para o mestrado e o desfecho do programa. "
        "<strong>Como decidir:</strong> pipeline fino pede trilha explícita IC→TCC→mestrado; "
        "evasão alta no mestrado pede revisão de acolhimento e de bolsas de pós.",
    "Painel de fomento — da semente IC ao portfólio":
        "<strong>O que é:</strong> a escala do fomento — bolsa de IC frente aos projetos FAPES/FACTO. "
        "<strong>Como decidir:</strong> a IC é semente barata com alto efeito alavanca; proteja o orçamento dela. "
        "O alerta de projetos com prazo vencido é ação imediata de gestão.",
}


def _decision_box(text: str | None) -> str:
    if not text:
        return ""
    return (
        f'<div class="decision"><span class="d-lead">▸ Como ler e decidir</span>{text}</div>'
    )


def section(title: str, sub: str, body: str, border_color: str = "",
            decision: str | None = None) -> str:
    style = f' style="border-color:{border_color};"' if border_color else ""
    dbox = _decision_box(decision if decision is not None else DECISIONS.get(title, ""))
    return (
        f'<div class="section"{style}>'
        f'<h2>{title}</h2>'
        f'<div class="sub">{sub}</div>'
        f'{body}'
        f'{dbox}'
        f'</div>'
    )


# ---------------------------------------------------------------------------
# Section generators
# ---------------------------------------------------------------------------

def _sec_stats(s: dict) -> str:
    pct_sem = round(s["sem_pesquisa"] / s["total"] * 100, 1)
    return f"""
  <div class="stats-grid">
    <div class="stat-card">
      <div class="number">{s['total']}</div>
      <div class="label">formandos únicos</div>
    </div>
    <div class="stat-card">
      <div class="number">{s['with_research']}</div>
      <div class="pct">{s['pct_research']}% do total</div>
      <div class="label">com participação em pesquisa</div>
    </div>
    <div class="stat-card">
      <div class="number">{s['sem_pesquisa']}</div>
      <div class="pct">{pct_sem}% do total</div>
      <div class="label">sem informação de participação em pesquisa</div>
      <div class="pct" style="margin-top:6px;font-style:italic;">não há registro que confirme se participaram ou não de projetos de pesquisa</div>
    </div>
    <div class="stat-card">
      <div class="number">{s['sem_registro']}</div>
      <div class="label">sem registro no SigPesq</div>
    </div>
  </div>
  {_decision_box(
      "<strong>O que é:</strong> o panorama macro — total de formandos e a fatia que passou por pesquisa. "
      "<strong>Como decidir:</strong> o <strong>% com pesquisa</strong> é o indicador-chave da instituição; "
      "defina meta de elevação ano a ano. Atenção: &lsquo;sem informação&rsquo; não é o mesmo que &lsquo;sem pesquisa&rsquo; — "
      "parte é apenas falta de registro, então trate a cobertura de cadastro antes de ler como baixa adesão."
  )}"""


def _sec_curso(s: dict) -> str:
    rows = ""
    for curso, total_c in sorted(s["curso_total"].items()):
        with_c = s["curso_with"].get(curso, 0)
        pct = round(with_c / total_c * 100, 1) if total_c else 0
        short = "ECA" if "Controle" in curso else "SI"
        rows += (
            f'<div style="margin-bottom:16px;">'
            f'<div style="font-size:12px;font-weight:600;margin-bottom:8px;">'
            f'{curso} <span style="color:var(--sub);">({total_c})</span></div>'
            f'{bar(f"com pesquisa", with_c, total_c, "var(--green)", "md")}'
            f'{bar(f"sem informação", total_c-with_c, total_c, "var(--gray)", "md")}'
            f'<div style="font-size:11px;color:var(--sub);margin-top:4px;">'
            f'{pct}% do {short} com participação confirmada em pesquisa</div>'
            f'</div>'
        )
    return section("Distribuição por curso", "formandos por curso e participação", rows)


def _sec_admission(s: dict) -> str:
    a = s.get("admission")
    if not a or not a.get("known"):
        return ""
    total = a["total"]
    GCOL = {
        "Ampla Concorrência": "var(--blue)",
        "Cotas / Reserva de vagas": "var(--green)",
        "Transferência": "var(--amber)",
        "Sem informação": "var(--gray)",
    }

    # --- top-level groups ---
    g_order = ["Ampla Concorrência", "Cotas / Reserva de vagas", "Transferência",
               "Sem informação"]
    g_max = max(a["group_counts"].values()) if a["group_counts"] else 1
    group_bars = "".join(
        bar(f"{g} ({round(a['group_counts'][g]/total*100,1)}%)",
            a["group_counts"][g], g_max, GCOL.get(g, "var(--sub)"), "md")
        for g in g_order if a["group_counts"].get(g)
    )

    # --- categories (raw labels) ---
    cats = a["category_counts"]
    c_max = max(cats.values()) if cats else 1
    cat_bars = "".join(
        bar(admission_label(k), v, c_max, "var(--green2)")
        for k, v in cats.items()
    )

    # --- quota attributes ---
    flags = a["flag_counts"]
    f_max = max(flags.values()) if flags else 1
    FCOL = {"Escola Pública": "var(--green)", "PPI": "var(--green2)",
            "Renda": "var(--amber)", "Ação Afirmativa": "var(--blue)",
            "Pessoa c/ Deficiência": "var(--red)"}
    flag_bars = "".join(
        bar(f"{k} ({round(v/total*100,1)}%)", v, f_max, FCOL.get(k, "var(--sub)"), "md")
        for k, v in flags.items()
    )

    # --- group × curso ---
    gc_rows = ""
    for curso, gc in sorted(a["group_curso"].items()):
        tot_c = sum(gc.values())
        short = "ECA" if "Controle" in curso else "SI"
        bars = "".join(
            bar(g, gc.get(g, 0), tot_c, GCOL.get(g, "var(--sub)"), "md")
            for g in g_order if gc.get(g)
        )
        gc_rows += (
            f'<div style="margin-bottom:16px;">'
            f'<div style="font-size:12px;font-weight:600;margin-bottom:8px;">'
            f'{short} — {curso} <span style="color:var(--sub);">({tot_c})</span></div>'
            f'{bars}</div>'
        )

    # --- cota × research participation ---
    gic = a["group_ic"]
    ic_rows = ""
    for g in g_order:
        d = gic.get(g)
        if not d or not d["total"]:
            continue
        ic_rows += (
            f'<div class="bar-row">'
            f'<span class="lbl md">{g}</span>'
            f'<div class="bar-track"><div class="bar-fill" '
            f'style="width:{d["pct"]:.1f}%;background:{GCOL.get(g,"var(--sub)")};"></div></div>'
            f'<span class="val" style="width:90px;color:{GCOL.get(g,"var(--sub)")};">'
            f'{d["ic"]}/{d["total"]} · {d["pct"]}%</span></div>'
        )

    # --- explicit numeric tables ---
    def _tbl(headers: list[str], rows: list[list[str]]) -> str:
        th = "".join(
            f'<th style="text-align:{"left" if i==0 else "right"};padding:6px 10px;'
            f'border-bottom:1px solid var(--border);color:var(--sub);font-weight:600;'
            f'font-size:11px;">{h}</th>'
            for i, h in enumerate(headers)
        )
        trs = ""
        for r in rows:
            tds = "".join(
                f'<td style="text-align:{"left" if i==0 else "right"};padding:5px 10px;'
                f'border-bottom:1px solid var(--border);font-size:11px;'
                f'{"color:var(--green);" if i>0 else ""}">{c}</td>'
                for i, c in enumerate(r)
            )
            trs += f"<tr>{tds}</tr>"
        return (
            '<table style="width:100%;border-collapse:collapse;margin-bottom:8px;">'
            f'<thead><tr>{th}</tr></thead><tbody>{trs}</tbody></table>'
        )

    grp_rows = [
        [g, str(a["group_counts"][g]), f'{round(a["group_counts"][g]/total*100,1)}%']
        for g in g_order if a["group_counts"].get(g)
    ]
    flag_rows = [
        [k, str(v), f'{round(v/total*100,1)}%'] for k, v in flags.items()
    ]
    ic_tbl_rows = [
        [g, f'{gic[g]["ic"]}/{gic[g]["total"]}', f'{gic[g]["pct"]}%']
        for g in g_order if gic.get(g) and gic[g]["total"]
    ]
    curso_cota_rows = []
    for curso, gc in sorted(a["group_curso"].items()):
        tot_c = sum(gc.values())
        cota_c = gc.get("Cotas / Reserva de vagas", 0)
        short = "ECA" if "Controle" in curso else "SI"
        curso_cota_rows.append(
            [f"{short} — {curso}", str(cota_c),
             f'{round(cota_c/tot_c*100,1)}%' if tot_c else "0%"]
        )

    # --- fellowship (bolsa) cross ---
    gf = a.get("group_fellowship", {})
    bolsa_rows = [
        [g, str(gf[g]["paid"]), f'{gf[g]["pct_paid_total"]}%',
         f'{gf[g]["pct_paid_ic"]}%', str(gf[g]["volunteer"]), str(gf[g]["no_ic"])]
        for g in g_order if gf.get(g) and gf[g]["total"]
    ]
    # sponsor mix per group
    sponsor_rows = []
    for g in g_order:
        d = gf.get(g)
        if not d or not d["sponsors"]:
            continue
        mix = " · ".join(f"{k} {v}" for k, v in d["sponsors"].items())
        sponsor_rows.append([g, mix])

    tables_block = (
        '<div class="grid2" style="margin-bottom:22px;">'
        f'<div><div style="font-size:12px;font-weight:600;margin-bottom:8px;">'
        f'Grupo de ingresso</div>{_tbl(["grupo","n","%"], grp_rows)}'
        f'<div style="font-size:12px;font-weight:600;margin:14px 0 8px;">'
        f'Atributos de cota <span style="color:var(--sub);font-weight:400;">'
        f'(aluno pode somar &gt;1)</span></div>{_tbl(["atributo","n","%"], flag_rows)}</div>'
        f'<div><div style="font-size:12px;font-weight:600;margin-bottom:8px;">'
        f'Cota × pesquisa <span style="color:var(--sub);font-weight:400;">'
        f'(participação em IC)</span></div>{_tbl(["grupo","com IC","%"], ic_tbl_rows)}'
        f'<div style="font-size:12px;font-weight:600;margin:14px 0 8px;">'
        f'Cotas por curso</div>{_tbl(["curso","cotas","%"], curso_cota_rows)}'
        f'<div class="note" style="margin-top:10px;">Cotistas têm <strong>maior</strong> '
        f'participação em iniciação científica que ampla concorrência.</div></div>'
        '</div>'
        f'<div style="font-size:12px;font-weight:600;margin:6px 0 8px;">'
        f'Cota × bolsa <span style="color:var(--sub);font-weight:400;">'
        f'(bolsa = fomento pago; % bolsa/IC = bolsistas entre os que fazem IC)</span></div>'
        f'{_tbl(["grupo","c/ bolsa","% total","% bolsa/IC","voluntário","sem IC"], bolsa_rows)}'
        f'<div style="font-size:12px;font-weight:600;margin:14px 0 8px;">'
        f'Financiador por grupo <span style="color:var(--sub);font-weight:400;">'
        f'(bolsas pagas)</span></div>{_tbl(["grupo","agências"], sponsor_rows)}'
        f'<div class="note" style="margin-bottom:22px;">Cotistas têm ~2× mais chance de '
        f'bolsa paga que ampla concorrência; a <strong>Fapes</strong> concentra o fomento '
        f'a cotistas.</div>'
    )

    body = (
        tables_block +
        '<div class="grid2">'
        f'<div><div style="font-size:12px;font-weight:600;margin-bottom:10px;">'
        f'Grupo de ingresso</div>{group_bars}'
        f'<div class="note" style="margin-top:14px;">'
        f'<strong>{a["known"]}</strong> de <strong>{total}</strong> formandos com forma de '
        f'ingresso identificada · <strong>{a["unknown"]}</strong> sem informação. '
        f'Dados deduplicados por matrícula (1 registro por aluno).</div></div>'
        f'<div><div style="font-size:12px;font-weight:600;margin-bottom:10px;">'
        f'Atributos de cota <span style="color:var(--sub);font-weight:400;">'
        f'(aluno pode somar mais de um)</span></div>{flag_bars}</div>'
        '</div>'
        f'<div style="margin-top:20px;font-size:12px;font-weight:600;margin-bottom:10px;">'
        f'Categorias detalhadas (rótulo original)</div>{cat_bars}'
        f'<div style="margin-top:20px;"><div class="grid2">'
        f'<div><div style="font-size:12px;font-weight:600;margin-bottom:10px;">'
        f'Grupo × curso</div>{gc_rows}</div>'
        f'<div><div style="font-size:12px;font-weight:600;margin-bottom:10px;">'
        f'Participação em pesquisa por grupo</div>{ic_rows}'
        f'<div class="note" style="margin-top:10px;">% de cada grupo que participou de '
        f'iniciação científica registrada no SigPesq.</div></div>'
        f'</div></div>'
        f'<div class="note" style="margin-top:18px;">'
        f'<strong>Legenda:</strong> '
        f'<strong>PPI</strong> = Pretos, Pardos e Indígenas · '
        f'<strong>Renda</strong> ≤ 1,5 salário mínimo per capita · '
        f'<strong>CD</strong> = Pessoa com Deficiência · '
        f'<strong>OE</strong> = Outras Escolas (não-públicas) · '
        f'<strong>M*</strong> = modalidades ENEM da Lei de Cotas (12.711/2015) · '
        f'<strong>PS</strong> = processo seletivo próprio.</div>'
    )
    return section(
        "Forma de ingresso (cotas)",
        "distribuição dos formandos por forma de ingresso, categorias e cruzamento com pesquisa",
        body,
        border_color="var(--green)",
    )


def _sec_fellowship(s: dict) -> str:
    counts = s["fellowship_counts"]
    if not counts:
        return ""
    max_v = max(counts.values())
    COLORS = {
        "PIBIC": "var(--green)", "PIVIC": "var(--green2)",
        "PIBITI": "var(--amber)", "PIVITI": "var(--blue)",
        "PIBIC-JR": "var(--gray)", "PROPÓS": "var(--gray)",
    }
    rows = "".join(
        bar(k, v, max_v, COLORS.get(k, "var(--sub)"))
        for k, v in sorted(counts.items(), key=lambda x: -x[1])
    )
    return section(
        "Tipo de bolsa",
        f"{s['total_with_fellowship']} formandos com pelo menos uma bolsa registrada",
        rows,
    )


GCOL_AG = {
    "Ampla Concorrência": "var(--blue)",
    "Cotas / Reserva de vagas": "var(--green)",
    "Transferência": "var(--amber)",
}


def _sec_agencies(s: dict) -> str:
    AGENCY_COLORS = {
        "Fapes": "var(--amber)", "CNPq": "var(--blue)",
        "Ifes": "var(--green2)", "Voluntário": "var(--sub)",
    }
    order = ["Fapes", "CNPq", "Ifes", "Voluntário"]
    cards = "".join(
        mini_card_agency(
            ag,
            AGENCY_COLORS.get(ag, "var(--gray)"),
            s["sponsor_counts"].get(ag, 0),
            s["sponsor_fellowship_unique"].get(ag, {}),
            s["sponsor_investment"].get(ag, 0),
        )
        for ag in order
        if ag in s["sponsor_counts"]
    )

    paid = {k: v for k, v in s["sponsor_counts"].items() if k != "Voluntário"}
    max_paid = max(paid.values()) if paid else 1
    total_paid = sum(paid.values())
    bars_paid = "".join(
        bar(k, v, max_paid, AGENCY_COLORS.get(k, "var(--gray)"), "md")
        + f'<span style="font-size:10px;color:var(--sub);"> {round(v/total_paid*100)}%</span>'
        if total_paid else bar(k, v, max_paid, AGENCY_COLORS.get(k, "var(--gray)"), "md")
        for k, v in sorted(paid.items(), key=lambda x: -x[1])
    )

    # ---- investment indicators ----
    invest = {k: v for k, v in s["sponsor_investment"].items() if v > 0}
    total_invest = sum(invest.values())
    max_invest = max(invest.values()) if invest else 1

    def _fmt_brl(v: float) -> str:
        return f'R$ {v:,.0f}'.replace(",", "X").replace(".", ",").replace("X", ".")

    invest_bars = "".join(
        f'<div style="display:flex;align-items:center;gap:10px;margin-bottom:8px;">'
        f'<span style="font-size:12px;color:var(--text);width:90px;flex-shrink:0;">{ag}</span>'
        f'<div style="flex:1;height:14px;background:#e7efe9;border-radius:3px;overflow:hidden;">'
        f'<div style="width:{round(v/max_invest*100)}%;height:100%;background:{AGENCY_COLORS.get(ag,"var(--sub)")};">'
        f'</div></div>'
        f'<span style="font-size:12px;font-weight:600;color:{AGENCY_COLORS.get(ag,"var(--sub)")};width:90px;text-align:right;">'
        f'{_fmt_brl(v)}</span>'
        f'<span style="font-size:11px;color:var(--sub);width:80px;text-align:right;">'
        f'{_fmt_brl(v / s["sponsor_counts"].get(ag, 1))}/aluno</span>'
        f'</div>'
        for ag, v in sorted(invest.items(), key=lambda x: -x[1])
    )

    invest_block = (
        f'<div style="background:#eef5f0;border:1px solid var(--border);border-radius:6px;'
        f'padding:14px 16px;margin-bottom:14px;">'
        f'<div style="display:flex;justify-content:space-between;align-items:baseline;margin-bottom:12px;">'
        f'<div style="font-size:11px;color:var(--sub);text-transform:uppercase;letter-spacing:1px;">'
        f'Investimento em bolsas</div>'
        f'<div style="font-size:15px;font-weight:700;color:var(--green);">{_fmt_brl(total_invest)}'
        f'<span style="font-size:10px;color:var(--sub);font-weight:400;"> total acumulado</span></div>'
        f'</div>'
        f'{invest_bars}'
        f'<div style="font-size:10px;color:var(--sub);margin-top:8px;">'
        f'Valores estimados com base nas mensalidades registradas por bolsa no SigPesq. '
        f'Bolsas sem valor cadastrado não são contabilizadas.</div>'
        f'</div>'
    ) if invest_bars else ""

    # ---- agency × forma de ingresso (cota vs ampla) ----
    adm = s.get("admission") or {}
    gf = adm.get("group_fellowship", {})
    g_order = ["Ampla Concorrência", "Cotas / Reserva de vagas", "Transferência"]
    GSHORT = {"Ampla Concorrência": "Ampla", "Cotas / Reserva de vagas": "Cotas",
              "Transferência": "Transf."}
    # build agency → {group: count} from per-group sponsor breakdown
    ag_grp: dict[str, dict[str, int]] = defaultdict(lambda: defaultdict(int))
    for g in g_order:
        for ag, n in (gf.get(g, {}).get("sponsors", {}) or {}).items():
            ag_grp[ag][g] += n
    ag_order = [a for a in ["Fapes", "Ifes", "CNPq"] if a in ag_grp] + \
               [a for a in ag_grp if a not in ("Fapes", "Ifes", "CNPq")]

    matrix_block = ""
    if ag_grp:
        head = (
            '<tr><th style="text-align:left;padding:6px 10px;border-bottom:1px solid '
            'var(--border);color:var(--sub);font-size:11px;">agência</th>'
            + "".join(
                f'<th style="text-align:right;padding:6px 10px;border-bottom:1px solid '
                f'var(--border);color:var(--sub);font-size:11px;">{GSHORT[g]}</th>'
                for g in g_order)
            + '<th style="text-align:right;padding:6px 10px;border-bottom:1px solid '
              'var(--border);color:var(--sub);font-size:11px;">total</th>'
            + '<th style="text-align:right;padding:6px 10px;border-bottom:1px solid '
              'var(--border);color:var(--sub);font-size:11px;">% cotas</th></tr>'
        )
        body_rows = ""
        col_tot = {g: 0 for g in g_order}
        for ag in ag_order:
            row = ag_grp[ag]
            tot = sum(row.get(g, 0) for g in g_order)
            cot = row.get("Cotas / Reserva de vagas", 0)
            for g in g_order:
                col_tot[g] += row.get(g, 0)
            cells = "".join(
                f'<td style="text-align:right;padding:5px 10px;border-bottom:1px solid '
                f'var(--border);font-size:11px;color:{GCOL_AG.get(g)};">{row.get(g,0)}</td>'
                for g in g_order
            )
            body_rows += (
                f'<tr><td style="padding:5px 10px;border-bottom:1px solid var(--border);'
                f'font-size:11px;color:{AGENCY_COLORS.get(ag,"var(--gray)")};font-weight:600;">'
                f'{ag}</td>{cells}'
                f'<td style="text-align:right;padding:5px 10px;border-bottom:1px solid '
                f'var(--border);font-size:11px;font-weight:600;">{tot}</td>'
                f'<td style="text-align:right;padding:5px 10px;border-bottom:1px solid '
                f'var(--border);font-size:11px;color:var(--green);">'
                f'{round(cot/tot*100) if tot else 0}%</td></tr>'
            )
        gtot = sum(col_tot.values())
        tot_cells = "".join(
            f'<td style="text-align:right;padding:5px 10px;font-size:11px;font-weight:700;'
            f'color:{GCOL_AG.get(g)};">{col_tot[g]}</td>' for g in g_order
        )
        body_rows += (
            f'<tr><td style="padding:5px 10px;font-size:11px;font-weight:700;">TOTAL</td>'
            f'{tot_cells}'
            f'<td style="text-align:right;padding:5px 10px;font-size:11px;font-weight:700;">{gtot}</td>'
            f'<td style="text-align:right;padding:5px 10px;font-size:11px;font-weight:700;'
            f'color:var(--green);">'
            f'{round(col_tot["Cotas / Reserva de vagas"]/gtot*100) if gtot else 0}%</td></tr>'
        )
        matrix_block = (
            f'<div style="background:#f7faf8;border:1px solid var(--border);'
            f'border-radius:6px;padding:14px 16px;margin-bottom:14px;">'
            f'<div style="font-size:11px;color:var(--sub);margin-bottom:10px;'
            f'text-transform:uppercase;letter-spacing:1px;">'
            f'Agência × forma de ingresso (bolsistas pagos)</div>'
            f'<table style="width:100%;border-collapse:collapse;">'
            f'<thead>{head}</thead><tbody>{body_rows}</tbody></table>'
            f'<div style="font-size:10px;color:var(--sub);margin-top:8px;">'
            f'Cada célula = formandos com bolsa paga daquela agência, por forma de ingresso. '
            f'A <strong>Fapes</strong> financia mais cotistas que ampla concorrência.</div>'
            f'</div>'
        )

    note = (
        f'De {s["total_with_fellowship"]} formandos com bolsa formal, '
        f'FAPES é a maior financiadora ({s["sponsor_counts"].get("Fapes",0)} alunos), '
        f'seguida por IFES ({s["sponsor_counts"].get("Ifes",0)}) e '
        f'CNPq ({s["sponsor_counts"].get("CNPq",0)}). '
        f'{s["sponsor_counts"].get("Voluntário",0)} participaram voluntariamente '
        f'(PIVIC/PIVITI). {s["multi_bolsa"]} formandos acumularam bolsas de agências distintas.'
    )

    body = (
        f'<div class="grid4" style="margin-bottom:20px;">{cards}</div>'
        f'<div style="background:#f7faf8;border:1px solid var(--border);'
        f'border-radius:6px;padding:14px 16px;margin-bottom:14px;">'
        f'<div style="font-size:11px;color:var(--sub);margin-bottom:10px;'
        f'text-transform:uppercase;letter-spacing:1px;">Participação relativa — bolsa paga</div>'
        f'{bars_paid}</div>'
        f'{matrix_block}'
        f'{invest_block}'
        f'<div class="note"><strong>Interpretação:</strong> {note}</div>'
    )
    return section(
        "Agências de fomento",
        "Formandos beneficiados por bolsa formal, agrupados pela entidade financiadora",
        body,
    )


def _sec_curso_sponsor(s: dict) -> str:
    AGENCY_COLORS = {
        "Fapes": "var(--amber)", "CNPq": "var(--blue)",
        "Ifes": "var(--green2)", "Voluntário": "var(--sub)",
    }
    sponsor_order = ["Fapes", "Ifes", "CNPq", "Voluntário"]
    cols = ""
    for curso, total_c in sorted(s["curso_total"].items()):
        short = "Eng. Controle e Automação" if "Controle" in curso else "Sistemas de Informação"
        sp = s["curso_sponsor"].get(curso, {})
        max_v = max(sp.values()) if sp else 1
        rows = "".join(
            bar(ag, sp.get(ag, 0), max_v,
                AGENCY_COLORS.get(ag, "var(--sub)"), "md")
            for ag in sponsor_order
            if sp.get(ag, 0) > 0
        )
        cols += (
            f'<div style="background:#f7faf8;border:1px solid var(--border);'
            f'border-radius:6px;padding:14px;">'
            f'<div style="font-size:12px;font-weight:600;margin-bottom:12px;">{short}</div>'
            f'{rows}</div>'
        )
    body = f'<div class="grid2">{cols}</div>'
    return section(
        "Curso × agência de fomento",
        "Formandos únicos por curso beneficiados por cada agência",
        body,
    )


def _sec_projects_duration(s: dict) -> str:
    pd = s["proj_dist"]
    max_p = max(pd.values()) if pd else 1
    PROJ_COLORS = ["var(--green)", "var(--green2)", "var(--amber)", "var(--amber)", "var(--red)"]
    proj_bars = "".join(
        bar(
            f"{k} projeto{'s' if k > 1 else ''}",
            v, max_p,
            PROJ_COLORS[min(k - 1, len(PROJ_COLORS) - 1)],
            "md",
        )
        for k, v in sorted(pd.items())
    )
    proj_note = (
        f'<div class="note">{s["with_projects"]} formandos com projetos registrados. '
        f'{s["sem_registro"]} sem registro no SigPesq.</div>'
    )
    proj_col = (
        f'<div class="section" style="margin-bottom:0;">'
        f'<h2>Projetos por formando</h2>'
        f'<div class="sub">dentre formandos com projetos registrados</div>'
        f'{proj_bars}{proj_note}'
        + _decision_box(
            "<strong>O que é:</strong> quantos projetos de IC cada formando acumulou. "
            "<strong>Como decidir:</strong> mais projetos = mais maturação científica e maior chance de seguir para a pós. "
            "Predomínio de 1 projeto só indica espaço para incentivar renovação e continuidade."
        )
        + '</div>'
    )

    # duration
    db = s["dur_bins"]
    total_d = s["durations_n"]
    pct_7_12 = round(db.get("7-12m", 0) / total_d * 100) if total_d else 0
    pct_le6 = round(db.get("≤6m", 0) / total_d * 100) if total_d else 0
    pct_gt12 = round(db.get(">12m", 0) / total_d * 100) if total_d else 0

    dur_col = (
        f'<div class="section" style="margin-bottom:0;">'
        f'<h2>Duração de bolsa</h2>'
        f'<div class="sub">{total_d} registros com data de início/fim</div>'
        f'{bar("7–12 meses", db.get("7-12m",0), total_d, "var(--green)", "md")} '
        f'<span style="font-size:10px;color:var(--sub);">{pct_7_12}%</span>'
        f'{bar("≤ 6 meses", db.get("≤6m",0), total_d, "var(--amber)", "md")} '
        f'<span style="font-size:10px;color:var(--sub);">{pct_le6}%</span>'
        + (f'{bar("> 12 meses", db.get(">12m",0), total_d, "var(--blue)", "md")} '
           f'<span style="font-size:10px;color:var(--sub);">{pct_gt12}%</span>'
           if db.get(">12m", 0) > 0 else "")
        + f'<div class="grid2" style="margin-top:14px;gap:8px;">'
        f'<div style="background:#f7faf8;border:1px solid var(--border);border-radius:4px;'
        f'padding:10px;text-align:center;">'
        f'<div style="font-size:20px;font-weight:700;color:var(--green);">{s["dur_mean"]}d</div>'
        f'<div style="font-size:10px;color:var(--sub);">média (~{s["dur_mean"]//30} meses)</div></div>'
        f'<div style="background:#f7faf8;border:1px solid var(--border);border-radius:4px;'
        f'padding:10px;text-align:center;">'
        f'<div style="font-size:20px;font-weight:700;color:var(--green2);">{s["dur_median"]}d</div>'
        f'<div style="font-size:10px;color:var(--sub);">mediana (~{s["dur_median"]//30} meses)</div></div>'
        f'</div>'
        + _decision_box(
            "<strong>O que é:</strong> por quanto tempo a bolsa durou. "
            "<strong>Como decidir:</strong> bolsas de 7–12 meses dão tempo para resultado; "
            "concentração em ≤6 meses indica fragmentação que reduz a efetividade — prefira menos bolsas, porém mais longas."
        )
        + '</div>'
    )

    return f'<div class="grid2" style="margin-bottom:24px;">{proj_col}{dur_col}</div>'


def _sec_progressao(s: dict) -> str:
    pg = s["prog"]
    cards = "".join(
        f'<div class="prog-card">'
        f'<div class="pt">{label}</div>'
        f'<div class="pn" style="color:{color};">{pg.get(key,0)}</div>'
        f'<div class="ps">{desc}</div>'
        f'</div>'
        for key, label, color, desc in [
            ("vol→pago", "Voluntário → Pago", "var(--green)",
             "iniciaram voluntário, conquistaram bolsa paga"),
            ("multi-agencia", "Multi-agência", "var(--amber)",
             "participaram de projetos financiados por 2 ou mais agências distintas (ex: FAPES em um projeto, CNPq em outro)"),
            ("pago→vol", "Pago → Voluntário", "var(--sub)",
             "tinham bolsa paga, continuaram sem"),
        ]
    )
    note = (
        f'<div class="note" style="border-color:var(--green);margin-top:14px;">'
        f'<strong>Voluntário → Pago:</strong> {pg.get("vol→pago",0)} formandos iniciaram '
        f'a pesquisa sem remuneração (PIVIC voluntário) e, em projetos subsequentes, '
        f'conquistaram bolsa paga — indicador de progressão na carreira de pesquisa.'
        f'<br><br>'
        f'<strong>Multi-agência:</strong> {pg.get("multi-agencia",0)} formandos '
        f'participaram de projetos financiados por agências diferentes ao longo da graduação — '
        f'por exemplo, uma bolsa FAPES em iniciação científica e outra CNPq em projeto distinto. '
        f'Isso reflete diversificação de vínculos de pesquisa, não acúmulo simultâneo de bolsas.'
        f'</div>'
    )
    return section(
        "Progressão de bolsa",
        f"Trajetória de formandos com múltiplas bolsas — {s['multi_bolsa']} com 2+ bolsas",
        f'<div class="grid3">{cards}</div>{note}',
    )


def _sec_orientadores(s: dict) -> str:
    sd = s["sup_dist"]
    max_sd = max(sd.values()) if sd else 1
    dist_bars = "".join(
        bar(
            f"{k} orientador{'es' if k > 1 else ''}",
            v, max_sd, "var(--green)" if k == 1 else ("var(--amber)" if k == 2 else "var(--red)"),
            "md",
        )
        for k, v in sorted(sd.items())
    )
    multi = sum(v for k, v in sd.items() if k > 1)
    dist_note = (
        f'<div style="margin-top:10px;font-size:11px;color:var(--sub);">'
        f'{multi} formandos ({round(multi/sum(sd.values())*100) if sd else 0}%) '
        f'atuaram com múltiplos orientadores.</div>'
    )

    top_rows = "".join(
        f'<div class="list-row"><span style="color:var(--text);">{name}</span>'
        f'<span style="color:var(--green);font-weight:600;">{count}</span></div>'
        for name, count in s["top_sups"][:8]
    )

    body = (
        f'<div class="grid2">'
        f'<div>{dist_bars}{dist_note}</div>'
        f'<div><div style="font-size:11px;color:var(--sub);margin-bottom:8px;'
        f'text-transform:uppercase;letter-spacing:1px;">Top orientadores</div>'
        f'{top_rows}</div>'
        f'</div>'
    )
    return section(
        "Orientadores por formando",
        "Formandos que tiveram mais de um orientador ao longo do curso",
        body,
    )


def _sec_rg(s: dict) -> str:
    stats = s.get("rg_top_stats") or []
    if not stats:
        return ""
    max_v = max((g["total"] for g in stats), default=1) or 1
    rows = ""
    for g in stats:
        rows += (
            f'<div style="margin-bottom:14px;">'
            f'<div style="font-size:12px;font-weight:600;margin-bottom:6px;">'
            f'{g["name"][:50]}</div>'
            f'{bar("total", g["total"], max_v, "var(--green2)", "sm")}'
            f'{bar("bolsas pagas", g["paid"], max_v, "var(--amber)", "sm")}'
            f'{bar("voluntários", g["vol"], max_v, "var(--gray)", "sm")}'
            f'{bar("cotistas", g.get("cota", 0), max_v, "var(--green)", "sm")}'
            f'{bar("não-cotistas", g.get("naocota", 0), max_v, "var(--blue)", "sm")}'
            f'</div>'
        )
    note = (
        '<div class="note"><strong>total</strong> = formandos vinculados ao grupo · '
        '<strong>bolsas pagas</strong> = com fomento financiado · '
        '<strong>voluntários</strong> = sem bolsa paga (PIVIC/PIVITI ou só vínculo) · '
        '<strong>cotistas</strong> = ingresso por reserva de vagas · '
        '<strong>não-cotistas</strong> = ampla concorrência / transferência.</div>'
    )
    return section(
        "Grupos de pesquisa",
        "formandos vinculados a grupos — total, bolsas pagas e voluntários",
        rows + note,
    )


def _sec_ka(s: dict) -> str:
    ka = s["ka_top"]
    if not ka:
        return ""
    max_v = ka[0][1] if ka else 1
    COLORS = {0: "var(--green)", 1: "var(--green)", 2: "var(--green2)",
              3: "var(--amber)", 4: "var(--amber)", 5: "var(--blue)"}
    rows = "".join(
        bar(name[:40], count, max_v, COLORS.get(i, "var(--sub)"))
        for i, (name, count) in enumerate(ka)
    )
    note = (
        '<div class="note">KAs derivadas dos grupos de pesquisa vinculados — '
        'não dos currículos individuais. Formandos são cadastrados como estudantes '
        'e não possuem currículo Lattes indexado no SigPesq.</div>'
    )
    return section(
        "Áreas de conhecimento",
        "Derivado dos grupos de pesquisa aos quais formandos estiveram vinculados",
        rows + note,
    )


def _sec_artigos() -> str:
    body = (
        f'<div style="display:grid;grid-template-columns:auto 1fr;gap:16px;align-items:start;">'
        f'<div style="background:#f7faf8;border:1px solid var(--border);border-radius:8px;'
        f'padding:20px 28px;text-align:center;">'
        f'<div style="font-size:40px;font-weight:700;color:var(--sub);">0</div>'
        f'<div style="font-size:11px;color:var(--sub);">artigos</div>'
        f'</div>'
        f'<div class="note" style="margin-top:0;">'
        f'<strong>Por que zero?</strong><br>'
        f'O campo <code>articles</code> é populado via extração do currículo Lattes — '
        f'disponível apenas para docentes e pesquisadores com Lattes vinculado. '
        f'Formandos são cadastrados como estudantes e não têm Lattes indexado no SigPesq. '
        f'Co-autorias existem mas não são acessíveis sem cruzar DOIs via API externa.<br><br>'
        f'<span style="color:var(--amber);">→ Para medir produção discente, indexar os '
        f'Lattes dos estudantes ou cruzar DOIs via API.</span>'
        f'</div></div>'
    )
    return section("Produção científica", "Artigos registrados para formandos no sistema", body)


def _sec_ic_timing(s: dict) -> str:
    t = s.get("ic_timing", {})
    if not t or not t.get("period_dist"):
        return ""

    avg = t["avg_semesters"]
    n = t["n_with_entry"]
    pd = t["period_dist"]
    yd = t["year_dist"]

    max_pd = max(pd.values()) if pd else 1
    max_yd = max(yd.values()) if yd else 1

    # -- KPI cards --
    early = sum(v for k, v in pd.items() if k <= 2)
    mid   = sum(v for k, v in pd.items() if 3 <= k <= 5)
    late  = sum(v for k, v in pd.items() if k >= 6)

    kpi_cards = (
        f'<div style="display:grid;grid-template-columns:repeat(4,1fr);gap:12px;margin-bottom:20px;">'

        f'<div style="background:#f7faf8;border:1px solid var(--border);border-radius:8px;padding:14px;text-align:center;">'
        f'<div style="font-size:28px;font-weight:700;color:var(--green);">{avg}</div>'
        f'<div style="font-size:10px;color:var(--sub);margin-top:4px;">semestres médios<br>até 1ª IC</div>'
        f'</div>'

        f'<div style="background:#f7faf8;border:1px solid var(--green);border-radius:8px;padding:14px;text-align:center;">'
        f'<div style="font-size:28px;font-weight:700;color:var(--green);">{early}</div>'
        f'<div style="font-size:10px;color:var(--sub);margin-top:4px;">iniciaram cedo<br>(1º–2º semestre)</div>'
        f'</div>'

        f'<div style="background:#f7faf8;border:1px solid var(--amber);border-radius:8px;padding:14px;text-align:center;">'
        f'<div style="font-size:28px;font-weight:700;color:var(--amber);">{mid}</div>'
        f'<div style="font-size:10px;color:var(--sub);margin-top:4px;">período intermediário<br>(3º–5º semestre)</div>'
        f'</div>'

        f'<div style="background:#f7faf8;border:1px solid var(--sub);border-radius:8px;padding:14px;text-align:center;">'
        f'<div style="font-size:28px;font-weight:700;color:var(--sub);">{late}</div>'
        f'<div style="font-size:10px;color:var(--sub);margin-top:4px;">iniciaram tarde<br>(6º semestre ou mais)</div>'
        f'</div>'

        f'</div>'
    )

    # -- Period distribution bars --
    COURSE_LEN = {"Sistemas de Informação": 8, "Engenharia de Controle e Automação": 10}
    default_len = 8

    def _period_color(k: int) -> str:
        if k <= 2:
            return "var(--green)"
        if k <= 5:
            return "var(--amber)"
        return "var(--sub)"

    period_bars = "".join(
        f'<div style="display:flex;align-items:center;gap:8px;margin-bottom:6px;">'
        f'<span style="font-size:11px;color:var(--text);width:70px;flex-shrink:0;">{k}º semestre</span>'
        f'<div style="flex:1;height:14px;background:#e7efe9;border-radius:3px;overflow:hidden;">'
        f'<div style="width:{round(v/max_pd*100)}%;height:100%;background:{_period_color(k)};"></div>'
        f'</div>'
        f'<span style="font-size:11px;font-weight:600;color:{_period_color(k)};width:24px;text-align:right;">{v}</span>'
        f'</div>'
        for k, v in sorted(pd.items())
    )

    # -- Year distribution bars --
    year_bars = "".join(
        f'<div style="display:flex;align-items:center;gap:8px;margin-bottom:6px;">'
        f'<span style="font-size:11px;color:var(--text);width:40px;flex-shrink:0;">{yr}</span>'
        f'<div style="flex:1;height:12px;background:#e7efe9;border-radius:3px;overflow:hidden;">'
        f'<div style="width:{round(v/max_yd*100)}%;height:100%;background:var(--blue);"></div>'
        f'</div>'
        f'<span style="font-size:11px;color:var(--blue);width:20px;text-align:right;">{v}</span>'
        f'</div>'
        for yr, v in sorted(yd.items())
    )

    def _subhead(title: str, sub: str) -> str:
        return (
            f'<div style="font-size:12px;font-weight:700;color:var(--text);'
            f'text-transform:uppercase;letter-spacing:.05em;margin-bottom:4px;">{title}</div>'
            f'<div style="font-size:11px;color:var(--sub);margin-bottom:14px;">{sub}</div>'
        )

    charts = (
        _subhead("Visão geral", "Todos os cursos somados")
        + f'<div style="display:grid;grid-template-columns:1fr 1fr;gap:24px;margin-bottom:16px;">'

        f'<div><div style="font-size:11px;color:var(--sub);text-transform:uppercase;'
        f'letter-spacing:.05em;margin-bottom:10px;">Semestre do curso na 1ª IC</div>'
        f'{period_bars}</div>'

        f'<div><div style="font-size:11px;color:var(--sub);text-transform:uppercase;'
        f'letter-spacing:.05em;margin-bottom:10px;">Ano de início (calendário)</div>'
        f'{year_bars}</div>'

        f'</div>'
    )

    # -- subseção por curso (a partir dos records, que já trazem o curso) --
    by_curso_recs: dict[str, list[int]] = defaultdict(list)
    for r in t.get("records", []):
        by_curso_recs[r.get("curso") or "—"].append(r.get("semesters_after", 0))

    def _curso_timing_block(curso: str, vals: list[int]) -> str:
        if not vals:
            return ""
        short = ("ECA — Eng. Controle e Automação" if "Controle" in curso
                 else "SI — Sistemas de Informação" if "Sistemas" in curso else curso)
        nn = len(vals)
        avg_c = round(sum(vals) / nn, 1)
        e = sum(1 for v in vals if v <= 2)
        m = sum(1 for v in vals if 3 <= v <= 5)
        lt = sum(1 for v in vals if v >= 6)
        dist_c = Counter(vals)
        maxd = max(dist_c.values()) if dist_c else 1
        bars_c = "".join(
            f'<div style="display:flex;align-items:center;gap:8px;margin-bottom:5px;">'
            f'<span style="font-size:10px;color:var(--text);width:64px;flex-shrink:0;">{k}º sem</span>'
            f'<div style="flex:1;height:12px;background:#e7efe9;border-radius:3px;overflow:hidden;">'
            f'<div style="width:{round(v/maxd*100)}%;height:100%;background:{_period_color(k)};"></div></div>'
            f'<span style="font-size:10px;font-weight:600;color:{_period_color(k)};width:20px;text-align:right;">{v}</span>'
            f'</div>'
            for k, v in sorted(dist_c.items())
        )
        return (
            f'<div style="background:#f7faf8;border:1px solid var(--border);border-radius:8px;padding:14px;">'
            f'<div style="font-size:12px;font-weight:600;margin-bottom:6px;">{short}</div>'
            f'<div style="font-size:11px;color:var(--sub);margin-bottom:10px;">'
            f'média <strong style="color:var(--green);">{avg_c}</strong> sem até 1ª IC · n={nn} · '
            f'<span style="color:var(--green);">{e} cedo</span> · '
            f'<span style="color:var(--amber);">{m} interm.</span> · '
            f'<span style="color:var(--sub);">{lt} tarde</span></div>'
            f'{bars_c}</div>'
        )

    curso_blocks_t = "".join(
        _curso_timing_block(c, by_curso_recs[c]) for c in sorted(by_curso_recs)
    )
    curso_subsec = (
        f'<div style="border-top:1px solid var(--border);margin:18px 0 14px;padding-top:16px;">'
        + _subhead(
            "Por curso",
            "Mesmo recorte separado por currículo (SI 8 sem · ECA 12 sem) — "
            "entrada tardia pesa mais no curso mais curto, onde sobra menos tempo de pesquisa."
          )
        + f'<div class="grid2">{curso_blocks_t}</div></div>'
    )

    note = (
        f'<div class="note" style="margin-top:0;">'
        f'<strong>Metodologia:</strong> período calculado pela diferença entre semestre de entrada '
        f'(extraído da matrícula — ex: <code>20181BSI…</code> → 2018/1) e data de início do '
        f'primeiro projeto SigPesq. Base: {n} de {s["with_research"]} formandos com pesquisa '
        f'e matrícula interpretável. '
        f'<span style="color:var(--green);">Verde = 1º–2º sem.</span> · '
        f'<span style="color:var(--amber);">Âmbar = 3º–5º sem.</span> · '
        f'<span style="color:var(--sub);">Cinza = 6º sem.+</span>'
        f'</div>'
    )

    return section(
        "Quando os formandos entraram na IC",
        f"Tempo entre matrícula e 1ª iniciação científica — média: {avg} semestres · base: {n} formandos",
        kpi_cards + charts + curso_subsec + note,
    )


def _sec_bolsistas(s: dict) -> str:
    bc = s.get("bolsistas_cross")
    if not bc or not bc.get("formaram"):
        return ""

    def _brl(v: float) -> str:
        return f'R$ {v:,.0f}'.replace(",", "X").replace(".", ",").replace("X", ".")

    def _kpi(num, lab, sub, col="var(--green)"):
        return (
            f'<div style="background:#f7faf8;border:1px solid var(--border);border-radius:8px;'
            f'padding:16px;text-align:center;">'
            f'<div style="font-size:30px;font-weight:800;color:{col};">{num}</div>'
            f'<div style="font-size:11px;color:var(--text);margin-top:4px;">{lab}</div>'
            f'<div style="font-size:10px;color:var(--sub);margin-top:2px;">{sub}</div></div>'
        )

    pct = round(bc["formaram"] / bc["total_bolsistas"] * 100, 1) if bc["total_bolsistas"] else 0
    novos = bc.get("novos_pesquisa", 0)
    kpis = (
        '<div class="grid4" style="margin-bottom:18px;">'
        + _kpi(bc["formaram"], "bolsistas FAPES formados", f"de {bc['total_bolsistas']} bolsistas", "var(--green)")
        + _kpi(bc["com_ic_sigpesq"], "já tinham IC no SigPesq", "dos que formaram", "var(--green2)")
        + _kpi(f"+{novos}", "novos em pesquisa", "bolsa é a única evidência", "var(--amber)")
        + _kpi(_brl(bc["valor_alocado_total"]), "alocado a formados", f"pago {_brl(bc['valor_pago_total'])}", "var(--blue)")
        + '</div>'
    )

    # ---- tipos de bolsa FAPES (com valores diferentes) ----
    por_bolsa = bc.get("por_bolsa", [])
    bolsa_block = ""
    if por_bolsa:
        bmax = max((x["formados"] for x in por_bolsa), default=1)
        brows = ""
        for x in por_bolsa:
            faixa = (_brl(x["valor_min"]) if x["valor_min"] == x["valor_max"]
                     else f'{_brl(x["valor_min"])} – {_brl(x["valor_max"])}')
            w = x["formados"] / bmax * 100
            brows += (
                f'<tr><td style="padding:6px 10px;border-bottom:1px solid var(--border);font-size:11px;">'
                f'<strong>{x["sigla"]}</strong> <span style="color:var(--sub);">{x["nome"]}</span></td>'
                f'<td style="padding:6px 10px;border-bottom:1px solid var(--border);font-size:11px;text-align:center;">'
                f'{x["formados"]}</td>'
                f'<td style="padding:6px 10px;border-bottom:1px solid var(--border);font-size:11px;text-align:right;">'
                f'{faixa}<span style="color:var(--sub);font-size:10px;">/mês</span></td></tr>'
            )
        bolsa_block = (
            '<div style="font-size:12px;font-weight:600;margin:6px 0 10px;">'
            'Tipos de bolsa FAPES dos formados <span style="color:var(--sub);font-weight:400;">'
            '(valores mensais distintos por modalidade)</span></div>'
            '<table style="width:100%;border-collapse:collapse;margin-bottom:16px;">'
            '<thead><tr>'
            '<th style="text-align:left;padding:6px 10px;border-bottom:1px solid var(--border);color:var(--sub);font-size:11px;">modalidade</th>'
            '<th style="padding:6px 10px;border-bottom:1px solid var(--border);color:var(--sub);font-size:11px;">formados</th>'
            '<th style="text-align:right;padding:6px 10px;border-bottom:1px solid var(--border);color:var(--sub);font-size:11px;">valor/mês</th>'
            f'</tr></thead><tbody>{brows}</tbody></table>'
        )

    # by curso / ingresso mini-bars
    pc = bc["por_curso"]; pi = bc["por_ingresso"]
    def _mini(d: dict, colors: dict) -> str:
        mx = max(d.values()) if d else 1
        return "".join(
            bar(k if len(k) < 24 else k[:22], v, mx, colors.get(k, "var(--sub)"), "md")
            for k, v in sorted(d.items(), key=lambda x: -x[1])
        )
    GCOL = {"Cotas / Reserva de vagas": "var(--green)", "Ampla Concorrência": "var(--blue)",
            "Transferência": "var(--amber)"}
    splits = (
        '<div class="grid2" style="margin-bottom:16px;">'
        f'<div><div style="font-size:12px;font-weight:600;margin-bottom:10px;">Por curso</div>'
        f'{_mini(pc, {})}</div>'
        f'<div><div style="font-size:12px;font-weight:600;margin-bottom:10px;">Por forma de ingresso</div>'
        f'{_mini(pi, GCOL)}</div></div>'
    )

    # table of graduated bolsistas
    rows = ""
    for m in bc["registros"]:
        short = "ECA" if "Controle" in m["curso"] else "SI"
        ic = '✓' if m["com_ic_sigpesq"] else '—'
        rows += (
            f'<tr><td style="padding:6px 10px;border-bottom:1px solid var(--border);font-size:11px;">'
            f'{m["nome"]}</td>'
            f'<td style="padding:6px 10px;border-bottom:1px solid var(--border);font-size:11px;text-align:center;">{short}</td>'
            f'<td style="padding:6px 10px;border-bottom:1px solid var(--border);font-size:11px;text-align:center;">{ic}</td>'
            f'<td style="padding:6px 10px;border-bottom:1px solid var(--border);font-size:11px;text-align:right;color:var(--sub);">'
            f'{m["total_projetos"]}p · {m["total_alocacoes"]}a</td></tr>'
        )
    table = (
        '<div style="font-size:12px;font-weight:600;margin:6px 0 10px;">'
        'Bolsistas que se formaram</div>'
        '<table style="width:100%;border-collapse:collapse;">'
        '<thead><tr>'
        '<th style="text-align:left;padding:6px 10px;border-bottom:1px solid var(--border);'
        'color:var(--sub);font-size:11px;">bolsista</th>'
        '<th style="padding:6px 10px;border-bottom:1px solid var(--border);color:var(--sub);font-size:11px;">curso</th>'
        '<th style="padding:6px 10px;border-bottom:1px solid var(--border);color:var(--sub);font-size:11px;">IC</th>'
        '<th style="text-align:right;padding:6px 10px;border-bottom:1px solid var(--border);color:var(--sub);font-size:11px;">proj·aloc</th>'
        f'</tr></thead><tbody>{rows}</tbody></table>'
    )

    note = (
        f'<div class="note" style="margin-top:14px;">Cruzamento dos <strong>'
        f'{bc["total_bolsistas"]}</strong> bolsistas <strong>FAPES</strong> do recorte IFES Serra '
        f'com os formandos (match por nome sem acento). <strong>{bc["formaram"]}</strong> já '
        f'concluíram a graduação; destes, <strong>{novos}</strong> não tinham IC no SigPesq e '
        f'<strong>passam a contar como participação em pesquisa</strong> (a bolsa é a evidência). '
        f'Os demais ({bc["nao_formandos"]}) são docentes, coordenadores, pós-graduação ou ainda '
        f'não formados. Cada modalidade FAPES tem valor mensal próprio. '
        f'Valores são <strong>alocados</strong>; pagamento efetivo consta zerado na fonte.</div>'
    )
    return section(
        "Bolsistas FAPES formados",
        "bolsistas de pesquisa FAPES do IFES Serra que já concluíram a graduação",
        kpis + bolsa_block + splits + table + note,
        border_color="var(--blue)",
    )


def _sec_ic_recovery(s: dict) -> str:
    r = s.get("ic_recovery")
    if not r:
        return ""
    before = r["sigpesq_ic"]
    after = r["combined_ic_n"]
    gained = after - before
    total = s["total"]

    def _kpi(num, lab, sub, col="var(--green)"):
        return (
            f'<div style="background:#f7faf8;border:1px solid var(--border);border-radius:8px;'
            f'padding:16px;text-align:center;">'
            f'<div style="font-size:30px;font-weight:700;color:{col};">{num}</div>'
            f'<div style="font-size:11px;color:var(--text);margin-top:4px;">{lab}</div>'
            f'<div style="font-size:10px;color:var(--sub);margin-top:2px;">{sub}</div></div>'
        )

    kpis = (
        f'<div class="grid4" style="margin-bottom:18px;">'
        + _kpi(before, "IC no SigPesq", f'{r["coverage_before_pct"]}% dos formandos', "var(--blue)")
        + _kpi(f'+{r["recovered_fuzzy_n"] + r["recovered_exact_n"]}', "recuperados via Lattes",
               "fora do SigPesq", "var(--amber)")
        + _kpi(after, "IC combinado", f'{r["coverage_after_pct"]}% dos formandos', "var(--green)")
        + _kpi(r["sem_ic_after"], "ainda sem IC", "nenhuma fonte", "var(--red)")
        + '</div>'
    )

    def _list(names, color):
        if not names:
            return '<div style="font-size:11px;color:var(--sub);">nenhum</div>'
        return "".join(
            f'<span style="display:inline-block;background:#f7faf8;border:1px solid {color};'
            f'border-radius:5px;padding:3px 9px;margin:0 5px 6px 0;font-size:11px;">{n}</span>'
            for n in names
        )

    exact = r["recovered_exact"]
    fuzzy = r["recovered_fuzzy"]
    lists = (
        (f'<div style="margin-bottom:14px;"><div style="font-size:12px;font-weight:600;'
         f'margin-bottom:8px;color:var(--green);">Match exato — alta confiança '
         f'({r["recovered_exact_n"]})</div>{_list(exact, "var(--green)")}</div>'
         if exact else "")
        + (f'<div><div style="font-size:12px;font-weight:600;margin-bottom:8px;color:var(--amber);">'
           f'Match aproximado (1º + último nome) — revisar manualmente '
           f'({r["recovered_fuzzy_n"]})</div>{_list(fuzzy, "var(--amber)")}</div>'
           if fuzzy else "")
    )

    note = (
        f'<div class="note" style="margin-top:14px;">Cruzamento dos <strong>'
        f'{r["sem_ic_sigpesq"]}</strong> formandos sem IC no SigPesq com a iniciação científica '
        f'registrada nos currículos Lattes. O casamento usa nome <strong>sem acento</strong> '
        f'(corrige grafias divergentes) — só esse ajuste elevou a contagem SigPesq de IC. '
        f'O match aproximado é candidato a confirmação manual, pois pode haver homônimos.</div>'
    )
    return section(
        "Recuperação de alunos com IC via Lattes",
        f"alunos com iniciação científica não encontrados no SigPesq, recuperados pelo Lattes",
        kpis + lists + note,
        border_color="var(--amber)",
    )


def _sec_lattes_cross(s: dict) -> str:
    lc = s.get("lattes_cross")
    if not lc:
        return ""

    sp_n = lc["sigpesq_n"]
    union_n = lc["union_n"]
    avg_sp = lc["avg_sigpesq"]
    avg_un = lc["avg_union"]
    delta = lc["delta_avg"]
    delta_pct = lc["delta_pct"]
    enriched = lc["enriched_count"]
    new_names = lc["new_via_lattes"]
    ic_form = lc["lattes_ic_formandos"]
    tcc_form = lc["lattes_tcc_formandos"]

    delta_color = "var(--green)" if delta >= 0 else "var(--red)"

    # Comparison table
    table_rows = f"""
      <tr>
        <td>Formandos com pesquisa (SigPesq)</td>
        <td style="text-align:right;font-weight:600;">{sp_n}</td>
        <td style="text-align:right;font-weight:600;">{lc['coverage_sigpesq_pct']}%</td>
      </tr>
      <tr>
        <td>Formandos com pesquisa (SigPesq + Lattes)</td>
        <td style="text-align:right;font-weight:600;">{union_n}</td>
        <td style="text-align:right;font-weight:600;">{lc['coverage_union_pct']}%</td>
      </tr>
      <tr>
        <td>Formandos IC via Lattes</td>
        <td style="text-align:right;">{ic_form}</td>
        <td></td>
      </tr>
      <tr>
        <td>Formandos TCC via Lattes</td>
        <td style="text-align:right;">{tcc_form}</td>
        <td></td>
      </tr>
      <tr>
        <td>Média projetos/formando — SigPesq</td>
        <td style="text-align:right;font-weight:600;">{avg_sp}</td>
        <td></td>
      </tr>
      <tr>
        <td>Média projetos/formando — União</td>
        <td style="text-align:right;font-weight:600;color:{delta_color};">{avg_un}</td>
        <td style="text-align:right;color:{delta_color};">+{delta_pct}%</td>
      </tr>
      <tr>
        <td>Formandos enriquecidos (Lattes adicionou TCC)</td>
        <td style="text-align:right;">{enriched}</td>
        <td></td>
      </tr>
    """

    comparison_table = (
        f'<table style="width:100%;border-collapse:collapse;font-size:13px;">'
        f'<thead><tr>'
        f'<th style="text-align:left;padding:8px 4px;border-bottom:1px solid var(--border);">Métrica</th>'
        f'<th style="text-align:right;padding:8px 4px;border-bottom:1px solid var(--border);">Valor</th>'
        f'<th style="text-align:right;padding:8px 4px;border-bottom:1px solid var(--border);">%</th>'
        f'</tr></thead><tbody>{table_rows}</tbody></table>'
    )

    # Distribution bars: SigPesq vs Union
    all_counts = sorted(set(lc["sp_dist"]) | set(lc["union_dist"]))
    max_val = max(
        max(lc["sp_dist"].values(), default=1),
        max(lc["union_dist"].values(), default=1),
    )
    dist_rows = ""
    for c in all_counts:
        sp_v = lc["sp_dist"].get(c, 0)
        un_v = lc["union_dist"].get(c, 0)
        sp_w = round(sp_v / max_val * 120)
        un_w = round(un_v / max_val * 120)
        dist_rows += (
            f'<tr style="border-bottom:1px solid var(--border);">'
            f'<td style="padding:5px 8px;font-weight:600;">{c}×</td>'
            f'<td style="padding:5px 8px;">'
            f'<div style="display:flex;align-items:center;gap:6px;">'
            f'<div style="width:{sp_w}px;height:10px;background:var(--sub);border-radius:2px;"></div>'
            f'<span style="font-size:12px;">{sp_v}</span></div></td>'
            f'<td style="padding:5px 8px;">'
            f'<div style="display:flex;align-items:center;gap:6px;">'
            f'<div style="width:{un_w}px;height:10px;background:var(--green);border-radius:2px;"></div>'
            f'<span style="font-size:12px;">{un_v}</span></div></td>'
            f'</tr>'
        )

    dist_table = (
        f'<table style="width:100%;border-collapse:collapse;font-size:13px;">'
        f'<thead><tr>'
        f'<th style="text-align:left;padding:6px 8px;border-bottom:1px solid var(--border);">Projetos</th>'
        f'<th style="padding:6px 8px;border-bottom:1px solid var(--border);">SigPesq</th>'
        f'<th style="padding:6px 8px;border-bottom:1px solid var(--border);">SigPesq + Lattes</th>'
        f'</tr></thead><tbody>{dist_rows}</tbody></table>'
    )

    # New formandos discovered via Lattes
    new_list = ""
    if new_names:
        items = "".join(f'<li style="margin:2px 0;">{n}</li>' for n in new_names[:20])
        suffix = f'<li style="color:var(--sub);">... e mais {len(new_names)-20}</li>' if len(new_names) > 20 else ""
        new_list = (
            f'<div style="margin-top:20px;">'
            f'<div style="font-size:12px;font-weight:600;color:var(--sub);text-transform:uppercase;'
            f'letter-spacing:.05em;margin-bottom:8px;">Formandos novos via Lattes ({len(new_names)})</div>'
            f'<ul style="margin:0;padding-left:20px;font-size:13px;columns:2;gap:24px;">'
            f'{items}{suffix}</ul></div>'
        )

    body = (
        f'<div style="display:grid;grid-template-columns:1fr 1fr;gap:24px;align-items:start;">'
        f'<div>{comparison_table}</div>'
        f'<div>{dist_table}</div>'
        f'</div>'
        f'{new_list}'
        f'<div class="note" style="margin-top:20px;">'
        f'<strong>Metodologia:</strong> SigPesq registra bolsas IC/IT com orientador. '
        f'Lattes captura adicionalmente TCCs (não registrados no SigPesq) e IC informais. '
        f'IC duplicada entre fontes toma o maior contagem individual (max). '
        f'TCC é aditivo — SigPesq não rastreia. '
        f'Impacto: média sobe de <strong>{avg_sp}</strong> → <strong style="color:{delta_color};">{avg_un}</strong> proj/formando '
        f'(<strong style="color:{delta_color};">+{delta_pct}%</strong>).'
        f'</div>'
    )

    return section(
        "Cruzamento Lattes × SigPesq",
        "Impacto do Lattes na cobertura de projetos por formando",
        body,
    )


def _sec_graduation_time(s: dict) -> str:
    gt = s.get("graduation_time", {})
    overall = gt.get("overall", {})
    by_curso = gt.get("by_curso", {})
    cats = gt.get("categories", {})
    if not overall:
        return ""

    mean_sem = overall["mean"]
    median_sem = overall["median"]
    n = overall["n"]
    total = s["total"]

    # Métrica comparável entre cursos = ATRASO (semestres além do previsto),
    # pois SI tem 8 semestres de currículo e ECA tem 12. A média absoluta
    # (10,3 sem) mistura as duas réguas e não significa nada isolada.
    od = gt.get("overall_delay", {})
    delay_mean = od.get("mean", 0) or 0
    delay_median = od.get("median", 0) or 0
    _dd = {int(k): v for k, v in od.get("dist", {}).items()}
    n_on  = sum(v for k, v in _dd.items() if k <= 0)   # no prazo / adiantado
    n_mod = sum(v for k, v in _dd.items() if 1 <= k <= 5)
    n_long = sum(v for k, v in _dd.items() if k >= 6)
    pct_on = round(n_on / n * 100, 1) if n else 0
    expected_map = gt.get("expected", {})

    _dcolor = ("var(--green)" if delay_mean <= 1
               else "var(--amber)" if delay_mean <= 3 else "var(--red)")
    _dval = f"{delay_mean:+.1f}".replace("+0.0", "0,0").replace(".", ",")
    _dlabel = ("no prazo médio" if delay_mean <= 0
               else f"{_dval} sem além do previsto")

    kpi = (
        f'<div style="display:grid;grid-template-columns:repeat(3,1fr);gap:12px;margin-bottom:8px;">'
        f'<div style="background:#f7faf8;border:1px solid {_dcolor};border-radius:8px;padding:14px;text-align:center;">'
        f'<div style="font-size:28px;font-weight:700;color:{_dcolor};">{_dval}</div>'
        f'<div style="font-size:10px;color:var(--sub);margin-top:4px;">atraso médio (sem)<br>{_dlabel}</div>'
        f'</div>'
        f'<div style="background:#f7faf8;border:1px solid var(--border);border-radius:8px;padding:14px;text-align:center;">'
        f'<div style="font-size:28px;font-weight:700;color:var(--green);">{pct_on}%</div>'
        f'<div style="font-size:10px;color:var(--sub);margin-top:4px;">no prazo ou adiantado<br>(≤ duração prevista do curso)</div>'
        f'</div>'
        f'<div style="background:#f7faf8;border:1px solid var(--border);border-radius:8px;padding:14px;text-align:center;">'
        f'<div style="font-size:28px;font-weight:700;color:var(--sub);">{n}</div>'
        f'<div style="font-size:10px;color:var(--sub);margin-top:4px;">formandos<br>com matrícula interpretável</div>'
        f'</div>'
        f'</div>'
        f'<div class="note" style="margin-bottom:16px;">'
        f'Currículo previsto: <strong>SI {expected_map.get("Sistemas de Informação", 8)} sem</strong> · '
        f'<strong>ECA {expected_map.get("Engenharia de Controle e Automação", 12)} sem</strong>. '
        f'O <strong>atraso</strong> (semestres além do previsto de cada curso) é a métrica comparável — '
        f'a média absoluta de {str(mean_sem).replace(".", ",")} sem mistura as duas réguas e induz a erro.'
        f'</div>'
    )

    # Category breakdown row — relativo ao currículo de CADA curso (via atraso)
    cat_row = (
        f'<div style="display:grid;grid-template-columns:repeat(3,1fr);gap:10px;margin-bottom:20px;">'

        f'<div style="background:#eef5f0;border:1px solid var(--green);border-radius:6px;padding:12px;text-align:center;">'
        f'<div style="font-size:22px;font-weight:700;color:var(--green);">{n_on}</div>'
        f'<div style="font-size:10px;color:var(--sub);margin-top:3px;">No prazo ou adiantado<br>'
        f'<span style="color:var(--green);">atraso ≤ 0 sem</span></div>'
        f'<div style="font-size:10px;color:var(--sub);margin-top:3px;">formou dentro da duração prevista do curso</div>'
        f'</div>'

        f'<div style="background:#eef5f0;border:1px solid var(--amber);border-radius:6px;padding:12px;text-align:center;">'
        f'<div style="font-size:22px;font-weight:700;color:var(--amber);">{n_mod}</div>'
        f'<div style="font-size:10px;color:var(--sub);margin-top:3px;">Atraso moderado<br>'
        f'<span style="color:var(--amber);">1–5 sem além do previsto</span></div>'
        f'<div style="font-size:10px;color:var(--sub);margin-top:3px;">reprovações ou trancamentos pontuais</div>'
        f'</div>'

        f'<div style="background:#eef5f0;border:1px solid var(--red);border-radius:6px;padding:12px;text-align:center;">'
        f'<div style="font-size:22px;font-weight:700;color:var(--red);">{n_long}</div>'
        f'<div style="font-size:10px;color:var(--sub);margin-top:3px;">Atraso longo<br>'
        f'<span style="color:var(--red);">≥ 6 sem além do previsto</span></div>'
        f'<div style="font-size:10px;color:var(--sub);margin-top:3px;">trancamento prolongado ou retomada tardia</div>'
        f'</div>'

        f'</div>'
    )

    curso_blocks = ""
    by_curso_cats = gt.get("by_curso_categories", {})
    for curso, cstats in sorted(by_curso.items()):
        if not cstats:
            continue
        short = "ECA" if "Controle" in curso else "SI"
        expected = expected_semesters(curso)  # SI=8, ECA=12
        cc = by_curso_cats.get(curso, {})
        dist = cstats.get("dist", {})
        max_dist = max(dist.values()) if dist else 1

        def _bar_color(k: int) -> str:
            if k < 4:
                return "var(--blue)"
            if k <= expected + 2:
                return "var(--green)"
            if k <= 24:
                return "var(--amber)"
            return "var(--red)"

        dist_bars = "".join(
            f'<div style="display:flex;align-items:center;gap:6px;margin-bottom:4px;">'
            f'<span style="font-size:10px;color:{_bar_color(k)};width:50px;flex-shrink:0;">{k} sem</span>'
            f'<div style="flex:1;height:10px;background:#e7efe9;border-radius:2px;overflow:hidden;">'
            f'<div style="width:{round(v / max_dist * 100)}%;height:100%;background:{_bar_color(k)};"></div>'
            f'</div>'
            f'<span style="font-size:10px;color:var(--sub);width:20px;text-align:right;">{v}</span>'
            f'</div>'
            for k, v in sorted(dist.items())
        )
        ci = gt.get("cohort_impact", {}).get(curso, {})
        atraso_m = ci.get("atraso_mean")
        atraso_med = ci.get("atraso_median")
        _am = f"{atraso_m:+.1f}".replace(".", ",") if atraso_m is not None else "—"
        _amed = f"{atraso_med:+d}" if isinstance(atraso_med, int) else "—"
        _ac = ("var(--green)" if (atraso_m or 0) <= 1
               else "var(--amber)" if (atraso_m or 0) <= 3 else "var(--red)")
        cat_summary = ""
        if ci.get("on_time"):
            cat_summary += f' · <span style="color:var(--green);">{ci["on_time"]} no prazo</span>'
        if ci.get("late"):
            cat_summary += f' · <span style="color:var(--amber);">{ci["late"]} atrasados</span>'
        curso_blocks += (
            f'<div style="background:#f7faf8;border:1px solid var(--border);border-radius:6px;padding:14px;">'
            f'<div style="font-size:12px;font-weight:600;margin-bottom:4px;">{short} '
            f'<span style="font-weight:400;color:var(--sub);">· previsto {expected} sem</span></div>'
            f'<div style="font-size:11px;color:var(--sub);margin-bottom:4px;">'
            f'atraso médio <strong style="color:{_ac};">{_am} sem</strong> (mediana {_amed}) · '
            f'média absoluta {cstats["mean"]} sem · n={cstats["n"]}'
            f'{cat_summary}</div>'
            f'{dist_bars}</div>'
        )

    # ---- IC vs sem-IC comparison card ----
    icv = gt.get("ic_vs_no_ic", {})
    ic_stats  = icv.get("ic", {})
    no_stats  = icv.get("no_ic", {})
    icv_curso = icv.get("by_curso", {})

    def _fmt(st: dict, key: str) -> str:
        v = st.get(key)
        return f"{v}" if v is not None else "—"

    def _delta_str(ic_val, no_val) -> str:
        if ic_val is None or no_val is None:
            return ""
        d = round(ic_val - no_val, 1)
        color = "var(--green)" if d <= 0 else "var(--amber)"
        sign = "+" if d > 0 else ""
        return f'<span style="color:{color};font-size:10px;"> ({sign}{d})</span>'

    def _row(label: str, ic_st: dict, no_st: dict, key: str) -> str:
        iv = ic_st.get(key)
        nv = no_st.get(key)
        return (
            f'<tr style="border-bottom:1px solid var(--border);">'
            f'<td style="padding:7px 10px;font-size:12px;color:var(--sub);">{label}</td>'
            f'<td style="padding:7px 10px;font-size:13px;font-weight:600;color:var(--green);text-align:center;">'
            f'{_fmt(ic_st, key)}{_delta_str(iv, nv)}</td>'
            f'<td style="padding:7px 10px;font-size:13px;color:var(--sub);text-align:center;">'
            f'{_fmt(no_st, key)}</td>'
            f'</tr>'
        )

    def _section_rows(label: str, ic_st: dict, no_st: dict) -> str:
        header = (
            f'<tr style="background:#eef5f0;">'
            f'<td colspan="3" style="padding:6px 10px;font-size:11px;font-weight:700;'
            f'color:var(--text);text-transform:uppercase;letter-spacing:.05em;">{label}</td>'
            f'</tr>'
        )
        return header + _row("Média (sem)", ic_st, no_st, "mean") + _row("Mediana (sem)", ic_st, no_st, "median")

    table_rows = _section_rows("Geral", ic_stats, no_stats)
    for curso, cv in sorted(icv_curso.items()):
        short = "ECA — Eng. Controle e Automação" if "Controle" in curso else "SI — Sistemas de Informação"
        table_rows += _section_rows(short, cv.get("ic", {}), cv.get("no_ic", {}))

    ic_n  = ic_stats.get("n", 0)
    no_n  = no_stats.get("n", 0)
    ic_table = (
        f'<div style="background:#eef5f0;border:1px solid var(--border);border-radius:8px;'
        f'overflow:hidden;margin-bottom:16px;">'
        f'<table style="width:100%;border-collapse:collapse;">'
        f'<thead><tr style="border-bottom:2px solid var(--border);">'
        f'<th style="padding:10px;text-align:left;font-size:12px;color:var(--sub);">Grupo</th>'
        f'<th style="padding:10px;text-align:center;font-size:12px;color:var(--green);">'
        f'Com IC <span style="font-weight:400;">({ic_n})</span></th>'
        f'<th style="padding:10px;text-align:center;font-size:12px;color:var(--sub);">'
        f'Sem IC <span style="font-weight:400;">({no_n})</span></th>'
        f'</tr></thead>'
        f'<tbody>{table_rows}</tbody>'
        f'</table>'
        f'<div style="padding:8px 10px;font-size:10px;color:var(--sub);">'
        f'Delta (verde) = alunos com IC formam mais rápido que sem IC.'
        f'</div>'
        f'</div>'
    )

    # ---- interpretive insight block ----
    ic_mean = ic_stats.get("mean")
    no_mean = no_stats.get("mean")
    ic_med  = ic_stats.get("median")
    no_med  = no_stats.get("median")

    if ic_mean is not None and no_mean is not None:
        delta_m   = round(ic_mean - no_mean, 1)
        abs_dm    = abs(delta_m)
        negligible_overall = abs_dm < 1.0

        # per-course deltas
        curso_deltas: list[dict] = []
        for curso, cv in sorted(icv_curso.items()):
            ic_c = cv.get("ic", {})
            no_c = cv.get("no_ic", {})
            ic_cm = ic_c.get("mean")
            no_cm = no_c.get("mean")
            ic_cmed = ic_c.get("median")
            no_cmed = no_c.get("median")
            if ic_cm is not None and no_cm is not None:
                d = round(ic_cm - no_cm, 1)
                curso_deltas.append({
                    "short": "ECA" if "Controle" in curso else "SI",
                    "ic_mean": ic_cm, "no_mean": no_cm,
                    "ic_med": ic_cmed, "no_med": no_cmed,
                    "delta": d, "n_ic": ic_c.get("n", 0), "n_no": no_c.get("n", 0),
                })

        # detect masking: any course with |delta| >= 3 AND opposite sign to overall
        masking_courses = [
            c for c in curso_deltas
            if abs(c["delta"]) >= 3 and (
                (negligible_overall) or
                (delta_m > 0 and c["delta"] < 0) or
                (delta_m < 0 and c["delta"] > 0)
            )
        ]
        strongest = max(curso_deltas, key=lambda c: abs(c["delta"])) if curso_deltas else None

        # ---- headline based on overall ----
        if negligible_overall:
            headline = "No agregado, IC não altera o tempo de formação — mas o dado por curso conta outra história"
            headline_color = "var(--amber)"
        elif delta_m < 0:
            headline = f"Alunos com IC se formam mais rápido — {abs_dm} semestre{'s' if abs_dm != 1 else ''} a menos no geral"
            headline_color = "var(--green)"
        else:
            headline = f"Alunos com IC levam {abs_dm} semestre{'s' if abs_dm != 1 else ''} a mais no geral"
            headline_color = "var(--amber)"

        # ---- overall paragraph ----
        _delta_color_overall = "var(--green)" if delta_m <= 0 else "var(--amber)"
        _delta_sign = "−" if delta_m < 0 else "+"
        _delta_plural = "s" if abs_dm != 1 else ""
        overall_text = (
            f"No total, formandos com IC levaram em média <strong>{ic_mean} semestres</strong> "
            f"({ic_mean / 2:.1f} anos) para concluir o curso, contra <strong>{no_mean} semestres</strong> "
            f"({no_mean / 2:.1f} anos) dos sem IC — diferença de "
            f"<strong style='color:{_delta_color_overall};'>"
            f"{_delta_sign}{abs_dm} semestre{_delta_plural}</strong>. "
            f"A mediana segue o mesmo padrão: {ic_med} sem (com IC) vs {no_med} sem (sem IC)."
        )

        # ---- per-course paragraphs ----
        curso_paras = ""
        for c in curso_deltas:
            d = c["delta"]
            abs_d = abs(d)
            d_color = "var(--green)" if d <= -1 else ("var(--amber)" if d >= 1 else "var(--sub)")
            d_anos = abs(d) / 2
            if abs_d < 1:
                verdict = "sem diferença relevante"
                verdict_detail = (
                    f"Com IC: {c['ic_mean']} sem · Sem IC: {c['no_mean']} sem "
                    f"(delta {'−' if d < 0 else '+'}{abs_d} sem — dentro da margem de variação normal)."
                )
            elif d < 0:
                verdict = f"com IC {abs_d} semestres mais rápido ({d_anos:.1f} anos a menos)"
                verdict_detail = (
                    f"Média com IC: <strong>{c['ic_mean']} sem</strong> ({c['ic_mean'] / 2:.1f} anos) · "
                    f"Sem IC: <strong>{c['no_mean']} sem</strong> ({c['no_mean'] / 2:.1f} anos). "
                    f"Mediana: {c['ic_med']} vs {c['no_med']} sem. Base: {c['n_ic']} com IC, {c['n_no']} sem IC."
                )
            else:
                verdict = f"com IC {abs_d} semestres mais lento ({d_anos:.1f} anos a mais)"
                verdict_detail = (
                    f"Média com IC: <strong>{c['ic_mean']} sem</strong> ({c['ic_mean'] / 2:.1f} anos) · "
                    f"Sem IC: <strong>{c['no_mean']} sem</strong> ({c['no_mean'] / 2:.1f} anos). "
                    f"Mediana: {c['ic_med']} vs {c['no_med']} sem. Base: {c['n_ic']} com IC, {c['n_no']} sem IC."
                )
            curso_paras += (
                f'<div style="margin-bottom:12px;">'
                f'<div style="font-size:12px;font-weight:700;color:var(--text);margin-bottom:4px;">'
                f'{c["short"]} — <span style="color:{d_color};">{verdict}</span></div>'
                f'<p style="font-size:12px;color:var(--sub);line-height:1.6;margin:0;">{verdict_detail}</p>'
                f'</div>'
            )

        # ---- masking warning ----
        masking_block = ""
        if masking_courses and negligible_overall:
            names = " e ".join(c["short"] for c in masking_courses)
            s_strongest = masking_courses[0]
            masking_block = (
                f'<div style="background:#fbf6e6;border:1px solid var(--amber);border-radius:6px;'
                f'padding:12px 14px;margin:12px 0;font-size:12px;line-height:1.6;">'
                f'<strong style="color:var(--amber);">⚠ Efeito de agregação:</strong> '
                f'a média geral ({delta_m:+} sem) oculta diferenças expressivas por curso. '
                f'No {names}, o delta chega a '
                f'<strong style="color:var(--amber);">{s_strongest["delta"]:+} semestres</strong> — '
                f'o agregado não é representativo de nenhum dos cursos individualmente.'
                f'</div>'
            )

        # ---- interpretation ----
        if strongest and abs(strongest["delta"]) >= 3 and strongest["delta"] < 0:
            takeaway = (
                f"O dado mais relevante está no {strongest['short']}: alunos que participaram de IC "
                f"concluíram o curso <strong>{abs(strongest['delta'])} semestres mais cedo</strong> "
                f"({abs(strongest['delta']) / 2:.1f} anos) do que os que não participaram. "
                f"Uma hipótese plausível é o <em>efeito de ancoragem</em>: a IC cria vínculo institucional, "
                f"reduz evasão e trancamentos, e mantém o aluno em progressão regular. "
                f"Alunos sem IC no {strongest['short']} apresentam média de {strongest['no_mean']} semestres "
                f"— muito acima do mínimo curricular de {'10' if strongest['short'] == 'ECA' else '8'} semestres — "
                f"sugerindo alta taxa de interrupções nesse grupo. "
                f"Importante: correlação não implica causalidade. Pode haver seleção — "
                f"alunos mais engajados ingressam em IC <em>e</em> terminam mais rápido."
            )
        elif negligible_overall:
            takeaway = (
                "IC não representa risco para a conclusão do curso em nenhum dos cursos analisados. "
                "Onde a diferença existe, favorece os alunos com IC. "
                "O dado sugere que conciliar pesquisa e graduação é viável, e possivelmente benéfico."
            )
        else:
            takeaway = (
                "IC não é fator de atraso. A diferença observada, quando existe, pode refletir "
                "perfil de aluno ou trajetória acadêmica mais intensa — não um custo imposto pela pesquisa."
            )

        insight_block = (
            f'<div style="background:#eef5f0;border:1px solid var(--border);border-left:4px solid {headline_color};'
            f'border-radius:8px;padding:18px 20px;margin-bottom:16px;">'
            f'<div style="font-size:14px;font-weight:700;color:{headline_color};margin-bottom:12px;">'
            f'{headline}</div>'
            f'<p style="font-size:13px;color:var(--text);line-height:1.7;margin-bottom:14px;">{overall_text}</p>'
            f'{masking_block}'
            f'<div style="border-top:1px solid var(--border);padding-top:14px;margin-bottom:12px;">'
            f'<div style="font-size:11px;color:var(--sub);text-transform:uppercase;letter-spacing:.05em;margin-bottom:10px;">Por curso</div>'
            f'{curso_paras}</div>'
            f'<div style="border-top:1px solid var(--border);padding-top:12px;">'
            f'<p style="font-size:12px;color:var(--sub);line-height:1.6;margin:0;">'
            f'<strong style="color:var(--text);">Interpretação:</strong> {takeaway}</p>'
            f'</div>'
            f'</div>'
        )
    else:
        insight_block = ""

    # fellowship impact moved to _sec_fellowship_impact(s)
    ft        = gt.get("fellowship_type", {})  # keep ft_* vars for insight_block only; block itself removed
    ft_paid   = ft.get("paid", {})
    ft_vol    = ft.get("volunteer", {})
    ft_no     = ft.get("no_ic", {})
    ft_by_curso = ft.get("by_curso", {})
    by_sponsor  = gt.get("by_sponsor", {})
    by_fel_name = gt.get("by_fel_name", {})

    SPONSOR_ORDER  = ["Fapes", "Ifes", "CNPq", "Voluntário", "Sem IC"]
    SPONSOR_COLORS = {
        "Fapes": "var(--amber)", "Ifes": "var(--green2)",
        "CNPq": "var(--blue)", "Voluntário": "var(--sub)", "Sem IC": "var(--gray)",
    }
    FEL_COLORS = {
        "PIBIC": "var(--green)", "PIVIC": "var(--amber)",
        "PIBITI": "var(--blue)", "PIVITI": "var(--sub)",
        "PIBIC-JR": "var(--gray)",
    }

    # --- KPI row: paid / volunteer / no_ic ---
    def _ft_kpi(label: str, st: dict, color: str, sub: str) -> str:
        mean = st.get("mean", "—")
        med  = st.get("median", "—")
        n_ft = st.get("n", 0)
        anos = f"{mean / 2:.1f} anos" if isinstance(mean, (int, float)) else "—"
        return (
            f'<div style="background:#eef5f0;border:1px solid {color};border-radius:8px;padding:14px;">'
            f'<div style="font-size:11px;font-weight:700;color:{color};text-transform:uppercase;'
            f'letter-spacing:.05em;margin-bottom:8px;">{label}</div>'
            f'<div style="font-size:26px;font-weight:700;color:{color};line-height:1;">{mean}</div>'
            f'<div style="font-size:10px;color:var(--sub);margin-top:2px;">sem médios ({anos})</div>'
            f'<div style="font-size:11px;color:var(--sub);margin-top:6px;">mediana {med} sem · n={n_ft}</div>'
            f'<div style="font-size:10px;color:var(--sub);margin-top:4px;">{sub}</div>'
            f'</div>'
        )

    ft_kpis = (
        f'<div style="display:grid;grid-template-columns:repeat(3,1fr);gap:12px;margin-bottom:18px;">'
        + _ft_kpi("Bolsa paga", ft_paid, "var(--green)", "Fapes · IFES · CNPq")
        + _ft_kpi("Voluntário", ft_vol, "var(--amber)", "PIVIC / sem bolsa formal")
        + _ft_kpi("Sem IC", ft_no, "var(--sub)", "não participou de pesquisa")
        + f'</div>'
    )

    # --- sponsor × course matrix ---
    def _sp_cell(st: dict, color: str) -> str:
        m = st.get("mean"); med = st.get("median"); n_ft = st.get("n", 0)
        if not m:
            return f'<td style="padding:7px 8px;text-align:center;color:var(--sub);font-size:11px;">—</td>'
        return (
            f'<td style="padding:7px 8px;text-align:center;">'
            f'<span style="font-size:13px;font-weight:700;color:{color};">{m}</span>'
            f'<span style="font-size:10px;color:var(--sub);"> sem</span><br>'
            f'<span style="font-size:10px;color:var(--sub);">med {med} · n={n_ft}</span>'
            f'</td>'
        )

    sp_cols = [sp for sp in SPONSOR_ORDER if sp in by_sponsor or sp == "Sem IC"]
    sp_header = "".join(
        '<th style="padding:8px;text-align:center;font-size:11px;color:' +
        SPONSOR_COLORS.get(sp, "var(--sub)") + f';">{sp}</th>'
        for sp in sp_cols
    )

    sp_rows = ""
    for curso in sorted(by_curso.keys()):
        short = "ECA" if "Controle" in curso else "SI"
        sp_rows += f'<tr style="border-bottom:1px solid var(--border);"><td style="padding:7px 8px;font-size:12px;font-weight:600;">{short}</td>'
        for sp in sp_cols:
            if sp == "Sem IC":
                cst = ft_by_curso.get(curso, {}).get("no_ic", {})
            else:
                cst = by_sponsor.get(sp, {}).get("by_curso", {}).get(curso, {})
            sp_rows += _sp_cell(cst, SPONSOR_COLORS.get(sp, "var(--sub)"))
        sp_rows += "</tr>"

    # overall row
    sp_rows += f'<tr style="background:#eef5f0;border-top:2px solid var(--border);"><td style="padding:7px 8px;font-size:11px;color:var(--sub);">Geral</td>'
    for sp in sp_cols:
        if sp == "Sem IC":
            cst = ft_no
        else:
            cst = by_sponsor.get(sp, {}).get("overall", {})
        sp_rows += _sp_cell(cst, SPONSOR_COLORS.get(sp, "var(--sub)"))
    sp_rows += "</tr>"

    sp_matrix = (
        f'<div style="background:#eef5f0;border:1px solid var(--border);border-radius:8px;'
        f'overflow:hidden;margin-bottom:14px;">'
        f'<div style="padding:8px 10px;font-size:11px;color:var(--sub);text-transform:uppercase;'
        f'letter-spacing:.05em;border-bottom:1px solid var(--border);">Agência × curso — média semestral</div>'
        f'<table style="width:100%;border-collapse:collapse;">'
        f'<thead><tr style="border-bottom:2px solid var(--border);">'
        f'<th style="padding:8px;text-align:left;font-size:11px;color:var(--sub);">Curso</th>'
        f'{sp_header}</tr></thead><tbody>{sp_rows}</tbody></table></div>'
    )

    # --- fellowship name bars (PIBIC / PIVIC / etc.) ---
    fel_items = [(fn, st) for fn, st in by_fel_name.items() if st.get("n", 0) >= 2]
    _fel_max = max((st.get("mean", 0) for _, st in fel_items), default=1)
    fel_bars = ""
    for fname, st in sorted(fel_items, key=lambda x: x[1].get("mean", 99)):
        m = st.get("mean", 0); med = st.get("median", 0); n_ft = st.get("n", 0)
        color = FEL_COLORS.get(fname, "var(--sub)")
        w = round(m / _fel_max * 100)
        fel_bars += (
            f'<div style="display:flex;align-items:center;gap:8px;margin-bottom:6px;">'
            f'<span style="font-size:11px;font-weight:600;color:{color};width:70px;flex-shrink:0;">{fname}</span>'
            f'<div style="flex:1;height:14px;background:#e7efe9;border-radius:3px;overflow:hidden;">'
            f'<div style="width:{w}%;height:100%;background:{color};"></div></div>'
            f'<span style="font-size:11px;font-weight:700;color:{color};width:30px;text-align:right;">{m}</span>'
            f'<span style="font-size:10px;color:var(--sub);">sem · med {med} · n={n_ft}</span>'
            f'</div>'
        )
    fel_section = (
        f'<div style="margin-bottom:14px;">'
        f'<div style="font-size:11px;color:var(--sub);text-transform:uppercase;letter-spacing:.05em;margin-bottom:8px;">'
        f'Por tipo de bolsa — média de semestres até formação</div>'
        f'{fel_bars}</div>'
    ) if fel_bars else ""

    # --- insight ---
    pm = ft_paid.get("mean"); vm = ft_vol.get("mean"); nm = ft_no.get("mean")
    pibic_m = by_fel_name.get("PIBIC", {}).get("mean")
    pivic_m = by_fel_name.get("PIVIC", {}).get("mean")

    if pm and vm and nm:
        pv_delta = round(vm - pm, 1)
        vn_delta = round(vm - nm, 1)

        # per-course deltas for volunteer vs paid (SI, where course is controlled)
        si_paid_m  = next((ft_by_curso.get(c, {}).get("paid",  {}).get("mean") for c in ft_by_curso if "Controle" not in c), None)
        si_vol_m   = next((ft_by_curso.get(c, {}).get("volunteer", {}).get("mean") for c in ft_by_curso if "Controle" not in c), None)
        si_no_m    = next((ft_by_curso.get(c, {}).get("no_ic", {}).get("mean") for c in ft_by_curso if "Controle" not in c), None)
        eca_vol_m  = next((ft_by_curso.get(c, {}).get("volunteer", {}).get("mean") for c in ft_by_curso if "Controle" in c), None)
        eca_no_m   = next((ft_by_curso.get(c, {}).get("no_ic", {}).get("mean") for c in ft_by_curso if "Controle" in c), None)

        si_lines = ""
        if si_paid_m and si_vol_m and si_no_m:
            si_pv = round(si_vol_m - si_paid_m, 1)
            si_vn = round(si_vol_m - si_no_m, 1)
            si_lines = (
                f"No SI (controlando o curso), a diferença é ainda mais clara: "
                f"bolsa paga={si_paid_m} sem · sem IC={si_no_m} sem · voluntário={si_vol_m} sem. "
                f"Voluntário leva <strong>{si_vn} semestre{'s' if si_vn != 1 else ''} a mais que sem IC</strong> "
                f"e <strong>{si_pv} sem a mais que bolsa paga</strong>. "
            )

        pibic_line = ""
        if pibic_m and pivic_m:
            pip_delta = round(pivic_m - pibic_m, 1)
            pibic_line = (
                f"Por tipo de bolsa: PIBIC (pago) média {pibic_m} sem vs PIVIC (voluntário) {pivic_m} sem "
                f"— diferença de <strong style='color:var(--green);'>{pip_delta} semestres ({pip_delta/2:.1f} anos)</strong>. "
            )

        ft_insight = (
            f'<div style="background:#eef5f0;border-left:4px solid var(--green);'
            f'border-radius:6px;padding:14px 16px;font-size:12px;line-height:1.8;">'
            f'<strong style="color:var(--green);font-size:13px;display:block;margin-bottom:8px;">'
            f'Bolsa paga acelera a formação — voluntário é mais lento até que sem IC</strong>'
            f'<p style="margin:0 0 10px;">'
            f'Formandos com bolsa paga concluíram em média em <strong>{pm} semestres</strong> ({pm/2:.1f} anos), '
            f'contra <strong>{vm} semestres</strong> ({vm/2:.1f} anos) dos voluntários e '
            f'<strong>{nm} semestres</strong> ({nm/2:.1f} anos) dos sem IC. '
            f'Delta bolsa paga → voluntário: <strong style="color:var(--amber);">+{pv_delta} semestres ({pv_delta/2:.1f} anos)</strong>. '
            f'Voluntário vs sem IC: <strong style="color:var(--amber);">+{vn_delta} semestres</strong> — '
            f'participar voluntariamente está <em>associado</em> a graduações mais longas que nem participar.'
            f'</p>'
            f'<p style="margin:0 0 10px;color:var(--sub);">{si_lines}{pibic_line}</p>'
            f'<p style="margin:0;color:var(--sub);">'
            f'<strong style="color:var(--text);">Hipóteses:</strong> '
            f'(1) Bolsistas pagos têm obrigações formais com a agência (relatórios periódicos, metas de produção) '
            f'que criam estrutura e reduzem trancamentos. '
            f'(2) Voluntários no ECA são alunos em situação acadêmica mais difícil — '
            f'recorrem à pesquisa voluntária quando não conseguem bolsa paga, e já acumulam atrasos. '
            f'(3) Seleção reversa: orientadores oferecem bolsas pagas a alunos de melhor desempenho, '
            f'que naturalmente concluem mais rápido. '
            f'Os dados não permitem separar causalidade de seleção sem dados de desempenho acadêmico.'
            f'</p></div>'
        )
    else:
        ft_insight = ""

    fellowship_block = ""  # rendered separately by _sec_fellowship_impact(s)

    note = (
        f'<div class="note">'
        f'<strong>Metodologia:</strong> tempo calculado em semestres entre entrada (extraída da matrícula '
        f'— ex: <code>20181BSI…</code> → 2018/1) e semestre de formatura. '
        f'Mínimo curricular: SI = 8 semestres, ECA = 10 semestres. '
        f'Acelerado (&lt;4 sem): provável transferência ou aproveitamento de disciplinas. '
        f'Prolongado (&gt;24 sem): provável trancamento ou reingresso. '
        f'Base: {n} de {total} formandos com matrícula interpretável.'
        f'</div>'
    )

    # ---- tempo de formação: cotistas × ampla concorrência ----
    by_adm = gt.get("by_admission", {})
    adm_block = ""
    if by_adm:
        ADM_ORDER = [
            ("Cotas / Reserva de vagas", "Cotistas", "var(--green)"),
            ("Ampla Concorrência", "Ampla concorrência", "var(--blue)"),
            ("Transferência", "Transferência", "var(--amber)"),
        ]
        _exp_map = gt.get("expected", {})
        def _sig(v) -> str:
            return f"+{v}" if v > 0 else (f"−{abs(v)}" if v < 0 else "0")
        cards = ""
        for key, label, col in ADM_ORDER:
            ov = (by_adm.get(key) or {}).get("overall") or {}
            dl = (by_adm.get(key) or {}).get("delay") or {}
            if not ov:
                continue
            cards += (
                f'<div style="background:#f7faf8;border:1px solid {col};border-radius:8px;'
                f'padding:14px;text-align:center;">'
                f'<div style="font-size:11px;color:{col};font-weight:600;margin-bottom:6px;">{label}</div>'
                f'<div style="font-size:28px;font-weight:700;color:{col};">{_sig(dl.get("mean", 0))}</div>'
                f'<div style="font-size:10px;color:var(--sub);margin-top:3px;">'
                f'semestres de atraso</div>'
                f'<div style="font-size:11px;color:var(--text);margin-top:6px;">'
                f'bruto {ov["mean"]} sem · n={ov["n"]}</div>'
                f'</div>'
            )
        # per-curso blocks: each curso shows cota / ampla / transf side by side
        cur_set = sorted({
            c for key, _, _ in ADM_ORDER
            for c in (by_adm.get(key, {}).get("by_curso", {}) or {})
        })
        curso_blocks_adm = ""
        for c in cur_set:
            short = "ECA" if "Controle" in c else "SI"
            cells = ""
            for key, label, col in ADM_ORDER:
                st = (by_adm.get(key, {}).get("by_curso", {}) or {}).get(c) or {}
                if st:
                    val = (
                        f'<div style="font-size:22px;font-weight:700;color:{col};">{st["mean"]}</div>'
                        f'<div style="font-size:10px;color:var(--sub);">{st["mean"]/2:.1f} anos · '
                        f'med {st["median"]} · n={st["n"]}</div>'
                    )
                else:
                    val = '<div style="font-size:18px;color:var(--sub);">—</div>'
                cells += (
                    f'<div style="background:#eef5f0;border:1px solid {col};border-radius:6px;'
                    f'padding:10px;text-align:center;">'
                    f'<div style="font-size:10px;color:{col};font-weight:600;margin-bottom:4px;">'
                    f'{label}</div>{val}</div>'
                )
            _ce = _exp_map.get(c, 12 if "Controle" in c else 8)
            curso_blocks_adm += (
                f'<div style="background:#f7faf8;border:1px solid var(--border);border-radius:8px;'
                f'padding:14px;margin-bottom:14px;">'
                f'<div style="font-size:12px;font-weight:600;margin-bottom:10px;">'
                f'{short} — {c} '
                f'<span style="color:var(--sub);font-weight:400;">· previsto {_ce} sem</span></div>'
                f'<div style="display:grid;grid-template-columns:repeat(3,1fr);gap:10px;">'
                f'{cells}</div></div>'
            )
        adm_block = (
            f'<div style="font-size:13px;font-weight:600;margin:8px 0 6px;">'
            f'Tempo de formação — cotistas × ampla concorrência</div>'
            f'<div style="font-size:11px;color:var(--sub);margin-bottom:12px;">'
            f'Cursos têm durações diferentes (SI = 8 sem, ECA = 12 sem); a comparação justa é '
            f'o <strong>atraso</strong> (semestres além do previsto). Menor = melhor.</div>'
            f'<div style="display:grid;grid-template-columns:repeat(3,1fr);gap:12px;margin-bottom:16px;">'
            f'{cards}</div>'
            f'<div style="font-size:12px;font-weight:600;margin:6px 0 10px;color:var(--sub);">'
            f'Por curso (média bruta em semestres)</div>{curso_blocks_adm}'
            f'<div class="note" style="margin-bottom:20px;">O <strong>atraso</strong> normaliza '
            f'pela duração de cada curso, permitindo comparar SI (8 sem) e ECA (12 sem) de forma '
            f'justa. Cotistas concluem mais perto do prazo que a ampla concorrência em ambos os cursos.</div>'
        )

    body = (kpi + cat_row + ic_table + insight_block
            + f'<div class="grid2" style="margin-bottom:16px;">{curso_blocks}</div>'
            + adm_block + note)
    return section(
        "Tempo de formação",
        f"Ingresso até a colação, medido como atraso sobre o currículo de cada curso "
        f"(SI {expected_map.get('Sistemas de Informação', 8)} sem · "
        f"ECA {expected_map.get('Engenharia de Controle e Automação', 12)} sem) — "
        f"atraso médio {_dval} sem · base: {n} formandos",
        body,
    )


def _sec_cohort_analysis(s: dict) -> str:
    ca = s.get("cohort_analysis", {})
    if not ca:
        return ""
    years = sorted(ca.keys())
    max_total = max(d["total"] for d in ca.values()) or 1
    max_ic_pct = max(d["ic_pct"] for d in ca.values()) or 1

    rows = ""
    for yr in years:
        d = ca[yr]
        ic_w = round(d["ic_pct"] / 100 * 140)
        total_w = round(d["total"] / max_total * 60)
        grad_color = "var(--green)" if (d["mean_grad"] or 99) <= 12 else ("var(--amber)" if (d["mean_grad"] or 99) <= 20 else "var(--red)")
        grad_str = f"{d['mean_grad']} sem" if d["mean_grad"] else "—"
        rows += (
            f'<tr style="border-bottom:1px solid var(--border);">'
            f'<td style="padding:6px 8px;font-size:12px;font-weight:600;color:var(--text);">{yr}</td>'
            f'<td style="padding:6px 8px;text-align:center;font-size:12px;">{d["total"]}</td>'
            f'<td style="padding:6px 8px;">'
            f'<div style="display:flex;align-items:center;gap:6px;">'
            f'<div style="width:{ic_w}px;height:10px;background:#e7efe9;border-radius:2px;overflow:hidden;">'
            f'<div style="width:100%;height:100%;background:var(--green);transform:scaleX({d["ic_pct"]/100:.2f});transform-origin:left;"></div></div>'
            f'<span style="font-size:11px;color:var(--green);font-weight:600;">{d["ic_pct"]}%</span>'
            f'<span style="font-size:10px;color:var(--sub);">({d["ic"]})</span></div></td>'
            f'<td style="padding:6px 8px;text-align:center;font-size:11px;color:var(--green);">{d["paid"]}</td>'
            f'<td style="padding:6px 8px;text-align:center;font-size:11px;color:var(--amber);">{d["volunteer"]}</td>'
            f'<td style="padding:6px 8px;text-align:center;font-size:13px;font-weight:700;color:{grad_color};">{grad_str}</td>'
            f'</tr>'
        )

    # trend insight
    recent = [d for yr, d in ca.items() if yr >= 2017]
    early  = [d for yr, d in ca.items() if yr < 2017]
    avg_ic_recent = round(sum(d["ic_pct"] for d in recent) / len(recent), 1) if recent else 0
    avg_ic_early  = round(sum(d["ic_pct"] for d in early) / len(early), 1) if early else 0
    trend_text = (
        f'Turmas mais recentes (≥2017): média de {avg_ic_recent}% com IC. '
        f'Turmas anteriores: {avg_ic_early}%. '
        + ("Participação crescendo nas coortes recentes." if avg_ic_recent > avg_ic_early
           else "Participação estável ou decrescente nas coortes recentes.")
    )

    body = (
        f'<div style="overflow-x:auto;">'
        f'<table style="width:100%;border-collapse:collapse;font-size:12px;">'
        f'<thead><tr style="border-bottom:2px solid var(--border);background:#eef5f0;">'
        f'<th style="padding:7px 8px;text-align:left;color:var(--sub);">Ingresso</th>'
        f'<th style="padding:7px 8px;text-align:center;color:var(--sub);">Formandos</th>'
        f'<th style="padding:7px 8px;text-align:left;color:var(--green);">% com IC</th>'
        f'<th style="padding:7px 8px;text-align:center;color:var(--amber);">Pago</th>'
        f'<th style="padding:7px 8px;text-align:center;color:var(--sub);">Voluntário</th>'
        f'<th style="padding:7px 8px;text-align:center;color:var(--sub);">Média grad.</th>'
        f'</tr></thead><tbody>{rows}</tbody></table></div>'
        f'<div class="note" style="margin-top:12px;">{trend_text} '
        f'Tempo médio de graduação decresce nas coortes mais recentes por construção (ingressaram mais tarde). '
        f'Coortes com poucos formandos (n=1–2) têm médias pouco representativas.</div>'
    )
    return section(
        "Análise de coorte por ano de ingresso",
        "Participação em IC e tempo de formação por turma de entrada",
        body,
    )


def _sec_supervisor_impact(s: dict) -> str:
    si = s.get("supervisor_impact", {})
    if not si:
        return ""
    items = sorted(si.items(), key=lambda x: x[1].get("mean", 99))
    if not items:
        return ""
    max_mean = max(st["mean"] for _, st in items) or 1

    rows = ""
    for name, st in items:
        m = st["mean"]; med = st["median"]; n = st["n"]
        color = "var(--green)" if m <= 10 else ("var(--amber)" if m <= 14 else "var(--red)")
        w = round(m / max_mean * 180)
        rows += (
            f'<div style="display:flex;align-items:center;gap:8px;margin-bottom:6px;">'
            f'<span style="font-size:11px;color:var(--text);width:220px;flex-shrink:0;">{name}'
            f'<span style="color:var(--sub);font-size:10px;"> (n={n})</span></span>'
            f'<div style="flex:1;height:12px;background:#e7efe9;border-radius:2px;overflow:hidden;">'
            f'<div style="width:{w}px;max-width:100%;height:100%;background:{color};"></div></div>'
            f'<span style="font-size:12px;font-weight:700;color:{color};width:50px;text-align:right;">{m} sem</span>'
            f'<span style="font-size:10px;color:var(--sub);">med {med}</span>'
            f'</div>'
        )

    overall_mean = s.get("graduation_time", {}).get("overall", {}).get("mean")
    note_text = (
        f'Ordenado do orientador com alunos de menor tempo médio de formação. '
        f'Média geral: {overall_mean} sem. '
        f'Apenas orientadores com ≥2 formandos nesta turma são exibidos. '
        f'Diferenças refletem perfil dos alunos orientados (curso, ano de ingresso) tanto quanto o impacto do orientador.'
    )

    return section(
        "Impacto do orientador no tempo de formação",
        "Tempo médio de formação dos formandos por orientador (≥2 alunos)",
        rows + f'<div class="note" style="margin-top:12px;">{note_text}</div>',
    )


def _sec_fellowship_value_impact(s: dict) -> str:
    fv = s.get("fellowship_value_impact", {})
    if not fv:
        return ""
    ORDER = ["Sem IC", "IC sem valor", "R$ 1–2k", "R$ 2–5k", "R$ 5k+"]
    COLORS = {
        "Sem IC": "var(--sub)", "IC sem valor": "var(--gray)",
        "R$ 1–2k": "var(--blue)", "R$ 2–5k": "var(--amber)", "R$ 5k+": "var(--green)",
    }
    items = [(k, fv[k]) for k in ORDER if k in fv]
    if not items:
        return ""
    max_mean = max(st["mean"] for _, st in items) or 1

    bars = ""
    for label, st in items:
        m = st["mean"]; med = st["median"]; n = st["n"]
        color = COLORS.get(label, "var(--sub)")
        w = round(m / max_mean * 100)
        bars += (
            f'<div style="display:flex;align-items:center;gap:8px;margin-bottom:8px;">'
            f'<span style="font-size:11px;font-weight:600;color:{color};width:110px;flex-shrink:0;">{label}</span>'
            f'<div style="flex:1;height:16px;background:#e7efe9;border-radius:3px;overflow:hidden;">'
            f'<div style="width:{w}%;height:100%;background:{color};"></div></div>'
            f'<span style="font-size:13px;font-weight:700;color:{color};width:50px;text-align:right;">{m}</span>'
            f'<span style="font-size:10px;color:var(--sub);">sem · med {med} · n={n}</span>'
            f'</div>'
        )

    sem_ic_m = fv.get("Sem IC", {}).get("mean")
    top_m = fv.get("R$ 5k+", {}).get("mean")
    bot_m = fv.get("IC sem valor", {}).get("mean")
    insight = ""
    if sem_ic_m and top_m and bot_m:
        d_top = round(top_m - sem_ic_m, 1)
        d_bot = round(bot_m - sem_ic_m, 1)
        color_top = "var(--green)" if d_top <= 0 else "var(--amber)"
        color_bot = "var(--amber)" if d_bot > 0 else "var(--green)"
        insight = (
            f'<div class="note" style="border-color:var(--green);margin-top:12px;">'
            f'<strong>Gradiente claro:</strong> quanto maior o investimento em bolsa, menor o tempo de formação. '
            f'R$5k+ conclui em {top_m} sem — '
            f'<span style="color:{color_top};">{abs(d_top)} sem {"mais rápido" if d_top < 0 else "mais lento"} que sem IC</span>. '
            f'IC sem valor registrado leva {bot_m} sem — '
            f'<span style="color:{color_bot};">{abs(d_bot)} sem {"mais lento" if d_bot > 0 else "mais rápido"} que sem IC</span>. '
            f'Hipóteses: bolsas de maior valor têm maior duração e obrigações mais formais; '
            f'"IC sem valor" pode incluir registros antigos ou projetos com bolsas não cadastradas no SigPesq.'
            f'</div>'
        )

    return section(
        "Investimento em bolsa × tempo de formação",
        "Tempo médio de formação agrupado pelo total recebido em bolsas (valor × meses)",
        bars + insight,
    )


def _sec_ic_continuity(s: dict) -> str:
    cont = s.get("ic_continuity", {})
    if not cont:
        return ""
    ORDER = ["Sem IC", "1 projeto", "2 projetos", "3+ projetos"]
    COLORS = {
        "Sem IC": "var(--sub)", "1 projeto": "var(--amber)",
        "2 projetos": "var(--green2)", "3+ projetos": "var(--green)",
    }
    items = [(k, cont[k]) for k in ORDER if k in cont]
    if not items:
        return ""
    max_mean = max(st["mean"] for _, st in items) or 1

    bars = ""
    for label, st in items:
        m = st["mean"]; med = st["median"]; n = st["n"]
        color = COLORS.get(label, "var(--sub)")
        w = round(m / max_mean * 100)
        bars += (
            f'<div style="display:flex;align-items:center;gap:8px;margin-bottom:8px;">'
            f'<span style="font-size:12px;font-weight:600;color:{color};width:110px;flex-shrink:0;">{label}</span>'
            f'<div style="flex:1;height:16px;background:#e7efe9;border-radius:3px;overflow:hidden;">'
            f'<div style="width:{w}%;height:100%;background:{color};"></div></div>'
            f'<span style="font-size:14px;font-weight:700;color:{color};width:50px;text-align:right;">{m}</span>'
            f'<span style="font-size:10px;color:var(--sub);">sem · med {med} · n={n}</span>'
            f'</div>'
        )

    m1 = cont.get("1 projeto", {}).get("mean")
    m3 = cont.get("3+ projetos", {}).get("mean")
    insight = ""
    if m1 and m3:
        delta = round(m1 - m3, 1)
        insight = (
            f'<div class="note" style="border-color:var(--green);margin-top:12px;">'
            f'<strong>Mais projetos, formação mais rápida:</strong> '
            f'formandos com 3+ projetos de IC concluíram em média {m3} sem, '
            f'contra {m1} sem dos que fizeram apenas 1 projeto — '
            f'<strong style="color:var(--green);">{delta} semestres a menos</strong>. '
            f'Persistência na pesquisa está associada a conclusão mais rápida, '
            f'possivelmente pelo vínculo contínuo com a instituição e orientador.'
            f'</div>'
        )

    return section(
        "Continuidade na IC × tempo de formação",
        "Formandos com 1, 2 ou 3+ projetos de IC versus sem IC",
        bars + insight,
    )


def _sec_ic_tcc_pipeline(s: dict) -> str:
    p = s.get("ic_tcc_pipeline", {})
    if not p:
        return ""

    both = p.get("both", 0)
    both_pct = p.get("both_pct", 0)
    ic_only = p.get("ic_only", 0)
    tcc_only = p.get("tcc_only", 0)
    same_sup = p.get("same_sup", 0)
    same_sup_pct = p.get("same_sup_pct", 0)
    ic_total = both + ic_only
    total = s.get("total", 0)
    records = p.get("pipeline_records", [])

    # ---- gráfico de funil: cada etapa retém uma fração da anterior ----
    stages = [
        {"label": "Formandos (turmas analisadas)", "value": total, "color": "var(--gray)"},
        {"label": "Com IC registrada (SigPesq)", "value": ic_total, "color": "var(--blue)"},
        {"label": "IC + TCC no Lattes", "value": both, "color": "var(--green2)"},
        {"label": "Continuidade: mesmo orientador IC→TCC", "value": same_sup, "color": "var(--green)"},
    ]
    top_v = stages[0]["value"] or 1
    funnel_rows = ""
    prev_v = None
    for st in stages:
        v = st["value"]
        w = max(round(v / top_v * 100), 14)  # largura mínima p/ legibilidade
        pct_top = round(v / top_v * 100, 1)
        if prev_v is not None:
            conv = round(v / prev_v * 100, 1) if prev_v else 0
            funnel_rows += (
                f'<div style="text-align:center;font-size:10px;color:var(--sub);margin:3px 0;">'
                f'▼ <strong style="color:var(--text);">{conv}%</strong> seguem para a próxima etapa</div>'
            )
        funnel_rows += (
            f'<div style="width:{w}%;min-width:150px;margin:0 auto;background:{st["color"]};'
            f'border-radius:6px;padding:11px 16px;display:flex;justify-content:space-between;'
            f'align-items:center;gap:12px;color:#fff;box-shadow:0 1px 3px rgba(16,40,24,.12);">'
            f'<span style="font-size:11px;font-weight:600;line-height:1.25;">{st["label"]}</span>'
            f'<span style="font-size:18px;font-weight:800;white-space:nowrap;">{v}'
            f'<span style="font-size:10px;font-weight:600;opacity:.85;"> · {pct_top}%</span></span>'
            f'</div>'
        )
        prev_v = v
    funnel = (
        f'<div style="margin-bottom:16px;padding:4px 0;">{funnel_rows}'
        f'<div style="text-align:center;font-size:10px;color:var(--sub);margin-top:8px;">'
        f'% ao lado de cada barra = fração do topo (formandos); % entre barras = conversão da etapa anterior.'
        f'</div></div>'
    )

    names_block = ""
    if records:
        items_html = "".join(
            f'<div style="padding:5px 8px;font-size:12px;background:#f7faf8;'
            f'border-radius:3px;margin-bottom:3px;">'
            f'<span style="color:var(--text);">{r["name"]}</span>'
            f'<span style="color:var(--sub);font-size:10px;"> — IC: {", ".join(r["ic_sup"][:2])}</span>'
            f'</div>'
            for r in records[:15]
        )
        suffix = f'<div style="font-size:10px;color:var(--sub);padding:4px 8px;">... e mais {len(records)-15}</div>' if len(records) > 15 else ""
        names_block = (
            f'<div style="margin-top:12px;">'
            f'<div style="font-size:11px;color:var(--sub);text-transform:uppercase;letter-spacing:.05em;margin-bottom:6px;">'
            f'Formandos com IC + TCC (mesmo orientador)</div>'
            f'{items_html}{suffix}</div>'
        )

    insight = (
        f'<div class="note" style="border-color:var(--green2);margin-top:12px;">'
        f'<strong>{both_pct}% dos formandos com IC também têm TCC registrado no Lattes</strong> — '
        f'alta taxa de conversão IC→TCC. '
        f'E <strong>{same_sup_pct}%</strong> desses ({same_sup} de {both}) mantiveram o <strong>mesmo orientador</strong> '
        f'da IC até o TCC — o vínculo de orientação se prolonga do projeto de iniciação ao trabalho de conclusão. '
        + (f'{p.get("tcc_only", 0)} formandos têm TCC mas não aparecem no SigPesq — '
           f'possível subregistro de IC ou orientações informais.' if tcc_only else "")
        + f'</div>'
    )

    return section(
        "Pipeline IC → TCC",
        "Formandos que fizeram IC e também registraram TCC no Lattes",
        funnel + names_block + insight,
        border_color="var(--green2)",
    )


def _sec_fellowship_impact(s: dict) -> str:
    gt = s.get("graduation_time", {})
    if not gt:
        return ""
    ft          = gt.get("fellowship_type", {})
    ft_paid     = ft.get("paid", {})
    ft_vol      = ft.get("volunteer", {})
    ft_no       = ft.get("no_ic", {})
    ft_by_curso = ft.get("by_curso", {})
    by_sponsor  = gt.get("by_sponsor", {})
    by_fel_name = gt.get("by_fel_name", {})
    if not ft_paid and not ft_vol:
        return ""

    SPONSOR_ORDER  = ["Fapes", "Ifes", "CNPq", "Voluntário", "Sem IC"]
    SPONSOR_COLORS = {
        "Fapes": "var(--amber)", "Ifes": "var(--green2)",
        "CNPq": "var(--blue)", "Voluntário": "var(--sub)", "Sem IC": "var(--gray)",
    }
    FEL_COLORS = {
        "PIBIC": "var(--green)", "PIVIC": "var(--amber)",
        "PIBITI": "var(--blue)", "PIVITI": "var(--sub)", "PIBIC-JR": "var(--gray)",
    }

    # KPI row
    def _kpi(label: str, st: dict, color: str, sub: str) -> str:
        mean = st.get("mean", "—"); med = st.get("median", "—"); n_k = st.get("n", 0)
        anos = f"{mean / 2:.1f} anos" if isinstance(mean, (int, float)) else "—"
        return (
            f'<div style="background:#eef5f0;border:1px solid {color};border-radius:8px;padding:14px;">'
            f'<div style="font-size:11px;font-weight:700;color:{color};text-transform:uppercase;'
            f'letter-spacing:.05em;margin-bottom:8px;">{label}</div>'
            f'<div style="font-size:26px;font-weight:700;color:{color};line-height:1;">{mean}</div>'
            f'<div style="font-size:10px;color:var(--sub);margin-top:2px;">sem médios ({anos})</div>'
            f'<div style="font-size:11px;color:var(--sub);margin-top:6px;">mediana {med} sem · n={n_k}</div>'
            f'<div style="font-size:10px;color:var(--sub);margin-top:4px;">{sub}</div>'
            f'</div>'
        )

    kpis = (
        f'<div style="display:grid;grid-template-columns:repeat(3,1fr);gap:12px;margin-bottom:18px;">'
        + _kpi("Bolsa paga", ft_paid, "var(--green)", "Fapes · IFES · CNPq")
        + _kpi("Voluntário", ft_vol, "var(--amber)", "PIVIC / sem bolsa formal")
        + _kpi("Sem IC", ft_no, "var(--sub)", "não participou de pesquisa")
        + f'</div>'
    )

    # sponsor × course matrix
    sp_cols = [sp for sp in SPONSOR_ORDER if sp in by_sponsor or sp == "Sem IC"]

    def _cell(st: dict, color: str) -> str:
        m = st.get("mean"); med = st.get("median"); n_c = st.get("n", 0)
        if not m:
            return f'<td style="padding:7px 8px;text-align:center;color:var(--sub);font-size:11px;">—</td>'
        return (
            f'<td style="padding:7px 8px;text-align:center;">'
            f'<span style="font-size:13px;font-weight:700;color:{color};">{m}</span>'
            f'<span style="font-size:10px;color:var(--sub);"> sem</span><br>'
            f'<span style="font-size:10px;color:var(--sub);">med {med} · n={n_c}</span>'
            f'</td>'
        )

    sp_header = "".join(
        '<th style="padding:8px;text-align:center;font-size:11px;color:' +
        SPONSOR_COLORS.get(sp, "var(--sub)") + f';">{sp}</th>'
        for sp in sp_cols
    )
    sp_rows = ""
    for curso in sorted(ft_by_curso.keys()):
        short = "ECA" if "Controle" in curso else "SI"
        sp_rows += f'<tr style="border-bottom:1px solid var(--border);"><td style="padding:7px 8px;font-size:12px;font-weight:600;">{short}</td>'
        for sp in sp_cols:
            cst = ft_by_curso.get(curso, {}).get("no_ic", {}) if sp == "Sem IC" else by_sponsor.get(sp, {}).get("by_curso", {}).get(curso, {})
            sp_rows += _cell(cst, SPONSOR_COLORS.get(sp, "var(--sub)"))
        sp_rows += "</tr>"
    # overall row
    sp_rows += '<tr style="background:#eef5f0;border-top:2px solid var(--border);"><td style="padding:7px 8px;font-size:11px;color:var(--sub);">Geral</td>'
    for sp in sp_cols:
        cst = ft_no if sp == "Sem IC" else by_sponsor.get(sp, {}).get("overall", {})
        sp_rows += _cell(cst, SPONSOR_COLORS.get(sp, "var(--sub)"))
    sp_rows += "</tr>"

    matrix = (
        f'<div style="background:#eef5f0;border:1px solid var(--border);border-radius:8px;overflow:hidden;margin-bottom:14px;">'
        f'<div style="padding:8px 10px;font-size:11px;color:var(--sub);text-transform:uppercase;'
        f'letter-spacing:.05em;border-bottom:1px solid var(--border);">Agência × curso — média semestral (anos de formação)</div>'
        f'<table style="width:100%;border-collapse:collapse;">'
        f'<thead><tr style="border-bottom:2px solid var(--border);">'
        f'<th style="padding:8px;text-align:left;font-size:11px;color:var(--sub);">Curso</th>'
        f'{sp_header}</tr></thead><tbody>{sp_rows}</tbody></table></div>'
    )

    # fellowship name bars
    fel_items = [(fn, st) for fn, st in by_fel_name.items() if st.get("n", 0) >= 2]
    fel_bars = ""
    if fel_items:
        _fel_max = max(st.get("mean", 0) for _, st in fel_items) or 1
        for fname, st in sorted(fel_items, key=lambda x: x[1].get("mean", 99)):
            m = st.get("mean", 0); med = st.get("median", 0); n_f = st.get("n", 0)
            color = FEL_COLORS.get(fname, "var(--sub)")
            w = round(m / _fel_max * 100)
            fel_bars += (
                f'<div style="display:flex;align-items:center;gap:8px;margin-bottom:6px;">'
                f'<span style="font-size:11px;font-weight:600;color:{color};width:70px;flex-shrink:0;">{fname}</span>'
                f'<div style="flex:1;height:14px;background:#e7efe9;border-radius:3px;overflow:hidden;">'
                f'<div style="width:{w}%;height:100%;background:{color};"></div></div>'
                f'<span style="font-size:11px;font-weight:700;color:{color};width:30px;text-align:right;">{m}</span>'
                f'<span style="font-size:10px;color:var(--sub);">sem · med {med} · n={n_f}</span>'
                f'</div>'
            )
    fel_section = (
        f'<div style="margin-bottom:14px;">'
        f'<div style="font-size:11px;color:var(--sub);text-transform:uppercase;letter-spacing:.05em;margin-bottom:8px;">'
        f'Por tipo de bolsa — média de semestres até formação (ordenado do mais rápido)</div>'
        f'{fel_bars}</div>'
    ) if fel_bars else ""

    # insight text
    pm = ft_paid.get("mean"); vm = ft_vol.get("mean"); nm = ft_no.get("mean")
    pibic_m = by_fel_name.get("PIBIC", {}).get("mean")
    pivic_m = by_fel_name.get("PIVIC", {}).get("mean")

    if pm and vm and nm:
        pv_delta = round(vm - pm, 1)
        vn_delta = round(vm - nm, 1)
        si_paid_m = next((ft_by_curso.get(c, {}).get("paid",      {}).get("mean") for c in ft_by_curso if "Controle" not in c), None)
        si_vol_m  = next((ft_by_curso.get(c, {}).get("volunteer", {}).get("mean") for c in ft_by_curso if "Controle" not in c), None)
        si_no_m   = next((ft_by_curso.get(c, {}).get("no_ic",     {}).get("mean") for c in ft_by_curso if "Controle" not in c), None)

        si_lines = ""
        if si_paid_m and si_vol_m and si_no_m:
            si_pv = round(si_vol_m - si_paid_m, 1)
            si_vn = round(si_vol_m - si_no_m, 1)
            si_lines = (
                f" No SI (curso controlado), o padrão se repete: bolsa paga={si_paid_m} sem · "
                f"sem IC={si_no_m} sem · voluntário={si_vol_m} sem. "
                f"Voluntário leva <strong>{si_vn} sem a mais que sem IC</strong> "
                f"e <strong>{si_pv} sem a mais que bolsa paga</strong>."
            )

        pibic_line = ""
        if pibic_m and pivic_m:
            pip_delta = round(pivic_m - pibic_m, 1)
            pibic_line = (
                f" Por tipo de bolsa: PIBIC (pago) média {pibic_m} sem vs PIVIC (voluntário) {pivic_m} sem "
                f"— <strong style='color:var(--green);'>{pip_delta} sem de diferença ({pip_delta / 2:.1f} anos)</strong>."
            )

        insight = (
            f'<div style="background:#eef5f0;border-left:4px solid var(--green);'
            f'border-radius:6px;padding:14px 16px;font-size:12px;line-height:1.8;">'
            f'<strong style="color:var(--green);font-size:13px;display:block;margin-bottom:8px;">'
            f'Bolsa paga acelera a formação — voluntário é mais lento até que sem IC</strong>'
            f'<p style="margin:0 0 10px;">'
            f'Formandos com bolsa paga concluíram em média em <strong>{pm} semestres</strong> ({pm / 2:.1f} anos), '
            f'contra <strong>{vm} semestres</strong> ({vm / 2:.1f} anos) dos voluntários e '
            f'<strong>{nm} semestres</strong> ({nm / 2:.1f} anos) dos sem IC. '
            f'Delta bolsa paga → voluntário: <strong style="color:var(--amber);">+{pv_delta} semestres ({pv_delta / 2:.1f} anos)</strong>. '
            f'Voluntário vs sem IC: <strong style="color:var(--amber);">+{vn_delta} semestres</strong> — '
            f'participar voluntariamente está <em>associado</em> a graduações mais longas que nem participar.'
            f'{si_lines}{pibic_line}</p>'
            f'<p style="margin:0;color:var(--sub);">'
            f'<strong style="color:var(--text);">Hipóteses:</strong> '
            f'(1) Bolsistas pagos têm obrigações formais com a agência (relatórios periódicos, metas), '
            f'o que pode reduzir trancamentos e manter progressão curricular regular. '
            f'(2) Voluntários no ECA concentram alunos em situação acadêmica mais difícil — '
            f'recorrem à pesquisa sem bolsa quando não conseguem bolsa paga, e já acumulam atrasos. '
            f'(3) Seleção reversa: orientadores oferecem bolsas pagas a alunos de melhor desempenho, '
            f'que naturalmente concluem mais rápido. '
            f'Os dados não permitem separar causalidade de seleção sem controlar desempenho acadêmico.'
            f'</p></div>'
        )
    else:
        insight = ""

    body = kpis + matrix + fel_section + insight
    return section(
        "Bolsa × tempo de formação",
        "Impacto da agência de fomento e modalidade de bolsa na duração do curso",
        body,
        border_color="var(--amber)",
    )


# ---------------------------------------------------------------------------
# HTML assembly
# ---------------------------------------------------------------------------

_SEM_PERIOD = {
    "1": ("Fevereiro", "Julho"),
    "2": ("Agosto",    "Dezembro"),
}

def _fmt_brl(value: float, cents: bool = False) -> str:
    """Formata número em moeda pt-BR: 48547973.83 → 'R$ 48.547.974'."""
    try:
        v = float(value or 0)
    except (TypeError, ValueError):
        v = 0.0
    if cents:
        s = f"{v:,.2f}".replace(",", "X").replace(".", ",").replace("X", ".")
    else:
        s = f"{round(v):,}".replace(",", ".")
    return f"R$ {s}"


def _sec_fapes_ecosystem(s: dict) -> str:
    """Orientadores de formandos que também coordenam projetos FAPES."""
    eco = s.get("integrated", {}).get("fapes_ecosystem", {})
    rows = sorted(eco.get("rows", []),
                  key=lambda r: (-r["formandos"], -r["n_proj"]))
    if not rows:
        return ""
    max_form = max((r["formandos"] for r in rows), default=0) or 1

    body = (
        '<div style="overflow-x:auto;">'
        '<div style="display:flex;font-size:10px;color:var(--sub);font-weight:600;'
        'padding:0 0 6px;border-bottom:1px solid #e7efe9;margin-bottom:8px;">'
        '<span style="width:240px;flex-shrink:0;">Orientador</span>'
        '<span style="flex:1;min-width:160px;text-align:right;">formandos mentorados</span>'
        '<span style="width:90px;text-align:right;">projetos FAPES</span>'
        '</div>'
    )
    for r in rows:
        w = round(r["formandos"] / max_form * 100)
        body += (
            '<div style="display:flex;align-items:center;font-size:11px;'
            'margin-bottom:5px;">'
            f'<span style="width:240px;flex-shrink:0;color:var(--text);">{r["nome"]}</span>'
            '<span style="flex:1;min-width:160px;display:flex;align-items:center;gap:6px;justify-content:flex-end;">'
            '<span style="flex:1;height:10px;background:#e7efe9;border-radius:2px;overflow:hidden;max-width:200px;">'
            f'<span style="display:block;width:{w}%;height:100%;background:var(--green);"></span></span>'
            f'<span style="font-weight:700;color:var(--green);width:36px;text-align:right;">{r["formandos"]}</span></span>'
            f'<span style="width:90px;text-align:right;color:var(--sub);">{r["n_proj"]}</span>'
            '</div>'
        )
    body += '</div>'

    note = (
        f'<div class="note" style="border-color:var(--blue);margin-top:14px;">'
        f'<strong>{eco["n_orientadores"]} dos orientadores de IC dos formandos também coordenam projetos FAPES</strong> '
        f'(de {eco["n_coord_total"]} coordenadores FAPES no campus). '
        f'Somam <strong>{_fmt_brl(eco["total_orc"])}</strong> em orçamento contratado — '
        f'<strong>{eco["pct_orc"]}% dos {_fmt_brl(eco["campus_orc"])}</strong> do portfólio FAPES do campus — '
        f'e {_fmt_brl(eco["total_bolsa"])} em bolsas ({eco["total_bolsa_q"]} cotas). '
        f'A bolsa de IC é a porta de entrada de um ecossistema de pesquisa muito maior: '
        f'os mesmos professores que mentoram graduandos comandam os projetos de grande porte. '
        f'<br><span style="color:var(--sub);">Cruzamento por nome normalizado entre orientadores SigPesq e coordenadores FAPES. '
        f'Recorte exclui os projetos de um coordenador específico, por decisão da gestão.</span>'
        f'</div>'
    )
    return section(
        "Ecossistema do orientador — IC × projetos FAPES",
        "Orientadores de formandos cruzados com o portfólio de projetos FAPES que coordenam",
        body + note,
        border_color="var(--blue)",
    )


def _sec_mestrado_pipeline(s: dict) -> str:
    """Funil graduação → mestrado PPComp + bolsas de pós-graduação."""
    mp = s.get("integrated", {}).get("mestrado_pipeline", {})
    if not mp:
        return ""

    def card(label, value, sub, color):
        return (
            '<div style="flex:1;min-width:120px;background:#f4f9f5;border-radius:8px;'
            'padding:14px 12px;text-align:center;">'
            f'<div style="font-size:26px;font-weight:800;color:{color};line-height:1;">{value}</div>'
            f'<div style="font-size:11px;font-weight:600;color:var(--text);margin-top:5px;">{label}</div>'
            f'<div style="font-size:10px;color:var(--sub);margin-top:2px;">{sub}</div>'
            '</div>'
        )

    cards = (
        '<div style="display:flex;flex-wrap:wrap;gap:10px;margin-bottom:14px;">'
        + card("Formandos → PPComp", mp["pipe_n"], f'{mp["pipe_pct"]}% dos formandos', "var(--green)")
        + card("Já defenderam", mp["pipe_def"], "do pipeline", "var(--blue)")
        + card("Ainda ativos", mp["pipe_ativos"], "do pipeline", "var(--amber)")
        + card("Mestrado (total)", mp["total"], f'taxa defesa {mp["taxa_defesa"]}%', "var(--text)")
        + card("Tempo médio", f'{mp["tempo_medio"]}a', "até a defesa", "var(--sub)")
        + '</div>'
    )

    # mestrado situação geral
    fapes = mp["bolsas_fapes"]; facto = mp["bolsas_facto"]; uniao = mp["bolsas_uniao"]
    bolsas_line = (
        '<div style="display:flex;gap:18px;flex-wrap:wrap;font-size:12px;'
        'padding:10px 0;border-top:1px solid #e7efe9;margin-top:4px;">'
        f'<span><strong>{uniao}</strong> bolsistas de pós (FAPES {fapes} · FACTO {facto})</span>'
        f'<span style="color:var(--sub);">valor FAPES mestrado: <strong style="color:var(--text);">{_fmt_brl(mp["bolsas_valor_fapes"])}</strong></span>'
        f'<span style="color:var(--sub);">defendidos {mp["defendidos"]} · ativos {mp["ativos"]} · evasão {mp["evasao"]}</span>'
        '</div>'
    )

    note = (
        '<div class="note" style="border-color:var(--green);margin-top:12px;">'
        f'<strong>O funil de pesquisa não termina no TCC.</strong> '
        f'{mp["pipe_n"]} formandos do Serra ({mp["pipe_pct"]}%) seguiram para o mestrado PPComp — '
        f'estendendo a trajetória IC → TCC → pós. Apenas 2 desse grupo evadiram. '
        f'A continuidade de <em>bolsa</em> (bolsista IC → bolsista mestrado) ainda é invisível: '
        f'exige chave de junção por aluno (Lattes/SigPesq), não só por nome. '
        f'<br><span style="color:var(--sub);">Fonte: mestrado/ppcomp_base_analitico.json (campo pipeline).</span>'
        '</div>'
    )
    return section(
        "Pipeline graduação → mestrado (PPComp)",
        "Quantos formandos seguiram para o mestrado e o desfecho do programa",
        cards + bolsas_line + note,
        border_color="var(--green)",
    )


def _sec_fomento_panel(s: dict) -> str:
    """Escala de fomento: da semente IC ao portfólio de projetos."""
    fp = s.get("integrated", {}).get("fomento_panel", {})
    if not fp:
        return ""

    steps = []
    steps.append(("Bolsa IC (graduação)", fp.get("ic_grad_invest", 0),
                  "investimento em bolsas de IC dos formandos", "var(--sub)"))
    if fp.get("bolsistas_alocado"):
        steps.append(("Bolsistas FAPES (alocado)", fp["bolsistas_alocado"],
                      "valor alocado a bolsistas do campus", "var(--gray)"))
    if fp.get("fapes_total"):
        ft = fp["fapes_total"]
        steps.append(("Projetos FAPES — Serra", ft.get("orcamento_contratado_total", 0),
                      f'{ft.get("quantidade_projetos", 0)} projetos · {_fmt_brl(ft.get("valor_bolsas_total", 0))} em bolsas · {ft.get("quantidade_bolsas_total", 0)} cotas',
                      "var(--blue)"))
    if fp.get("facto_aprovado"):
        _fp_n = fp.get("facto_proj", 0)
        steps.append(("Fundação FACTO — IFES Serra", fp["facto_aprovado"],
                      f'{_fp_n} projeto{"s" if _fp_n != 1 else ""} executado{"s" if _fp_n != 1 else ""} pelo Campus Serra (valor aprovado)',
                      "var(--green)"))

    max_v = max((v for _, v, _, _ in steps), default=0) or 1
    bars = ""
    for label, value, sub, color in steps:
        # escala log-ish: usar raiz para não esmagar a barra da IC
        w = round((value / max_v) ** 0.45 * 100)
        bars += (
            '<div style="margin-bottom:12px;">'
            '<div style="display:flex;justify-content:space-between;align-items:baseline;margin-bottom:3px;">'
            f'<span style="font-size:12px;font-weight:600;color:{color};">{label}</span>'
            f'<span style="font-size:14px;font-weight:800;color:{color};">{_fmt_brl(value)}</span>'
            '</div>'
            '<div style="height:14px;background:#e7efe9;border-radius:3px;overflow:hidden;">'
            f'<div style="width:{max(w, 2)}%;height:100%;background:{color};"></div></div>'
            f'<div style="font-size:10px;color:var(--sub);margin-top:2px;">{sub}</div>'
            '</div>'
        )

    # flag de gestão: projetos em andamento com prazo encerrado
    flag = ""
    enc = fp.get("fapes_prazo_encerrado") or {}
    if enc.get("quantidade_projetos"):
        flag = (
            '<div class="note" style="border-color:var(--amber);margin-top:8px;">'
            f'<strong>Alerta de gestão:</strong> {enc["quantidade_projetos"]} projetos FAPES '
            f'constam "Em Andamento" mas com prazo já encerrado — '
            f'{_fmt_brl(enc.get("orcamento_contratado_total", 0))} contratados, '
            f'{enc.get("quantidade_bolsas_total", 0)} bolsas. Divergência status × prazo a revisar.'
            '</div>'
        )

    note = (
        '<div class="note" style="border-color:var(--blue);margin-top:12px;">'
        '<strong>A IC é a semente, não a árvore.</strong> '
        'A escala (eixo comprimido por raiz para caber a IC) mostra que a bolsa de iniciação '
        'científica é uma fração mínima do fomento à pesquisa — os projetos FAPES do campus '
        'movem dezenas de milhões, comandados em boa parte pelos mesmos orientadores de IC. '
        'Apenas os projetos FACTO executados pelo Campus Serra entram no recorte (a base FACTO original é institucional). '
        '<br><span style="color:var(--sub);">FACTO = valor aprovado (não executado), filtrado por instituição executora = Campus Serra. '
        'Fontes: projetos-fapes, projetos-facto, bolsistas, formandos.</span>'
        '</div>'
    )
    return section(
        "Painel de fomento — da semente IC ao portfólio",
        "Escala comparada: bolsa de IC vs projetos FAPES vs fundação FACTO — recorte IFES Serra",
        bars + flag + note,
        border_color="var(--blue)",
    )


def _sem_full_period(semester: str) -> str:
    """'2024_1' → '1º Semestre de 2024 (Fevereiro – Julho)'"""
    try:
        year, sem = semester.split("_")
        start, end = _SEM_PERIOD.get(sem, ("?", "?"))
        ordinal = "1º" if sem == "1" else "2º"
        return f"{ordinal} Semestre de {year} ({start}–{end})"
    except Exception:
        return semester.replace("_", ".")


def render_html(s: dict, semester: str, generated_at: str,
                semesters: list | None = None) -> str:
    sem_label = semester.replace("_", ".")
    sem_period = _sem_full_period(semester)
    if semesters and len(semesters) > 1:
        ss = sorted(semesters)
        sem_label = f"{ss[0].replace('_', '.')} – {ss[-1].replace('_', '.')}"
        sem_period = f"Análise consolidada — {len(semesters)} semestres"
    body_parts = [
        _sec_stats(s),
        _sec_curso(s),
        _sec_admission(s),
        _sec_fellowship(s),
        _sec_agencies(s),
        _sec_curso_sponsor(s),
        _sec_projects_duration(s),
        _sec_progressao(s),
        _sec_orientadores(s),
        _sec_rg(s),
        _sec_ka(s),
        _sec_artigos(),
        _sec_graduation_time(s),
        _sec_fellowship_impact(s),
        _sec_cohort_analysis(s),
        _sec_supervisor_impact(s),
        _sec_fellowship_value_impact(s),
        _sec_ic_continuity(s),
        _sec_ic_tcc_pipeline(s),
        _sec_ic_timing(s),
        _sec_ic_recovery(s),
        _sec_bolsistas(s),
        _sec_lattes_cross(s),
        _sec_fapes_ecosystem(s),
        _sec_mestrado_pipeline(s),
        _sec_fomento_panel(s),
    ]
    body = "\n".join(p for p in body_parts if p)

    return f"""<!DOCTYPE html>
<html lang="pt-BR">
<head>
<meta charset="UTF-8">
<meta name="viewport" content="width=device-width, initial-scale=1.0">
<title>Formandos × Pesquisa — IFES Serra {sem_label} — {sem_period}</title>
<link rel="preconnect" href="https://fonts.googleapis.com">
<link href="https://fonts.googleapis.com/css2?family=Inter:wght@400;500;600;700;800&display=swap" rel="stylesheet">
<style>{CSS}</style>
</head>
<body>
<header>
  <div class="eyebrow">IFES Serra · Relatório de Pesquisa</div>
  <h1>Formandos × Pesquisa — {sem_label}</h1>
  <p style="font-size:16px;font-weight:600;color:var(--green);margin-bottom:6px;">{sem_period}</p>
  <p>Análise da participação em iniciação científica e tecnológica dos formandos do semestre {sem_label}.</p>
</header>
{body}
<footer>Gerado em {generated_at} · fonte: SigPesq / exportações canônicas</footer>
</body>
</html>"""


# ---------------------------------------------------------------------------
# Entry point
# ---------------------------------------------------------------------------

def main() -> None:
    parser = argparse.ArgumentParser()
    parser.add_argument("--semester", default="2024_1",
                        choices=list(SEMESTER_FILE_MAP.keys()))
    parser.add_argument("--all", dest="all_semesters", action="store_true",
                        help="Combine all available semesters into one report")
    parser.add_argument("--out", default=None,
                        help="Output HTML path (default: data/exports/formandos/)")
    args = parser.parse_args()

    if args.all_semesters:
        # Files are overlapping cumulative snapshots (same matrícula recurs across
        # files, all "Formado"). Dedup by matrícula — the admission join key — so
        # each student counts once. Earliest occurrence wins (keeps oldest
        # grad_semester); admission is already resolved via the global map.
        seen: dict[str, dict] = {}  # matrícula → formando
        semesters_used = []
        for sem in sorted(SEMESTER_FILE_MAP.keys()):
            path = DATA_FORMANDOS / SEMESTER_FILE_MAP[sem]
            if not path.exists():
                continue
            batch = load_formandos(sem)
            for f in batch:
                key = f["matricula"] or f["nome"].strip().lower()
                if key not in seen:
                    seen[key] = f
            semesters_used.append(sem)
            print(f"  {sem}: {len(batch)} formandos")
        formandos = list(seen.values())
        print(f"  Total único: {len(formandos)} alunos ({len(semesters_used)} semestres)")
        grad_semester = semesters_used[-1] if semesters_used else args.semester
    else:
        print(f"Loading formandos for {args.semester}...")
        formandos = load_formandos(args.semester)
        semesters_used = [args.semester]
        grad_semester = args.semester
        print(f"  {len(formandos)} formandos")

    print("Loading exports...")
    adv_projects = load_json("advisorships_canonical.json")
    rgs = load_json("research_groups_canonical.json")

    print("Loading Lattes CVs...")
    lattes = load_lattes()
    print(f"  {len(lattes['ic'])} IC records, {len(lattes['tcc'])} TCC records")

    bolsistas = load_bolsistas()
    print(f"Loading bolsistas... {len(bolsistas)} bolsistas únicos")

    print("Computing statistics...")
    stats = compute(formandos, adv_projects, rgs, lattes=lattes,
                    grad_semester=grad_semester, bolsistas=bolsistas)

    now = datetime.now().strftime("%Y-%m-%d %H:%M")
    _sem_key = "all" if args.all_semesters else args.semester
    html = render_html(stats, grad_semester, now, semesters=semesters_used)

    out_path = (
        Path(args.out)
        if args.out
        else OUT_DIR / f"formandos_pesquisa_{_sem_key}_generated.html"
    )
    out_path.parent.mkdir(parents=True, exist_ok=True)
    out_path.write_text(html, encoding="utf-8")
    print(f"Written: {out_path}")

    # JSON export — same stem, machine-readable stats payload
    json_path = out_path.with_suffix(".json")
    payload = {
        "semester": _sem_key,
        "semesters": semesters_used,
        "generated_at": now,
        "stats": stats,
    }
    json_path.write_text(
        json.dumps(payload, ensure_ascii=False, indent=2, default=str),
        encoding="utf-8",
    )
    print(f"Written: {json_path}")


if __name__ == "__main__":
    main()
